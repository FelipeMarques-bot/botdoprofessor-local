import csv
import io
import os
import re
import tempfile
import urllib.request
from dataclasses import dataclass
from datetime import date as _date_cls
from datetime import datetime as _datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

LogFn = Callable[[str], None]


@dataclass
class RegistroNota:
    escola: str
    turno: str
    turma: str
    trimestre: str
    aluno: str
    atividade: str
    nota: float
    data_realizacao: str = ""


def _normalize(s: str) -> str:
    import unicodedata
    text = (s or "").strip().lower()
    text = text.replace("º", "o").replace("°", "o").replace("ª", "a")
    text = text.replace("\u00a0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = "".join(ch for ch in text if unicodedata.category(ch) not in {"Cf", "Cc"})
    return re.sub(r"\s+", " ", text)


def _to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    if re.fullmatch(r"-?\d+(\.\d+)?", text):
        try:
            return float(text)
        except ValueError:
            return None
    return None


IGNORE_COLS = {
    "Nome", "Status", "Status Fluxo", "Media", "Media Final",
    "Observacoes", "Observacoes Pedagogicas",
}


def _is_grade_column(col_name: str) -> bool:
    clean = col_name.strip()
    if not clean or clean in IGNORE_COLS:
        return False
    lowered = re.sub(r"[^a-z0-9]+", " ", clean.lower()).strip()
    words = lowered.split()
    if not words:
        return False
    prefix_blacklist = ("status", "media", "obs", "coment", "nome", "chamada", "frequencia")
    for word in words:
        if word == "id":
            return False
        if any(word.startswith(bw) for bw in prefix_blacklist):
            return False
    return True


def _is_date_column(col_name: str) -> bool:
    """Colunas de data (ex: 'Data realização 1'). Nao sao colunas de nota."""
    return _normalize(col_name or "").startswith("data")


def _date_suffix(col_name: str) -> Optional[int]:
    """Extrai o numero final de uma coluna (ex: 'Data realização 2' -> 2)."""
    norm = _normalize(col_name or "")
    match = re.search(r"(\d+)\s*$", norm)
    return int(match.group(1)) if match else None


def _format_data_realizacao(value: object) -> str:
    """Normaliza um valor de data para o formato DD/MM/AAAA."""
    if value is None:
        return ""
    if isinstance(value, _datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, _date_cls):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if not text:
        return ""
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if match:
        return f"{int(match.group(1)):02d}/{int(match.group(2)):02d}/{match.group(3)}"
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        return f"{int(match.group(3)):02d}/{int(match.group(2)):02d}/{match.group(1)}"
    return text


def _pair_date_columns(activities, dates):
    """Associa colunas 'Data realização N' as colunas de atividade.

    activities: lista de (chave, nome_da_atividade) em ordem de coluna.
    dates: lista de (chave, nome_da_coluna, sufixo_numerico_ou_None).
    Retorna {chave_atividade: chave_data}.
    """
    mapping = {}
    used = set()
    assigned = set()

    for dkey, _dname, suffix in dates:
        if suffix is None:
            continue
        for akey, aname in activities:
            if akey in used:
                continue
            if re.search(rf"(^|\D){suffix}(\D|$)", _normalize(aname)):
                mapping[akey] = dkey
                used.add(akey)
                assigned.add(dkey)
                break

    for dkey, _dname, suffix in dates:
        if dkey in assigned or suffix is None:
            continue
        pos = suffix - 1
        if 0 <= pos < len(activities) and activities[pos][0] not in used:
            mapping[activities[pos][0]] = dkey
            used.add(activities[pos][0])
            assigned.add(dkey)

    for dkey, _dname, _suffix in dates:
        if dkey in assigned:
            continue
        for akey, _aname in activities:
            if akey not in used:
                mapping[akey] = dkey
                used.add(akey)
                assigned.add(dkey)
                break

    return mapping


def ler_notas_excel(caminho: str, logger: Optional[LogFn] = None) -> List[RegistroNota]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl nao instalado. Rode: pip install openpyxl")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    registros: List[RegistroNota] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h or "") for h in rows[0]]

        col_escola = _find_col_index(headers, ["escola", "escola"])
        col_turno = _find_col_index(headers, ["turno", "turno"])
        col_turma = _find_col_index(headers, ["turma", "turma", "classe", "sala"])
        col_trimestre = _find_col_index(headers, ["trimestre", "trimestre", "bimestre", "periodo"])
        col_nome = _find_col_index(headers, ["nome", "nome", "aluno", "estudante", "nome aluno"])

        meta_cols = {
            c for c in (col_escola, col_turno, col_turma, col_trimestre, col_nome) if c is not None
        }
        grade_cols = {}
        for i, h in enumerate(headers):
            if i in meta_cols:
                continue
            if _is_date_column(h):
                continue
            if _is_grade_column(h):
                grade_cols[i] = h.strip()

        if col_nome is None:
            continue

        escola = ""
        turno = ""
        turma = ""
        trimestre = ""
        info_rows = [r for r in rows[1:] if any(v is not None for v in r)]

        for row in info_rows:
            if col_escola is not None and not escola:
                escola = str(row[col_escola] or "").strip()
            if col_turno is not None and not turno:
                turno = str(row[col_turno] or "").strip()
            if col_turma is not None and not turma:
                turma = str(row[col_turma] or "").strip()
            if col_trimestre is not None and not trimestre:
                trimestre = str(row[col_trimestre] or "").strip()

        if not escola:
            escola = os.path.splitext(os.path.basename(caminho))[0]
        if not turno:
            turno = "Nao informado"
        if not turma:
            turma = sheet_name
        if not trimestre:
            trimestre = "Nao informado"

        date_map = _pair_date_columns(
            [(i, grade_cols[i]) for i in sorted(grade_cols)],
            [
                (i, headers[i], _date_suffix(headers[i]))
                for i in range(len(headers))
                if _is_date_column(headers[i])
            ],
        )

        for row in info_rows:
            nome = str(row[col_nome] or "").strip()
            if not nome:
                continue

            for col_idx, atividade in grade_cols.items():
                nota = _to_float(row[col_idx] if col_idx < len(row) else None)
                if nota is None:
                    continue
                data_col = date_map.get(col_idx)
                data_realizacao = ""
                if data_col is not None and data_col < len(row):
                    data_realizacao = _format_data_realizacao(row[data_col])
                registros.append(RegistroNota(
                    escola=escola,
                    turno=turno,
                    turma=turma,
                    trimestre=trimestre,
                    aluno=nome,
                    atividade=atividade,
                    nota=nota,
                    data_realizacao=data_realizacao,
                ))

        if logger:
            logger(f"Planilha '{sheet_name}': {len([r for r in info_rows if str(r[col_nome] or '').strip()])} alunos, {len(grade_cols)} atividades com notas")

    wb.close()
    return registros


def _ts_chave(value: object) -> float:
    if isinstance(value, _datetime):
        return value.timestamp()
    if isinstance(value, _date_cls):
        return _datetime(value.year, value.month, value.day).timestamp()
    return 0.0


def _parse_acertos(value: object) -> Tuple[Optional[float], Optional[float]]:
    """Interpreta a celula de acertos/pontuacao de um relatorio do Google Forms.

    Aceita '2/20' (texto), 19 / 19.0 (numero bruto) ou '19'.
    Retorna (acertos, total); total vem None quando nao ha o formato 'X/Y'.
    """
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None
    text = str(value).strip()
    if not text:
        return None, None
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", text)
    if match:
        return _to_float(match.group(1)), _to_float(match.group(2))
    return _to_float(text), None


def detectar_total_questoes_forms(caminho: str) -> Optional[int]:
    """Tenta descobrir o total de questoes de um relatorio do Google Forms.

    Prioriza os denominadores das celulas no formato 'X/Y'; se nao houver,
    usa a maior pontuacao encontrada. Retorna None se nada for detectado.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception:
        return None

    denominadores: List[float] = []
    brutas: List[float] = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h or "") for h in rows[0]]
            col_score = _find_col_index(headers, ["score", "pontuacao", "acertos"])
            if col_score is None:
                continue
            for row in rows[1:]:
                acertos, total = _parse_acertos(
                    row[col_score] if col_score < len(row) else None
                )
                if total:
                    denominadores.append(total)
                elif acertos is not None:
                    brutas.append(acertos)
    finally:
        wb.close()

    if denominadores:
        from collections import Counter as _Counter
        return int(_Counter(denominadores).most_common(1)[0][0])
    if brutas:
        return int(max(brutas))
    return None


def detectar_turmas_forms(caminho: str) -> List[str]:
    """Turmas distintas no relatorio do Google Forms (coluna 'Turma').

    Grafias diferentes da mesma turma ('6º1' x '6°1') sao unificadas,
    mantendo a primeira grafia vista. Retorna lista ordenada.
    """
    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception:
        return []

    turmas: Dict[str, str] = {}
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            headers = [str(h or "") for h in header]
            col_turma = _find_col_index(headers, ["turma"])
            if col_turma is None:
                continue
            for row in rows:
                val = row[col_turma] if col_turma < len(row) else None
                txt = str(val or "").strip()
                if not txt:
                    continue
                chave = _normalize(txt)
                if chave and chave not in turmas:
                    turmas[chave] = txt
    finally:
        wb.close()
    return [turmas[k] for k in sorted(turmas)]


def ler_notas_google_forms(
    caminho: str,
    valor_questao: float = 0.5,
    atividade: str = "",
    n_questoes: Optional[int] = None,
    logger: Optional[LogFn] = None,
) -> List[RegistroNota]:
    """Le relatorios de avaliacoes gerados pelo Google Forms.

    Estrutura esperada (export .xlsx do Forms): Timestamp, Score/Pontuacao,
    Nome, Escola, Turma e uma coluna por questao. A nota final de cada aluno e
    calculada como acertos * valor_questao (ex.: 19 acertos x 0,5 = 9,5).
    Aceita pontuacao bruta (19) ou fracao ('19/20'). Se o mesmo aluno responder
    mais de uma vez, mantem apenas a resposta mais recente.
    Nao altera a leitura das planilhas comuns nem do Notion.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl nao instalado. Rode: pip install openpyxl")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    registros: List[RegistroNota] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h or "") for h in rows[0]]
        col_nome = _find_col_index(headers, ["nome", "aluno", "estudante"])
        col_score = _find_col_index(headers, ["score", "pontuacao", "acertos"])
        col_escola = _find_col_index(headers, ["escola"])
        col_turno = _find_col_index(headers, ["turno"])
        col_turma = _find_col_index(headers, ["turma", "classe", "sala"])
        col_trimestre = _find_col_index(headers, ["trimestre", "bimestre", "periodo"])
        col_timestamp = next(
            (i for i, h in enumerate(headers) if _normalize(h) == "timestamp"), None
        )

        if col_nome is None or col_score is None:
            if logger:
                logger(f"Aba '{sheet_name}': sem colunas Nome/Score do Google Forms. Pulando.")
            continue

        brutos_por_linha: Dict[int, Tuple[Optional[float], Optional[float]]] = {}
        denominadores: List[float] = []
        brutas: List[float] = []
        for idx, row in enumerate(rows[1:], start=1):
            acertos, total = _parse_acertos(
                row[col_score] if col_score < len(row) else None
            )
            brutos_por_linha[idx] = (acertos, total)
            if total:
                denominadores.append(total)
            elif acertos is not None:
                brutas.append(acertos)

        if n_questoes:
            total_questoes = int(n_questoes)
        elif denominadores:
            from collections import Counter as _Counter
            total_questoes = int(_Counter(denominadores).most_common(1)[0][0])
        elif brutas:
            total_questoes = int(max(brutas))
        else:
            total_questoes = 0

        nome_atividade = (atividade or "").strip() or f"Avaliacao - {sheet_name}"

        # Dedup: mantem apenas a resposta mais recente de cada aluno.
        selecionadas: Dict[str, Tuple[int, float]] = {}
        ordem: List[str] = []
        duplicadas = 0
        for idx, row in enumerate(rows[1:], start=1):
            acertos, _total = brutos_por_linha.get(idx, (None, None))
            if acertos is None:
                continue
            nome = str(row[col_nome] or "").strip()
            if not nome:
                continue
            chave = _normalize(nome)
            ts = row[col_timestamp] if col_timestamp is not None and col_timestamp < len(row) else None
            ts_num = _ts_chave(ts)
            if chave in selecionadas:
                duplicadas += 1
                if ts_num >= selecionadas[chave][1]:
                    selecionadas[chave] = (idx, ts_num)
            else:
                ordem.append(chave)
                selecionadas[chave] = (idx, ts_num)

        registros_sheet: List[RegistroNota] = []
        notas_altas = 0
        for chave in ordem:
            idx, _tsn = selecionadas[chave]
            row = rows[idx]
            acertos, _total = brutos_por_linha.get(idx, (None, None))
            if acertos is None:
                continue
            nome = str(row[col_nome] or "").strip()
            nota = round(float(acertos) * float(valor_questao), 2)
            if nota > 10:
                notas_altas += 1
            registros_sheet.append(RegistroNota(
                escola=str(row[col_escola] or "").strip() if col_escola is not None else "",
                turno=str(row[col_turno] or "").strip() if col_turno is not None else "",
                turma=str(row[col_turma] or "").strip() if col_turma is not None else "",
                trimestre=str(row[col_trimestre] or "").strip() if col_trimestre is not None else "",
                aluno=nome,
                atividade=nome_atividade,
                nota=nota,
                data_realizacao="",
            ))

        registros.extend(registros_sheet)

        if logger and registros_sheet:
            msg = (
                f"Google Forms '{sheet_name}': {len(registros_sheet)} aluno(s), "
                f"{duplicadas} resposta(s) duplicada(s) ignorada(s), "
                f"{total_questoes} questoes x {valor_questao} pts"
            )
            if notas_altas:
                msg += (
                    f" | ATENCAO: {notas_altas} nota(s) acima de 10 - "
                    "confira o valor por questao"
                )
            logger(msg)

    wb.close()
    return registros


def _csv_skip_comments(f):
    for line in f:
        stripped = line.lstrip()
        if stripped.startswith("#"):
            continue
        yield line


def _csv_cell(row: Dict[str, Any], col: Optional[str]) -> str:
    if col is None:
        return ""
    val = row.get(col)
    if isinstance(val, list):
        return ",".join(str(v) for v in val)
    return val or ""


def _sniff_delimiter(content: str) -> str:
    first_line = content.splitlines()[0] if content else ""
    counts = {",": 0, ";": 0, "\t": 0}
    for ch in counts:
        counts[ch] = first_line.count(ch)
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def ler_notas_csv(caminho: str, logger: Optional[LogFn] = None) -> List[RegistroNota]:
    with open(caminho, encoding="utf-8-sig") as f:
        content = f.read()
    return _ler_csv_content(content, logger=logger)


def _ler_csv_content(content: str, logger: Optional[LogFn] = None) -> List[RegistroNota]:
    registros: List[RegistroNota] = []
    delimiter = _sniff_delimiter(content)
    reader = csv.DictReader(_csv_skip_comments(io.StringIO(content)), delimiter=delimiter)
    headers = reader.fieldnames or []

    nome_col = _find_header(headers, ["nome", "aluno", "estudante"])
    escola_col = _find_header(headers, ["escola"])
    turno_col = _find_header(headers, ["turno"])
    turma_col = _find_header(headers, ["turma", "classe", "sala"])
    trimestre_col = _find_header(headers, ["trimestre", "bimestre", "periodo"])

    if not nome_col:
        if logger:
            logger("CSV: coluna 'Nome' nao encontrada")
        return registros

    meta_cols = {c for c in (nome_col, escola_col, turno_col, turma_col, trimestre_col) if c}
    grade_cols = {
        h: h
        for h in headers
        if h not in meta_cols
        and _is_grade_column(h)
        and not _is_date_column(h)
        and _normalize(h) not in ("nome", "aluno")
    }

    date_map = _pair_date_columns(
        [(h, h) for h in grade_cols],
        [(h, h, _date_suffix(h)) for h in headers if _is_date_column(h)],
    )

    for row in reader:
        extra = row.pop(None, None)
        if isinstance(extra, list) and extra and headers:
            last = headers[-1]
            sufixo = ",".join(str(v) for v in extra)
            base = row.get(last)
            row[last] = (base + "," + sufixo) if base else sufixo
        nome = _csv_cell(row, nome_col).strip()
        if not nome:
            continue
        for col_h, atividade in grade_cols.items():
            raw = _csv_cell(row, col_h).strip()
            nota = _to_float(raw)
            if nota is None:
                continue
            data_h = date_map.get(col_h)
            data_realizacao = _format_data_realizacao(_csv_cell(row, data_h)) if data_h else ""
            registros.append(RegistroNota(
                escola=_csv_cell(row, escola_col).strip() or "Nao informado",
                turno=_csv_cell(row, turno_col).strip() or "Nao informado",
                turma=_csv_cell(row, turma_col).strip() or "Nao informado",
                trimestre=_csv_cell(row, trimestre_col).strip() or "Nao informado",
                aluno=nome,
                atividade=atividade,
                nota=nota,
                data_realizacao=data_realizacao,
            ))

    if logger:
        logger(f"CSV: {len(registros)} notas carregadas")
    return registros


def ler_notas_google_sheets(url_or_id: str, logger: Optional[LogFn] = None) -> List[RegistroNota]:
    sheet_id = _extract_sheet_id(url_or_id)
    if not sheet_id:
        raise ValueError(f"URL/ID de planilha invalida: {url_or_id}")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    try:
        # Exporta como XLSX: traz TODAS as abas (o export CSV so traz a primeira,
        # o que fazia planilhas com aba de instrucoes retornarem 0 notas).
        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        with urllib.request.urlopen(export_url, timeout=60) as resp:
            with open(tmp_path, "wb") as f:
                f.write(resp.read())

        registros = ler_notas_excel(tmp_path, logger=logger)

        # Fallback: exporta a primeira aba em CSV (planilhas que so tem uma aba
        # ou que nao exportam xlsx por algum motivo).
        if not registros:
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            with urllib.request.urlopen(export_url, timeout=60) as resp:
                content = resp.read().decode("utf-8-sig")
            registros = _ler_csv_content(content, logger=logger)

        os.unlink(tmp_path)
        if logger:
            logger(f"Google Sheets: {len(registros)} notas carregadas")
        return registros
    except RuntimeError:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise
    except Exception as exc:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise RuntimeError(f"Falha ao baixar planilha do Google Sheets: {exc}")


def _find_header(headers: List[str], candidates: List[str]) -> Optional[str]:
    h_norm = {h: _normalize(h) for h in headers}
    for cand in candidates:
        cand_norm = _normalize(cand)
        if not cand_norm:
            continue
        for h, hn in h_norm.items():
            if hn and cand_norm == hn:
                return h
    for cand in candidates:
        cand_norm = _normalize(cand)
        if not cand_norm:
            continue
        for h, hn in h_norm.items():
            if hn and (cand_norm in hn or hn in cand_norm):
                return h
    return None


def _find_col_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    h_norm = {i: _normalize(h) for i, h in enumerate(headers)}
    for cand in candidates:
        cand_norm = _normalize(cand)
        if not cand_norm:
            continue
        for i, hn in h_norm.items():
            if hn and cand_norm == hn:
                return i
    for cand in candidates:
        cand_norm = _normalize(cand)
        if not cand_norm:
            continue
        for i, hn in h_norm.items():
            if hn and (cand_norm in hn or hn in cand_norm):
                return i
    return None


def _detect_spreadsheet_type(path: str) -> str:
    """Detecta o tipo de planilha pelo conteudo do arquivo, nao pela extensao."""
    try:
        with open(path, "rb") as f:
            head = f.read(8)
    except OSError:
        return "excel"
    if head.startswith(b"PK\x03\x04"):
        return "excel"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "excel"
    return "csv"


def ler_notas_google_drive(url_or_id: str, logger: Optional[LogFn] = None) -> List[RegistroNota]:
    try:
        import gdown
    except ImportError:
        raise RuntimeError(
            "gdown nao instalado. Rode: pip install gdown"
        )

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    try:
        if re.match(r"^https?://", url_or_id):
            output = gdown.download(url_or_id, tmp_path, quiet=True, fuzzy=True)
        else:
            output = gdown.download(id=url_or_id, output=tmp_path, quiet=True, fuzzy=True)

        if output is None or not os.path.exists(tmp_path):
            raise RuntimeError(
                "Nao foi possivel baixar o arquivo do Google Drive. "
                "Verifique se o link esta correto e se o arquivo esta compartilhado "
                "como 'Qualquer um com o link pode ver'."
            )

        size = os.path.getsize(tmp_path)
        if size == 0:
            raise RuntimeError("O arquivo baixado do Google Drive esta vazio.")

        if logger:
            logger(f"Google Drive: download ok ({_fmt_size(size)})")

        tipo = _detect_spreadsheet_type(tmp_path)
        if tipo == "csv":
            registros = ler_notas_csv(tmp_path, logger=logger)
        else:
            registros = ler_notas_excel(tmp_path, logger=logger)

        os.unlink(tmp_path)
        if logger:
            logger(f"Google Drive: {len(registros)} notas carregadas")
        return registros
    except RuntimeError:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise
    except Exception as exc:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise RuntimeError(
            f"Falha ao baixar arquivo do Google Drive: {exc}. "
            "Confira se o link esta certo e o arquivo esta publico "
            "('Qualquer um com o link pode ver')."
        )


def _fmt_size(n: int) -> str:
    if n >= 1024 * 1024:
        return f"{n / (1024 * 1024):.1f} MB"
    if n >= 1024:
        return f"{n / 1024:.1f} KB"
    return f"{n} B"


def _normalize_str(s: str) -> str:
    import unicodedata
    return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch)).strip()


def gerar_template_notas_xlsx() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exemplo - 6º Ano"
    headers = [
        "Escola", "Turno", "Turma", "Trimestre",
        "Nome do Aluno", "Atividade 1", "Atividade 2", "Atividade 3",
        "Data realização 1", "Data realização 2", "Data realização 3",
        "Observações 1", "Observações 2", "Observações 3",
    ]
    ws.append(headers)
    ws.append([
        "Juvenal", "Matutino", "6º Ano", "1º Trimestre",
        "Exemplo Aluno", "8,5", "7,0", "9,0",
        "01/03/2026", "15/04/2026", "20/05/2026",
        "", "", "",
    ])
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 18)

    _add_notas_instrucoes(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _add_notas_instrucoes(wb):
    ws2 = wb.create_sheet("Instrucoes", 0)
    instrucoes = [
        ["INSTRUCOES - PLANILHA DE NOTAS", "", ""],
        ["", "", ""],
        ["O que e cada coluna:", "", ""],
        ["", "", ""],
        ["Coluna", "O que preencher", "Exemplo"],
        ["Escola", "Nome da escola (como aparece no portal)", "Juvenal"],
        ["Turno", "Matutino, Vespertino ou Noturno", "Matutino"],
        ["Turma", "Turma completa (ex: 6o Ano, 7o Ano)", "6o Ano"],
        ["Trimestre", "1o Trimestre, 2o Trimestre ou 3o Trimestre", "1o Trimestre"],
        ["Nome do Aluno", "Nome completo do aluno", "Joao Silva"],
        ["Atividade 1", "Nota da primeira atividade (use virgula para decimal)", "8,5"],
        ["Atividade 2", "Nota da segunda atividade", "7,0"],
        ["Atividade 3", "Nota da terceira atividade", "9,0"],
        ["Data realizacao 1", "Data que a atividade 1 foi aplicada (dd/mm/aaaa)", "01/03/2026"],
        ["Data realizacao 2", "Data que a atividade 2 foi aplicada (dd/mm/aaaa)", "15/04/2026"],
        ["Data realizacao 3", "Data que a atividade 3 foi aplicada (dd/mm/aaaa)", "20/05/2026"],
        ["Observacoes 1", "Observacao sobre a atividade 1 (opcional)", ""],
        ["Observacoes 2", "Observacao sobre a atividade 2 (opcional)", ""],
        ["Observacoes 3", "Observacao sobre a atividade 3 (opcional)", ""],
        ["", "", ""],
        ["DICAS IMPORTANTES:", "", ""],
        ["1. Os nomes das atividades (Atividade 1, 2, 3) devem ser IGUAIS aos nomes", "que aparecem no portal SGE para aquela turma/trimestre.", ""],
        ["2. Preencha uma linha por aluno. Repita os dados de escola/turma/trimestre", "para cada aluno.", ""],
        ["3. Use virgula (,) como separador decimal nas notas. Ex: 8,5", "", ""],
        ["4. Datas sempre no formato: dia/mes/ano (ex: 01/03/2026)", "", ""],
        ["5. Deixe em branco o que nao for preencher.", "", ""],
    ]
    for row in instrucoes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 24


def gerar_template_notas_csv() -> str:
    buf = io.StringIO()
    _write_csv_comments(buf, [
        "INSTRUCOES - Planilha de Notas",
        "Escola: nome da escola. Ex: Juvenal",
        "Turno: Matutino, Vespertino ou Noturno",
        "Turma: 6o Ano, 7o Ano, etc.",
        "Trimestre: 1o Trimestre, 2o Trimestre ou 3o Trimestre",
        "Nome do Aluno: nome completo do aluno",
        "Atividade 1/2/3: nota de cada atividade (use virgula p/ decimal). Ex: 8,5",
        "Data realizacao 1/2/3: data da atividade no formato dd/mm/aaaa",
        "Observacoes 1/2/3: opcional",
        "",
        "IMPORTANTE: os nomes das atividades devem ser IGUAIS aos nomes no portal SGE",
        "",
    ])
    writer = csv.writer(buf)
    writer.writerow([
        "Escola", "Turno", "Turma", "Trimestre",
        "Nome do Aluno", "Atividade 1", "Atividade 2", "Atividade 3",
        "Data realização 1", "Data realização 2", "Data realização 3",
        "Observações 1", "Observações 2", "Observações 3",
    ])
    writer.writerow([
        "Juvenal", "Matutino", "6º Ano", "1º Trimestre",
        "Exemplo Aluno", "8,5", "7,0", "9,0",
        "01/03/2026", "15/04/2026", "20/05/2026",
        "", "", "",
    ])
    return buf.getvalue()


def gerar_template_sequencias_xlsx() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sequências"
    headers = [
        "Name", "Escola", "Ano", "Periodo inicio", "Periodo fim",
        "N aulas", "Titulo Documento", "Link do Arquivo",
        "Ativo", "Observações",
    ]
    ws.append(headers)
    ws.append([
        "SD - 6º Ano - Matemática", "Juvenal", "6º Ano",
        "01/03/2026", "31/03/2026", 4,
        "Sequência Didática - Matemática - 6º Ano",
        "https://drive.google.com/your-file-link-here",
        "Sim", "",
    ])
    ws.append([
        "SD - 7º Ano - Português", "Arapongas", "7º Ano",
        "01/04/2026", "30/04/2026", 4,
        "Sequência Didática - Português - 7º Ano",
        "https://drive.google.com/your-file-link-here",
        "Sim", "",
    ])
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 22)

    _add_seq_instrucoes(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _add_seq_instrucoes(wb):
    ws2 = wb.create_sheet("Instrucoes", 0)
    instrucoes = [
        ["INSTRUCOES - PLANILHA DE SEQUENCIAS DIDATICAS", "", ""],
        ["", "", ""],
        ["O que e cada coluna:", "", ""],
        ["", "", ""],
        ["Coluna", "O que preencher", "Exemplo"],
        ["Name", "Nome da sequencia (para identificacao)", "SD - 6o Ano - Matematica"],
        ["Escola", "Nome da escola (como aparece no portal)", "Juvenal"],
        ["Ano", "Ano escolar (6o Ano, 7o Ano, 8o Ano, 9o Ano)", "6o Ano"],
        ["Periodo inicio", "Data de inicio da sequencia (dd/mm/aaaa)", "01/03/2026"],
        ["Periodo fim", "Data de termino da sequencia (dd/mm/aaaa)", "31/03/2026"],
        ["N aulas", "Numero de aulas da sequencia", "4"],
        ["Titulo Documento", "Titulo do documento que aparecera no portal", "Sequencia Didatica - Matematica - 6o Ano"],
        ["Link do Arquivo", "Link do Google Drive para o PDF do plano de aula", "https://drive.google.com/..."],
        ["Ativo", "'Sim' para processar, 'Nao' para pular esta linha", "Sim"],
        ["Observacoes", "Observacao opcional (nao vai para o portal)", ""],
        ["", "", ""],
        ["DICAS IMPORTANTES:", "", ""],
        ["1. Coloque 'Nao' na coluna Ativo para pular linhas sem deleta-las.", "", ""],
        ["2. Os links do Google Drive devem ser compartilhados como 'Qualquer um com o link pode ver'.", "", ""],
        ["3. Datas sempre no formato: dia/mes/ano (ex: 01/03/2026)", "", ""],
        ["4. Deixe em branco o que nao for preencher.", "", ""],
    ]
    for row in instrucoes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 24


def _write_csv_comments(buf, lines):
    for line in lines:
        buf.write(f"#{line}\n")


def gerar_template_sequencias_csv() -> str:
    buf = io.StringIO()
    _write_csv_comments(buf, [
        "INSTRUCOES - Planilha de Sequencias Didaticas",
        "Name: nome da sequencia para identificacao. Ex: SD - 6o Ano - Matematica",
        "Escola: nome da escola. Ex: Juvenal",
        "Ano: 6o Ano, 7o Ano, 8o Ano ou 9o Ano",
        "Periodo inicio: data de inicio (dd/mm/aaaa). Ex: 01/03/2026",
        "Periodo fim: data de termino (dd/mm/aaaa). Ex: 31/03/2026",
        "N aulas: numero de aulas. Ex: 4",
        "Titulo Documento: titulo que aparecera no portal",
        "Link do Arquivo: link do Google Drive para o PDF",
        "Ativo: Sim para processar, Nao para pular",
        "Observacoes: opcional",
        "",
        "IMPORTANTE: links do Drive devem estar como 'Qualquer um com o link pode ver'",
        "",
    ])
    writer = csv.writer(buf)
    writer.writerow([
        "Name", "Escola", "Ano", "Periodo inicio", "Periodo fim",
        "N aulas", "Titulo Documento", "Link do Arquivo",
        "Ativo", "Observações",
    ])
    writer.writerow([
        "SD - 6º Ano - Matemática", "Juvenal", "6º Ano",
        "01/03/2026", "31/03/2026", 4,
        "Sequência Didática - Matemática - 6º Ano",
        "https://drive.google.com/your-file-link-here",
        "Sim", "",
    ])
    return buf.getvalue()


def ler_sequencias_excel(caminho: str, logger: Optional[LogFn] = None) -> list:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl nao instalado")
    from datetime import datetime as _dt

    wb = openpyxl.load_workbook(caminho, data_only=True)
    registros = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        h_norm = {_normalize_str(h): i for i, h in enumerate(headers)}

        def _col(names):
            for n in names:
                nn = _normalize_str(n)
                if nn in h_norm:
                    return h_norm[nn]
            return None

        col_name = _col(["Name", "Nome", "Titulo"])
        col_escola = _col(["Escola"])
        col_ano = _col(["Ano"])
        col_periodo_inicio = _col(["Periodo inicio", "Periodo início", "Data inicio", "Data início"])
        col_periodo_fim = _col(["Periodo fim", "Periodo fim", "Data fim"])
        col_n_aulas = _col(["N aulas", "Nº aulas", "Numero de aulas"])
        col_titulo_doc = _col(["Titulo Documento", "Título Documento", "Titulo do Documento"])
        col_link = _col(["Link do Arquivo", "Link do arquivo", "Arquivo PDF", "URL"])
        col_ativo = _col(["Ativo"])

        if col_name is None:
            if logger:
                logger(f"Aba '{sheet_name}': coluna 'Name' nao encontrada. Pulando.")
            continue

        for row_idx, row in enumerate(rows[1:], start=2):
            name = str(row[col_name] or "").strip()
            if not name:
                continue
            escola = str(row[col_escola] or "").strip() if col_escola is not None else ""
            ano = str(row[col_ano] or "").strip() if col_ano is not None else ""
            periodo_inicio = _fmt_date(str(row[col_periodo_inicio] or "")) if col_periodo_inicio is not None else ""
            periodo_fim = _fmt_date(str(row[col_periodo_fim] or "")) if col_periodo_fim is not None else ""
            n_aulas_raw = row[col_n_aulas] if col_n_aulas is not None else None
            n_aulas = 4
            if n_aulas_raw is not None:
                try:
                    n_aulas = int(float(str(n_aulas_raw).replace(",", ".")))
                except (ValueError, TypeError):
                    pass
            titulo_doc = str(row[col_titulo_doc] or "").strip() if col_titulo_doc is not None else name
            link = str(row[col_link] or "").strip() if col_link is not None else ""
            ativo_raw = str(row[col_ativo] or "").strip().lower() if col_ativo is not None else "sim"
            ativo = ativo_raw in ("sim", "s", "yes", "1", "true", "verdadeiro")

            if not ativo:
                if logger:
                    logger(f"Linha {row_idx}: registro '{name}' ignorado (Ativo = '{ativo_raw}').")
                continue

            registros.append({
                "page_id": "",
                "ano": ano,
                "escola": escola,
                "turno": "",
                "turma": "",
                "titulo_documento": titulo_doc,
                "arquivo_nome": f"{name}.pdf",
                "arquivo_url": link,
                "link_arquivo": link,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
                "n_aulas": n_aulas,
                "status": "",
            })

        if logger:
            logger(f"Aba '{sheet_name}': {len(registros)} sequencias carregadas.")

    wb.close()
    return registros


def ler_sequencias_csv(caminho: str, logger: Optional[LogFn] = None) -> list:
    registros = []
    from datetime import datetime as _dt

    with open(caminho, encoding="utf-8-sig") as f:
        reader = csv.DictReader(_csv_skip_comments(f))
        headers = reader.fieldnames or []
        h_norm = {_normalize_str(h): h for h in headers}

        def _col(names):
            for n in names:
                nn = _normalize_str(n)
                if nn in h_norm:
                    return h_norm[nn]
            return None

        col_name = _col(["Name", "Nome", "Titulo"])
        col_escola = _col(["Escola"])
        col_ano = _col(["Ano"])
        col_periodo_inicio = _col(["Periodo inicio", "Periodo início", "Data inicio", "Data início"])
        col_periodo_fim = _col(["Periodo fim", "Periodo fim", "Data fim"])
        col_n_aulas = _col(["N aulas", "Nº aulas", "Numero de aulas"])
        col_titulo_doc = _col(["Titulo Documento", "Título Documento", "Titulo do Documento"])
        col_link = _col(["Link do Arquivo", "Link do arquivo", "Arquivo PDF", "URL"])
        col_ativo = _col(["Ativo"])

        if not col_name:
            if logger:
                logger("CSV: coluna 'Name' nao encontrada")
            return registros

        for row in reader:
            name = (row.get(col_name) or "").strip()
            if not name:
                continue

            escola = (row.get(col_escola) or "").strip() if col_escola else ""
            ano = (row.get(col_ano) or "").strip() if col_ano else ""
            periodo_inicio = _fmt_date(row.get(col_periodo_inicio, "")) if col_periodo_inicio else ""
            periodo_fim = _fmt_date(row.get(col_periodo_fim, "")) if col_periodo_fim else ""
            n_aulas = 4
            if col_n_aulas:
                try:
                    n_aulas = int(float(str(row.get(col_n_aulas, "4")).replace(",", ".")))
                except (ValueError, TypeError):
                    pass
            titulo_doc = (row.get(col_titulo_doc) or name).strip() if col_titulo_doc else name
            link = (row.get(col_link) or "").strip() if col_link else ""
            ativo_raw = (row.get(col_ativo) or "sim").strip().lower() if col_ativo else "sim"
            ativo = ativo_raw in ("sim", "s", "yes", "1", "true", "verdadeiro")
            if not ativo:
                continue

            registros.append({
                "page_id": "",
                "ano": ano,
                "escola": escola,
                "turno": "",
                "turma": "",
                "titulo_documento": titulo_doc,
                "arquivo_nome": f"{name}.pdf",
                "arquivo_url": link,
                "link_arquivo": link,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
                "n_aulas": n_aulas,
                "status": "",
            })

    if logger:
        logger(f"CSV: {len(registros)} sequencias carregadas")
    return registros


def _fmt_date(value: str) -> str:
    from datetime import datetime as _dt
    raw = value.strip()
    if not raw:
        return ""
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", raw):
        return raw
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return _dt.strptime(raw, "%Y-%m-%d").strftime("%d/%m/%Y")
    try:
        import openpyxl
        from datetime import datetime as _py_dt
        if isinstance(value, _py_dt):
            return value.strftime("%d/%m/%Y")
    except ImportError:
        pass
    return raw


def _extract_sheet_id(url_or_id: str) -> Optional[str]:
    url_or_id = url_or_id.strip()
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_or_id)
    if match:
        return match.group(1)
    if re.match(r"^[a-zA-Z0-9_-]{20,}$", url_or_id):
        return url_or_id
    return None


def registros_para_linhas(registros: List[RegistroNota]) -> List[Dict[str, Any]]:
    """Converte RegistroNota em linhas editaveis para o data_editor do painel."""
    return [
        {
            "escola": r.escola,
            "turno": r.turno,
            "turma": r.turma,
            "trimestre": r.trimestre,
            "aluno": r.aluno,
            "atividade": r.atividade,
            "nota": r.nota,
            "data_realizacao": r.data_realizacao,
            "status": "",
        }
        for r in registros
    ]


def linhas_para_registros(
    linhas: List[Dict[str, Any]],
    defaults: Optional[Dict[str, str]] = None,
    logger: Optional[LogFn] = None,
) -> List[RegistroNota]:
    """Converte linhas editadas no painel em RegistroNota validos.

    Linhas sem aluno/atividade ou com nota invalida (None ou fora de 0-10)
    sao ignoradas. Campos vazios usam os defaults (filtros do painel) e,
    se ainda assim vazios, ficam como 'Nao informado'.
    """
    defaults = defaults or {}
    registros: List[RegistroNota] = []

    def _fill(ln: Dict[str, Any], campo: str) -> str:
        val = str(ln.get(campo) or "").strip()
        return val or defaults.get(campo) or "Nao informado"

    for ln in linhas:
        aluno = str(ln.get("aluno") or "").strip()
        atividade = str(ln.get("atividade") or "").strip()
        if not aluno or not atividade:
            continue
        nota = _to_float(ln.get("nota"))
        if nota is None or nota < 0 or nota > 10:
            if logger:
                logger(
                    f"Linha ignorada (nota invalida): aluno='{aluno}' atividade='{atividade}'"
                )
            continue
        registros.append(RegistroNota(
            escola=_fill(ln, "escola"),
            turno=_fill(ln, "turno"),
            turma=_fill(ln, "turma"),
            trimestre=_fill(ln, "trimestre"),
            aluno=aluno,
            atividade=atividade,
            nota=nota,
            data_realizacao=str(ln.get("data_realizacao") or "").strip(),
        ))
    return registros


def filtrar_linhas_por_filtros(
    linhas: List[Dict[str, Any]],
    filtros: Dict[str, str],
    logger: Optional[LogFn] = None,
) -> Tuple[List[Dict[str, Any]], int]:
    """Mantem apenas as linhas que casam com os filtros ATIVOS do painel.

    Os filtros do painel tambem preenchem campos vazios em
    linhas_para_registros; aqui eles passam a RESTRINGIR: se o professor
    selecionar Escola/Turno/Turma/Trimestre, linhas de outras escolas/turmas
    sao ignoradas no lancamento. Comparacao tolerante (minusculas, sem
    acentos, '6º1' = '6°1'). Campo vazio na linha herda o proprio filtro e
    portanto sempre casa. Retorna (linhas_mantidas, qtd_descartada).
    """
    ativos = {c: _normalize(v) for c, v in (filtros or {}).items() if _normalize(v)}
    if not ativos:
        return list(linhas), 0
    mantidas: List[Dict[str, Any]] = []
    descartadas = 0
    for ln in linhas:
        ok = True
        for campo, alvo in ativos.items():
            valor_linha = _normalize(str(ln.get(campo) or "")) or alvo
            if valor_linha != alvo:
                ok = False
                break
        if ok:
            mantidas.append(ln)
        else:
            descartadas += 1
    if descartadas and logger:
        logger(f"[FILTRO] {descartadas} linha(s) fora dos filtros selecionados foram ignoradas.")
    return mantidas, descartadas


def valores_distintos(linhas: List[Dict[str, Any]], campo: str) -> List[str]:
    """Valores distintos de uma coluna das linhas do painel, ordenados.

    Grafias equivalentes ('6º1' x '6°1') sao unificadas mantendo a primeira
    vista; usados para oferecer opcoes exatas nos filtros do painel.
    """
    vistos: Dict[str, str] = {}
    for ln in linhas:
        txt = str(ln.get(campo) or "").strip()
        if not txt:
            continue
        chave = _normalize(txt)
        if chave and chave not in vistos:
            vistos[chave] = txt
    return [vistos[k] for k in sorted(vistos)]


def carregar_notas(
    fonte: str,
    caminho_ou_url: str,
    logger: Optional[LogFn] = None,
) -> List[RegistroNota]:
    if fonte == "notion":
        msg = "Use lancar_notas_sge.carregar_notas_notion() para fonte Notion"
        raise ValueError(msg)

    if fonte == "excel":
        if not os.path.exists(caminho_ou_url):
            raise FileNotFoundError(f"Arquivo nao encontrado: {caminho_ou_url}")
        return ler_notas_excel(caminho_ou_url, logger=logger)

    if fonte == "csv":
        if not os.path.exists(caminho_ou_url):
            raise FileNotFoundError(f"Arquivo nao encontrado: {caminho_ou_url}")
        return ler_notas_csv(caminho_ou_url, logger=logger)

    if fonte == "google_sheets":
        return ler_notas_google_sheets(caminho_ou_url, logger=logger)

    if fonte == "google_drive":
        return ler_notas_google_drive(caminho_ou_url, logger=logger)

    raise ValueError(f"Fonte desconhecida: {fonte}. Use: excel, csv, google_sheets, google_drive, notion")
