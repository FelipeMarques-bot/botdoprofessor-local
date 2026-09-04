"""
Interface grafica amigavel para o Bot do Professor - Lancamento de Notas.
Usa Streamlit para abrir no navegador.
Nao requer conhecimentos tecnicos.

Uso:
    streamlit run painel.py
"""

import json
import os
import re
import subprocess
import sys
import tempfile
import time
import urllib.request
import urllib.error
import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import streamlit as st

from autofix import attempt_autofix, apply_fix

LICENSE_SERVER_URL = "https://botdoprofessor.onrender.com"
from leitor_planilhas import (
    gerar_template_notas_xlsx, gerar_template_notas_csv,
    gerar_template_sequencias_xlsx, gerar_template_sequencias_csv,
    ler_sequencias_excel, ler_sequencias_csv,
)

# Caminho sintetico da fonte "planilha no painel": usada pelo StatusStore
# para persistir o status (Lancada/Falha) das linhas editadas entre execucoes.
_PAINEL_SOURCE_PATH = os.path.join(
    tempfile.gettempdir(), "sge_bot_uploads", "_planilha_painel.xlsx"
)

# === CONFIGURACAO DA PAGINA ===
st.set_page_config(
    page_title="Bot do Professor - Lancamento de Notas",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

# === ESTILO CSS ===
_BASE_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    * { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; }

    .main-header {
        font-size: 2rem; font-weight: 700; margin-bottom: 0.25rem;
        background: linear-gradient(135deg, #2563eb, #7c3aed);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .sub-header { font-size: 1rem; color: #666; margin-bottom: 1.5rem; }

    .card {
        background: #f8f9fa; border-radius: 12px; padding: 1.5rem;
        margin-bottom: 1rem; border: 1px solid #e9ecef;
        transition: box-shadow 0.2s ease, transform 0.2s ease;
    }
    .card:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.08); transform: translateY(-1px); }

    .stButton button {
        width: 100%; border-radius: 8px; padding: 0.5rem 1rem; font-weight: 600;
        transition: all 0.2s ease; position: relative; overflow: hidden;
    }
    .stButton button:hover {
        transform: translateY(-1px); box-shadow: 0 4px 12px rgba(37,99,235,0.3);
    }
    .stButton button:active { transform: translateY(0); }
    .stButton button[kind="primary"] {
        background: linear-gradient(135deg, #2563eb, #7c3aed);
        border: none; color: white;
    }
    .stButton button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1d4ed8, #6d28d9);
        box-shadow: 0 4px 16px rgba(37,99,235,0.4);
    }

    .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] {
        border-radius: 8px; transition: border-color 0.2s ease, box-shadow 0.2s ease;
    }
    .stTextInput input:focus, .stTextArea textarea:focus {
        border-color: #2563eb; box-shadow: 0 0 0 3px rgba(37,99,235,0.15);
    }

    .st-bb, .st-bc, .st-bd, .st-be, .st-bf, .st-bg, .st-bh { border-radius: 8px; }

    div[data-testid="stMetric"] {
        background: #f8f9fa; border-radius: 10px; padding: 1rem;
        border: 1px solid #e9ecef; transition: all 0.2s ease;
    }
    div[data-testid="stMetric"]:hover {
        box-shadow: 0 4px 12px rgba(0,0,0,0.06); transform: translateY(-1px);
    }

    section[data-testid="stSidebar"] .stMarkdown h3 { font-size: 0.9rem; text-transform: uppercase; letter-spacing: 0.05em; color: #888; }

    .stAlert { border-radius: 10px; border: none; }
    .st-bw { border-radius: 8px; }

    @media (max-width: 768px) {
        .main-header { font-size: 1.5rem; }
        .row-widget.stColumns { flex-direction: column; }
    }

    @keyframes pulse-glow {
        0%, 100% { box-shadow: 0 0 0 0 rgba(37,99,235,0.4); }
        50% { box-shadow: 0 0 0 8px rgba(37,99,235,0); }
    }
    .stButton button[kind="primary"]:not(:disabled) {
        animation: pulse-glow 2s infinite;
    }

    .link-entry {
        background: #f0f4ff; border-radius: 8px; padding: 0.75rem;
        margin-bottom: 0.5rem; border-left: 3px solid #2563eb;
    }
    .link-entry-remove { color: #ef4444; cursor: pointer; font-size: 0.8rem; }
</style>
"""

_LIGHT_CSS = """
<style>
    .stApp { background: #fafbfc; color: #1a1a2e; }
    div[data-testid="stMetric"] { background: #ffffff; }
    .link-entry { background: #f0f4ff; }
</style>
"""

_DARK_THEME_CSS = """
<style>
    .stApp { background: #0e1117; color: #e0e0e0; }
    div[data-testid="stMetric"] { background: #1e2130; border-color: #2d3040; }
    div[data-testid="stMetric"]:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.3); }
    .main-header { background: linear-gradient(135deg, #60a5fa, #a78bfa); -webkit-background-clip: text; background-clip: text; }
    .sub-header { color: #888; }
    .card { background: #1e2130; border-color: #2d3040; }
    .card:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.3); }
    .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] { background-color: #262a3a; color: #e0e0e0; border-color: #3a3d50; }
    .st-bd, .st-cf, .st-cg, .st-ch, .st-ci, .st-cj { background-color: #1e2130; color: #e0e0e0; }
    .st-dg, .st-dh, .st-di, .st-dj, .st-dk, .st-dl { background-color: #262a3a; color: #e0e0e0; }
    .stMarkdown, .stText, p, span, label { color: #c0c0d0; }
    h1, h2, h3, h4, h5, h6 { color: #e8e8f0; }
    .st-bb, .st-bc, .st-bd, .st-be, .st-bf, .st-bg, .st-bh { background-color: #262a3a; }
    .st-cx, .st-cy, .st-cz, .st-d0, .st-d1, .st-d2 { border-color: #3a3d50; }
    .stAlert { background-color: #1e2130; color: #e0e0e0; }
    section[data-testid="stSidebar"] { background-color: #12151f; }
    section[data-testid="stSidebar"] .stMarkdown, section[data-testid="stSidebar"] p { color: #b0b0c0; }
    .link-entry { background: #1a1f35; border-left-color: #3b82f6; }
</style>
"""

st.markdown(_BASE_CSS, unsafe_allow_html=True)


# === SESSAO DE ESTADO ===
if "logs" not in st.session_state:
    st.session_state.logs = []
if "executando" not in st.session_state:
    st.session_state.executando = False
if "resultado" not in st.session_state:
    st.session_state.resultado = None
if "config_path" not in st.session_state:
    st.session_state.config_path = ""
if "notion_dbs" not in st.session_state:
    st.session_state.notion_dbs = []


def log(msg: str):
    try:
        timestamp = datetime.now().strftime("%H:%M:%S")
        st.session_state.logs.append(f"[{timestamp}] {msg}")
    except Exception:
        pass
    try:
        path = st.session_state.get("log_file", "")
        if path:
            with open(path, "a", encoding="utf-8") as f:
                f.write(f"[{timestamp}] {msg}\n")
    except Exception:
        pass


def _start_log_file() -> str:
    """Cria (ou retorna) o arquivo de log persistente da execucao atual."""
    log_dir = os.path.join(os.getcwd(), "artifacts", "logs")
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = os.path.join(log_dir, f"execucao_{timestamp}.log")
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"=== Execucao {datetime.now().isoformat()} ===\n")
    except Exception:
        pass
    return path


def _stats_globais_path() -> str:
    stats_dir = os.path.join(os.path.expanduser("~"), ".bot_local")
    os.makedirs(stats_dir, exist_ok=True)
    return os.path.join(stats_dir, "stats_globais.json")


def _registrar_stats_resultado(resultado: dict) -> None:
    """Acumula sucessos/falhas globais para calcular a taxa de acerto do bot."""
    if not isinstance(resultado, dict):
        return
    sucesso = int(
        resultado.get("notas_preenchidas")
        or resultado.get("preenchidas")
        or resultado.get("planejamentos")
        or 0
    )
    falhas = int(resultado.get("falhas", 0) or 0)
    if sucesso <= 0 and falhas <= 0:
        return
    path = _stats_globais_path()
    try:
        try:
            with open(path, encoding="utf-8") as f:
                dados = json.load(f)
        except Exception:  # noqa: BLE001
            dados = {"sucesso": 0, "falha": 0}
        dados["sucesso"] = int(dados.get("sucesso", 0)) + sucesso
        dados["falha"] = int(dados.get("falha", 0)) + falhas
        with open(path, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False)
    except Exception:  # noqa: BLE001
        pass


def _taxa_acerto_global() -> tuple:
    """Retorna (percentual, total_lancamentos); (None, 0) se sem historico."""
    try:
        with open(_stats_globais_path(), encoding="utf-8") as f:
            dados = json.load(f)
        sucesso = int(dados.get("sucesso", 0))
        falha = int(dados.get("falha", 0))
        total = sucesso + falha
        if total <= 0:
            return None, 0
        return round(100.0 * sucesso / total), total
    except Exception:  # noqa: BLE001
        return None, 0


from contextlib import contextmanager


@st.cache_resource(show_spinner=False)
def _nav_persistente_estado():
    """Navegadores reutilizados entre execucoes, um por portal (chave = URL)."""
    return {"portais": {}, "atexit_ok": False}


def _fechar_navegador_persistente(url_portal: str = "") -> None:
    """Fecha sessoes persistidas. Sem url informada, fecha TODOS os portais."""
    est = _nav_persistente_estado()
    chaves = [url_portal] if url_portal else list(est["portais"].keys())
    for chave in chaves:
        sess = est["portais"].pop(chave, None)
        if not sess:
            continue
        for obj_key in ("page", "ctx", "browser"):
            obj = sess.get(obj_key)
            if obj is not None:
                try:
                    obj.close()
                except Exception:
                    pass
        pw = sess.get("pw")
        if pw is not None:
            try:
                pw.stop()
            except Exception:
                pass


def _obter_page_persistente(headless_flag, url_portal: str, cpf: str, senha: str):
    """Page viva do portal indicado; troca de URL/credenciais descarta a sessao."""
    import atexit

    est = _nav_persistente_estado()
    if not est.get("atexit_ok"):
        atexit.register(_fechar_navegador_persistente)
        est["atexit_ok"] = True

    chave = (url_portal or "").strip().rstrip("/").lower()
    sess = est["portais"].get(chave)
    alvo_headless = bool(headless_flag)

    if sess and (sess.get("cpf") != (cpf or "") or sess.get("senha") != (senha or "")):
        _fechar_navegador_persistente(chave)
        sess = None
    if sess and sess.get("headless") is not None and sess["headless"] != alvo_headless:
        _fechar_navegador_persistente(chave)
        sess = None

    try:
        if sess is None or sess["browser"] is None or sess["browser"].is_connected():
            if sess is not None:
                _fechar_navegador_persistente(chave)
            from playwright.sync_api import sync_playwright
            sess = {
                "pw": sync_playwright().start(),
                "headless": alvo_headless,
                "cpf": cpf or "",
                "senha": senha or "",
                "ctx": None,
                "page": None,
            }
            sess["browser"] = sess["pw"].chromium.launch(headless=alvo_headless)
            est["portais"][chave] = sess
        pagina = sess.get("page")
        viva = False
        if pagina is not None and not pagina.is_closed():
            try:
                pagina.evaluate("1")
                viva = True
            except Exception:
                viva = False
        if not viva:
            sess["ctx"] = sess["browser"].new_context()
            sess["page"] = sess["ctx"].new_page()
            from lancar_notas_sge import ACTION_TIMEOUT_MS
            sess["page"].set_default_timeout(ACTION_TIMEOUT_MS)
        return sess["page"]
    except Exception:
        _fechar_navegador_persistente(chave)
        raise


@contextmanager
def _sessao_navegador(headless_flag, url_portal: str, cpf: str, senha: str):
    """Entrega a Page persistente do portal sem fecha-la ao fim da execucao.

    O navegador (e o login) sobrevivem entre execucoes enquanto o painel estiver
    aberto. Se algo falhar no meio, a sessao deste portal e descartada para a
    proxima execucao comecar limpa.
    """
    page = _obter_page_persistente(headless_flag, url_portal, cpf, senha)
    try:
        yield page
    except Exception:
        _fechar_navegador_persistente((url_portal or "").strip().rstrip("/").lower())
        raise


def _validar_template(fonte: str, caminho: str, tipo: str) -> tuple:
    try:
        if tipo == "notas":
            if fonte == "excel":
                import openpyxl
                wb = openpyxl.load_workbook(caminho, data_only=True)
                headers = [str(h or "") for h in next(wb.active.iter_rows(values_only=True))]
                wb.close()
            else:
                with open(caminho, encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
            req = ["nome", "aluno"]
            found = any(any(r in h.lower() for r in req) for h in headers)
            if not found:
                return False, ["Coluna 'Nome' ou 'Aluno' nao encontrada no cabecalho"]
            return True, []
        else:
            if fonte == "excel":
                import openpyxl
                wb = openpyxl.load_workbook(caminho, data_only=True)
                headers = [str(h or "") for h in next(wb.active.iter_rows(values_only=True))]
                wb.close()
            else:
                with open(caminho, encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
            h_lower = [h.lower() for h in headers]
            avisos = []
            if not any("ano" in h for h in h_lower):
                avisos.append("Coluna 'Ano' nao encontrada. O ano sera extraido da Turma.")
            if not any("escola" in h for h in h_lower):
                avisos.append("Coluna 'Escola' nao encontrada.")
            if not any("nome" in h or "name" in h for h in h_lower):
                avisos.append("Coluna 'Name'/'Nome' nao encontrada.")
            return len(avisos) < 3, avisos
    except Exception as exc:
        return False, [f"Erro ao ler arquivo: {exc}"]


def _planilha_status_store():
    from status_store import StatusStore
    return StatusStore(_PAINEL_SOURCE_PATH, logger=None)


def _aplicar_status_salvo(linhas: List[Dict]) -> None:
    """Preenche a coluna Status das linhas com o status persistido localmente."""
    from status_store import StatusStore
    store = StatusStore(_PAINEL_SOURCE_PATH, logger=None)
    for ln in linhas:
        status = store.obter_status_tolerante(
            ln.get("escola", ""), ln.get("turno", ""), ln.get("turma", ""),
            ln.get("trimestre", ""), ln.get("aluno", ""), ln.get("atividade", ""),
        )
        if status:
            ln["status"] = status


def _semear_status_manual(status_store) -> None:
    """Respeita status Lancada/Falha editados manualmente nas linhas do painel."""
    from leitor_planilhas import _to_float
    for ln in st.session_state.get("planilha_linhas", []):
        status = str(ln.get("status") or "").strip()
        if not status:
            continue
        aluno = str(ln.get("aluno") or "").strip()
        atividade = str(ln.get("atividade") or "").strip()
        if not aluno or not atividade:
            continue
        nota = _to_float(ln.get("nota"))
        if nota is None:
            nota = 0.0
        args = (
            str(ln.get("escola") or ""), str(ln.get("turno") or ""),
            str(ln.get("turma") or ""), str(ln.get("trimestre") or ""),
            aluno, atividade,
        )
        if status == "Lancada":
            status_store.marcar_lancada(*args, nota)
        elif status == "Falha":
            status_store.marcar_falha(*args, nota, erro="manual")


def _sincronizar_status_painel(status_store) -> None:
    """Grava no painel o status final registrado pelo StatusStore apos o lancamento."""
    linhas = st.session_state.get("planilha_linhas", [])
    for ln in linhas:
        status = status_store.obter_status_tolerante(
            ln.get("escola", ""), ln.get("turno", ""), ln.get("turma", ""),
            ln.get("trimestre", ""), ln.get("aluno", ""), ln.get("atividade", ""),
        )
        if status:
            ln["status"] = status
    st.session_state.planilha_linhas = linhas


def _revisao_dir() -> str:
    d = os.path.join(os.getcwd(), "artifacts", "revisao")
    os.makedirs(d, exist_ok=True)
    return d


def _item_revisao_id(escola: str, turma: str, trimestre: str, atividade: str, aluno: str) -> str:
    import hashlib
    chave = "|".join([
        str(escola).lower(), str(turma).lower(), str(trimestre).lower(),
        str(atividade).lower(), str(aluno).lower(),
    ])
    return hashlib.sha1(chave.encode("utf-8")).hexdigest()[:12]


def _coletar_divergencia(page, contexto, atividade: str, data_realizacao: str, aluno: str,
                         nota_esperada, nota_lida, coluna_sge: str = "", logger=None) -> None:
    """Enfileira uma divergencia de leitura para confirmacao do usuario (com screenshot)."""
    if "revisao_fila" not in st.session_state:
        st.session_state.revisao_fila = []
    fila = st.session_state.revisao_fila
    item_id = _item_revisao_id(contexto.escola, contexto.turma, contexto.trimestre, atividade, aluno)
    shot = os.path.join(_revisao_dir(), f"{item_id}.png")
    try:
        from lancar_notas_sge import _capturar_evidencia_divergencia
        _capturar_evidencia_divergencia(page, aluno, coluna_sge, shot, logger=logger)
    except Exception as exc:  # noqa: BLE001
        if logger is not None:
            logger(f"[REVISAO] Falha ao capturar evidencia: {exc}")
    item = {
        "id": item_id,
        "escola": contexto.escola,
        "turno": contexto.turno,
        "turma": contexto.turma,
        "trimestre": contexto.trimestre,
        "atividade": atividade,
        "data_realizacao": data_realizacao,
        "aluno": aluno,
        "nota_esperada": nota_esperada,
        "nota_lida": str(nota_lida or ""),
        "screenshot": shot,
        "coluna_sge": coluna_sge,
        "decisao": None,
        "valor_corrigido": "",
        "resolvido": False,
    }
    for i, it in enumerate(fila):
        if it.get("id") == item_id:
            fila[i].update(item)
            return
    fila.append(item)


def _coletar_ausente(page, contexto, atividade: str, data_realizacao: str, aluno: str,
                     nota_esperada, coluna_sge: str = "", logger=None) -> None:
    """Enfileira um aluno NAO localizado na grade do SGE (nome/imagem ilegivel)
    para revisao do usuario: ele pode corrigir o nome (letra ilegivel) ou enviar
    outra imagem (imagem ilegivel)."""
    if "revisao_fila" not in st.session_state:
        st.session_state.revisao_fila = []
    fila = st.session_state.revisao_fila
    from lancar_notas_sge import _coletar_ausente as _montar_ausente
    item = _montar_ausente(
        page, contexto, atividade, aluno, nota_esperada,
        coluna_sge=coluna_sge, data_realizacao=data_realizacao,
        logger=logger,
    )
    for i, it in enumerate(fila):
        if it.get("id") == item["id"]:
            fila[i].update(item)
            return
    fila.append(item)


def _coletar_retentativa_ausentes() -> int:
    """Prepara a re-tentativa dos ausentes com nome corrigido (letra ilegivel).

    Monta a lista de pendentes em st.session_state.revisao_pendentes e agenda o
    novo lancamento (revisao_retentar_ausentes + revisao_aplicar). Retorna quantos
    itens validos serao re-tentados (0 se nenhum nome foi corrigido)."""
    fila = st.session_state.get("revisao_fila", [])
    pend = []
    for it in fila:
        if it.get("tipo") != "ausente":
            continue
        if it.get("decisao") != "retentar":
            continue
        nome_corrigido = (it.get("aluno_corrigido") or "").strip()
        if not nome_corrigido:
            continue
        pend.append({
            "tipo": "ausente",
            "escola": it.get("escola"), "turno": it.get("turno"),
            "turma": it.get("turma"), "trimestre": it.get("trimestre"),
            "atividade": it.get("atividade"),
            "aluno": it.get("aluno"),
            "aluno_corrigido": nome_corrigido,
            "nota_esperada": it.get("nota_esperada"),
            "data_realizacao": it.get("data_realizacao", ""),
            "decisao": "retentar",
        })
    if not pend:
        return 0
    st.session_state.revisao_pendentes = pend
    st.session_state.revisao_retentar_ausentes = True
    st.session_state.revisao_aplicar = True
    return len(pend)


def _salvar_nova_imagem_ausente(item, uploaded) -> str:
    """Salva a nova imagem enviada pelo professor (imagem ilegivel) em
    artifacts/revisao e retorna o caminho (ou '' em caso de falha)."""
    if uploaded is None:
        return ""
    try:
        rev_dir = _revisao_dir()
        os.makedirs(rev_dir, exist_ok=True)
        ext = os.path.splitext(uploaded.name or "")[1].lower() or ".png"
        dest = os.path.join(rev_dir, f"{item.get('id', 'x')}_nova{ext}")
        with open(dest, "wb") as f:
            f.write(uploaded.getvalue())
        return dest
    except Exception:  # noqa: BLE001
        return ""


def _alimentar_fila_revisao(resultado: dict, logger=None) -> None:
    """Enfileira os itens NAO-CONFIRMADO retornados pela execucao (re-auditoria
    pos-lancamento e falhas de verificacao pos-preenchimento) para confirmacao do usuario."""
    itens = resultado.get("itens_nao_confirmados") or []
    if not itens:
        return
    if "revisao_fila" not in st.session_state:
        st.session_state.revisao_fila = []
    fila = st.session_state.revisao_fila
    novos = 0
    for it in itens:
        item_id = it.get("id")
        if not item_id:
            continue
        encontrado = next((x for x in fila if x.get("id") == item_id), None)
        if encontrado is not None:
            if encontrado.get("decisao") is None:
                encontrado.update(it)
            continue
        novo = dict(it)
        novo.setdefault("decisao", None)
        novo.setdefault("valor_corrigido", "")
        novo.setdefault("resolvido", False)
        fila.append(novo)
        novos += 1
    if novos > 0:
        st.session_state.revisao_fase = "pendente"
        if logger is not None:
            logger(f"[REVISAO] {novos} item(ns) de revisao aguardando confirmacao manual.")


def _aplicar_decisoes_revisao():
    """Coleta itens decididos (confirmar/corrigir) e agenda gravacao forcada no SGE."""
    fila = st.session_state.get("revisao_fila", [])
    pendentes = []
    for item in fila:
        if item.get("decisao") not in ("confirmar", "corrigir"):
            continue
        pendentes.append({
            "escola": item.get("escola"), "turno": item.get("turno"),
            "turma": item.get("turma"), "trimestre": item.get("trimestre"),
            "atividade": item.get("atividade"),
            "aluno": item.get("aluno"),
            "decisao": item.get("decisao"),
            "valor_corrigido": item.get("valor_corrigido", ""),
            "nota_esperada": item.get("nota_esperada"),
        })
    if not pendentes:
        return False
    st.session_state.revisao_pendentes = pendentes
    st.session_state.revisao_aplicar = True
    return True


def _remodelar_estrutura_com_ia(evidencia: dict) -> str:
    """Analisa o screenshot do alarme [ESTRUTURA-CHANGED] via IA e sugere novos seletores.

    Consulta a IA SOMENTE aqui (no desvio), nunca no fluxo normal. Retorna o texto
    JSON da sugestao para o usuario revisar antes de aprovar qualquer override.
    """
    try:
        from ai_assist import analyze_portal_failure
        shot = (evidencia.get("evidencia") or {}).get("screenshot", "")
        if not shot or not os.path.exists(shot):
            return "Screenshot de evidencia nao encontrado em artifacts/estrutura/."
        with open(shot, "rb") as f:
            bytes_shot = f.read()
        result = analyze_portal_failure(
            bytes_shot,
            error="ESTRUTURA-CHANGED: grade de notas do SGE nao reconhecida",
            operation="remodelar_estrutura",
            context="",
        )
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as exc:  # noqa: BLE001
        return f"Erro na analise da IA: {exc}"


def _to_suffix_selector(sel: str) -> str:
    """Normaliza um seletor sugerido pela IA para ficar especifico ao aluno ({suffix}).

    Seletores genericos de atributo (ex.: input[name*='NOTA']) ganham um sufixo
    extra para casar apenas a linha do aluno quando o chain substituir {suffix}.
    """
    sel = (sel or "").strip()
    if not sel:
        return ""
    if "{suffix}" in sel:
        return sel
    if sel.endswith("]") and "[" in sel:
        return sel + "[name*='_{suffix}']"
    return sel


def _salvar_estrutura_override_do_sugestao(sugestao_texto: str) -> bool:
    """Extrai seletores da sugestao da IA e persiste o override LOCAL (so com OK do usuario).

    Formato salvo em artifacts/estrutura/estrutura_override.json:
      {"slot_selector": "...", "grade_selectors": ["..."], "fonte_ia": true}
    """
    try:
        data = json.loads(sugestao_texto)
    except Exception:  # noqa: BLE001
        return False

    from lancar_notas_sge import _salvar_estrutura_override

    override: dict = {}
    grade = data.get("grade_flow") or {}
    slot_sel = _to_suffix_selector(grade.get("student_name_selector") or "")
    if slot_sel and "{suffix}" not in slot_sel:
        override["slot_selector"] = slot_sel
    grade_sel_raw = grade.get("grade_input_selector") or ""
    grade_selectors: list = []
    if isinstance(grade_sel_raw, str) and grade_sel_raw.strip():
        grade_selectors.append(_to_suffix_selector(grade_sel_raw))
    elif isinstance(grade_sel_raw, list):
        for g in grade_sel_raw:
            if isinstance(g, str) and g.strip():
                grade_selectors.append(_to_suffix_selector(g))

    for fx in data.get("suggested_fixes") or []:
        if isinstance(fx, dict) and fx.get("selector"):
            grade_selectors.append(_to_suffix_selector(str(fx["selector"])))
    grade_selectors = [s for s in grade_selectors if s]
    if grade_selectors:
        override["grade_selectors"] = grade_selectors

    if not override:
        return False
    override["fonte_ia"] = True
    return _salvar_estrutura_override(override)


def _ler_arquivo_texto(path: str) -> str:
    try:
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
    except Exception:  # noqa: BLE001
        pass
    return ""


def salvar_config():
    config = {
        "sge_url": st.session_state.get("sge_url", ""),
        "sge_cpf": st.session_state.get("sge_cpf", ""),
        "sge_senha": st.session_state.get("sge_senha", ""),
        "notion_token": st.session_state.get("notion_token", ""),
        "root_page_id": st.session_state.get("root_page_id", ""),
        "gemini_key": st.session_state.get("gemini_key", ""),
        "gemini_model": st.session_state.get("gemini_model", "gemini-2.5-flash"),
        "openai_key": st.session_state.get("openai_key", ""),
        "openai_model": st.session_state.get("openai_model", "gpt-4o"),
        "anthropic_key": st.session_state.get("anthropic_key", ""),
        "anthropic_model": st.session_state.get("anthropic_model", "claude-sonnet-4-20250514"),
        "ai_provider": st.session_state.get("ai_provider", "local"),
        "ollama_model": st.session_state.get("ollama_model", "llama3.2-vision"),
        "fonte": st.session_state.get("fonte", "notion"),
        "escola": st.session_state.get("escola", ""),
        "turno": st.session_state.get("turno", ""),
        "turma": st.session_state.get("turma", ""),
        "trimestre": st.session_state.get("trimestre", ""),
        "avaliacao_nome": st.session_state.get("avaliacao_nome", ""),
        "avaliacao_data": st.session_state.get("avaliacao_data", ""),
        "tipo": st.session_state.get("tipo", "notas"),
        "seq_titulo_documento": st.session_state.get("seq_titulo_documento", ""),
        "seq_periodo_inicio": st.session_state.get("seq_periodo_inicio", ""),
        "seq_periodo_fim": st.session_state.get("seq_periodo_fim", ""),
        "seq_n_aulas": st.session_state.get("seq_n_aulas", 4),
        "seq_drive_links": st.session_state.get("seq_drive_links", []),
    }
    config_dir = Path.home() / ".sge_bot"
    config_dir.mkdir(exist_ok=True)
    config_path = config_dir / "config.json"
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    st.session_state.config_path = str(config_path)
    return str(config_path)


def carregar_config():
    config_paths = [
        Path.home() / ".sge_bot" / "config.json",
        Path(".env"),
        Path(".env.local"),
    ]
    for path in config_paths:
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError):
                try:
                    from dotenv import load_dotenv
                    load_dotenv(path)
                except ImportError:
                    pass
    return {}


def carregar_env():
    for env_file in [".env", ".env.local", Path.home() / ".sge_bot" / ".env"]:
        if os.path.exists(str(env_file)):
            try:
                from dotenv import load_dotenv
                load_dotenv(str(env_file))
            except ImportError:
                pass

    for key in ["SGE_LOGIN_URL", "SGE_CPF", "SGE_SENHA", "NOTION_TOKEN", "ROOT_PAGE_ID", "GEMINI_API_KEY", "OPENAI_API_KEY", "ANTHROPIC_API_KEY", "AI_PROVIDER", "OLLAMA_HOST"]:
        val = os.environ.get(key, "")
        if val:
            return val
    return ""


def _listar_turmas_notion():
    """Lista as turmas disponiveis no Notion para os filtros atuais.

    Quando um filtro generico (ex.: "6o Ano") casa com mais de uma turma
    (ex.: "6º Ano 1" e "6º Ano 2"), o lancamento automatico monta blocos de
    todas elas e a navegacao entre turmas no SGE pode falhar. Este helper
    alimenta o seletor que permite lancar UMA turma por vez.
    """
    if st.session_state.notion_token:
        os.environ["NOTION_TOKEN"] = st.session_state.notion_token
    if st.session_state.root_page_id:
        os.environ["ROOT_PAGE_ID"] = st.session_state.root_page_id

    # lancar_notas_sge le NOTION_TOKEN/ROOT_PAGE_ID no import (globais de
    # modulo); sem reload, o token preenchido na tela nao e visto pelo botao.
    import importlib
    import lancar_notas_sge as _sge_mod
    importlib.reload(_sge_mod)
    from lancar_notas_sge import _normalize, listar_contextos_disponiveis

    escola = (st.session_state.get("escola") or "").strip()
    turno = (st.session_state.get("turno") or "").strip()
    trimestre = (st.session_state.get("trimestre") or "").strip()

    def _sub(value: str, expected: str) -> bool:
        if not expected:
            return True
        return _normalize(expected) in _normalize(value)

    turmas = set()
    try:
        for ctx in listar_contextos_disponiveis():
            if not _sub(ctx.get("escola", ""), escola):
                continue
            if not _sub(ctx.get("turno", ""), turno):
                continue
            if not _sub(ctx.get("trimestre", ""), trimestre):
                continue
            if ctx.get("turma"):
                turmas.add(ctx["turma"])
    except Exception as exc:  # noqa: BLE001
        st.error(f"Falha ao carregar turmas do Notion: {exc}")
        return []
    return sorted(turmas)


# Carrega config salva para preencher campos automaticamente
config_salva = carregar_config()
for k, v in config_salva.items():
    if v and k not in st.session_state:
        st.session_state[k] = v
# Default para ai_provider se nao veio da config
if "ai_provider" not in st.session_state:
    st.session_state.ai_provider = "local"

# Defaults para campos de sequencia (Google Drive)
for _k, _v in [("seq_titulo_documento", ""), ("seq_periodo_inicio", ""), ("seq_periodo_fim", ""), ("seq_n_aulas", 4)]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# Defaults para preferencias de exibicao
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False
if "headless_mode" not in st.session_state:
    st.session_state.headless_mode = True

# === VALIDACAO DE LICENCA ===
BOT_LOCAL_CONFIG = Path.home() / ".bot_local" / "config.json"

def _load_license_cache():
    if not BOT_LOCAL_CONFIG.exists():
        return {}
    try:
        with open(BOT_LOCAL_CONFIG, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        key = cfg.get("license_key", "")
        validated = cfg.get("license_validated_at", "")
        plan = cfg.get("license_plan", "")
        expires = cfg.get("license_expires_at", "")
        if key and validated:
            try:
                dt = datetime.fromisoformat(validated)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if (datetime.now(timezone.utc) - dt).days < 7:
                    return {"key": key, "plan": plan, "expires": expires}
            except (ValueError, TypeError):
                pass
    except Exception:
        pass
    return {}

def _save_license_cache(key, plan, expires_at):
    try:
        BOT_LOCAL_CONFIG.parent.mkdir(parents=True, exist_ok=True)
        cfg = {}
        if BOT_LOCAL_CONFIG.exists():
            with open(BOT_LOCAL_CONFIG, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        cfg["license_key"] = key.strip().upper()
        cfg["license_plan"] = plan or ""
        cfg["license_expires_at"] = expires_at or ""
        cfg["license_validated_at"] = datetime.now(timezone.utc).isoformat()
        with open(BOT_LOCAL_CONFIG, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

def _validate_license_online(key):
    data = json.dumps({"license_key": key.strip().upper()}).encode("utf-8")
    req = urllib.request.Request(
        f"{LICENSE_SERVER_URL}/api/license/public-validate",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        resp = urllib.request.urlopen(req, timeout=15)
        result = json.loads(resp.read().decode("utf-8"))
        return result.get("valid", False), result
    except (urllib.error.URLError, OSError, json.JSONDecodeError) as e:
        return False, {"error": str(e)}

def _is_network_error(data):
    """True se a falha for de rede/servidor (nao uma revogacao real de licenca)."""
    err = (data or {}).get("error", "") or ""
    lowered = err.lower()
    markers = ("servidor", "indisponivel", "timed out", "connection",
               "resolve", "network", "errno", "http error")
    return (not err) or any(m in lowered for m in markers)

if "license_valid" not in st.session_state:
    cached = _load_license_cache()
    if cached:
        valid, data = _validate_license_online(cached["key"])
        if valid:
            st.session_state.license_valid = True
            st.session_state.license_key = cached["key"]
            st.session_state.license_plan = data.get("plan", cached.get("plan", ""))
        elif _is_network_error(data):
            st.session_state.license_valid = True
            st.session_state.license_key = cached["key"]
            st.session_state.license_plan = cached.get("plan", "")
        else:
            st.session_state.license_valid = False
            st.session_state.license_key = ""
            st.session_state.license_plan = ""
            st.session_state.license_error = data.get("error", "Assinatura finalizada ou cancelada")
            st.session_state.license_resubscribe_url = data.get("resubscribe_url", "")
    else:
        st.session_state.license_valid = False
        st.session_state.license_key = ""
        st.session_state.license_plan = ""

if not st.session_state.license_valid:
    st.markdown("""
    <style>
        .blocker-bg {
            position: fixed; top: 0; left: 0; width: 100%; height: 100%;
            background: linear-gradient(135deg, #0a0e27, #1a1f3a);
            z-index: -1;
        }
        .blocker-card {
            background: #12162e; border-radius: 16px; padding: 32px;
            max-width: 440px; margin: 0 auto; text-align: center;
            box-shadow: 0 8px 32px rgba(0,0,0,0.5);
        }
        .blocker-card h1 {
            font-size: 1.8rem; font-weight: 700; margin-bottom: 4px;
            background: linear-gradient(135deg, #2563eb, #7c3aed);
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        }
        .blocker-card p { color: #94a3b8; font-size: 0.95rem; }
    </style>
    <div class="blocker-bg"></div>
    <div style="height:64px"></div>
    <div class="blocker-card">
        <h1>BotDoProfessor</h1>
        <p>Ative sua licenca para acessar o painel</p>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        err_msg = st.session_state.get("license_error", "")
        if err_msg:
            st.error(err_msg)
        lic_key = st.text_input("Chave de licenca", key="lic_input",
                                placeholder="Cole sua chave aqui")
        if st.button("Validar", type="primary", use_container_width=True):
            if not lic_key.strip():
                st.error("Digite a chave de licenca")
            else:
                with st.spinner("Validando..."):
                    valid, data = _validate_license_online(lic_key)
                    if valid:
                        _save_license_cache(lic_key, data.get("plan", ""), data.get("expires_at", ""))
                        st.session_state.license_valid = True
                        st.session_state.license_key = lic_key.upper()
                        st.session_state.license_plan = data.get("plan", "")
                        st.rerun()
                    else:
                        st.error(data.get("error", "Chave invalida ou expirada"))
                        url = data.get("resubscribe_url")
                        if url:
                            st.markdown(
                                f'<a href="{url}" target="_blank" '
                                'style="display:inline-block;margin-top:6px;">'
                                '<button style="background:#e94560;color:white;border:none;'
                                'padding:10px 18px;border-radius:8px;font-weight:bold;'
                                'cursor:pointer;width:100%;">'
                                'Corrigir pagamento / Assinar / Reassinar</button></a>',
                                unsafe_allow_html=True,
                            )
        resub = st.session_state.get("license_resubscribe_url") or "https://botdoprofessor.onrender.com/checkout"
        st.markdown(
            f'<a href="{resub}" target="_blank">'
            'Nao tem chave? Compre ou reassine aqui</a>',
            unsafe_allow_html=True,
        )
    st.stop()


# === BARRA LATERAL ===
with st.sidebar:
    plan_label = st.session_state.get("license_plan", "").capitalize() if st.session_state.get("license_plan") else ""
    if plan_label:
        st.markdown(
            f"<div style='background:#1a2a1a;border:1px solid #2a5a2a;border-radius:8px;padding:6px 12px;margin-bottom:12px;font-size:0.85rem;color:#5ae05a;text-align:center;'>"
            f"Licenca {plan_label}</div>",
            unsafe_allow_html=True,
        )
    st.markdown("### Configuracoes")

    with st.expander("Portal do Professor", expanded=True):
        portal_options = ["SGE", "Professor Online", "Novo Portal", "Auto (detecta pela escola)"]
        portal_index = 0
        saved_portal = st.session_state.get("portal_selecionado", "SGE")
        if saved_portal in portal_options:
            portal_index = portal_options.index(saved_portal)
        portal_selecionado = st.selectbox(
            "Portal",
            options=portal_options,
            index=portal_index,
            key="portal_selecionar",
            help="SGE ou Professor Online: bot ja conhece o acesso. Novo Portal: a IA local aprende o acesso. Auto: detecta o portal pela escola filtrada."
        )
        st.session_state.portal_selecionado = portal_selecionado

        if portal_selecionado == "Auto (detecta pela escola)":
            st.caption(
                "**Auto**: o bot consulta o registro de escolas (~/.sge_bot/escolas.json). "
                "Se a escola filtrada for conhecida do **Professor Online**, executa nele; "
                "caso contrario, usa o **SGE**. As credenciais do Professor Online ficam salvas na config."
            )

        if portal_selecionado == "Professor Online":
            portal_default_url = "https://professoronline.sed.sc.gov.br"
            portal_url_key = "po_url_input"
            portal_cpf_key = "po_cpf_input"
            portal_senha_key = "po_senha_input"
            portal_env = os.environ.get("PO_URL", os.environ.get("PO_BASE_URL", ""))
        elif portal_selecionado == "Novo Portal":
            portal_default_url = ""
            portal_url_key = "np_url_input"
            portal_cpf_key = "np_cpf_input"
            portal_senha_key = "np_senha_input"
            portal_env = os.environ.get("NP_URL", "")
        else:
            portal_default_url = "https://www.sge8147.com.br/hportalprofessor.aspx"
            portal_url_key = "sge_url_input"
            portal_cpf_key = "sge_cpf_input"
            portal_senha_key = "sge_senha_input"
            portal_env = carregar_env()

        if st.session_state.get("portal_url_para") != portal_selecionado:
            st.session_state["portal_url_para"] = portal_selecionado
            st.session_state["portal_url_value"] = portal_env or portal_default_url

        portal_url = st.text_input(
            "URL do Portal do Professor",
            value=st.session_state.get("portal_url_value", portal_env or portal_default_url),
            key=portal_url_key,
            placeholder="https://...",
        )
        st.session_state.portal_url_value = portal_url

        col1, col2 = st.columns(2)
        with col1:
            cpf = st.text_input("CPF", value=st.session_state.get("portal_cpf_value", config_salva.get("sge_cpf", "")), key=portal_cpf_key, type="password")
            st.session_state.portal_cpf_value = cpf
        with col2:
            senha = st.text_input("Senha", value=st.session_state.get("portal_senha_value", config_salva.get("sge_senha", "")), key=portal_senha_key, type="password")
            st.session_state.portal_senha_value = senha

        if st.button("Fechar navegador do portal", help="Encerra a sessao que fica aberta entre execucoes. Use se o portal travar ou para entrar com outro usuario."):
            _fechar_navegador_persistente()
            st.toast("Navegador fechado. A proxima execucao fara novo login.")
        st.caption("Entre execucoes o bot reaproveita o portal ja aberto, sem repetir login. Trocou de portal ou de senha? A sessao e renovada automaticamente.")

        if (senha or "").strip().lower() in {"123456", "12345678", "12345", "123", "senha", "sua_senha", "teste", "test", "password"}:
            st.warning(
                "A senha atual parece ser de **teste/placeholder**. Para o login funcionar, informe "
                "a senha REAL do portal e o CPF completo (11 digitos, apenas numeros)."
            )

        if portal_selecionado == "Novo Portal":
            portal_nome = st.text_input(
                "Nome do novo portal (opcional)",
                value=st.session_state.get("portal_nome_value", ""),
                key="np_nome_input",
                help="Ex.: 'Prefeitura X', 'SGP Municipal'. Se vazio, o nome e derivado da URL.",
            )
            st.session_state.portal_nome_value = portal_nome
            st.warning(
                "Portal novo: o bot ainda nao conhece este acesso. "
                "Ao clicar em EXECUTAR, o **modo aprendizado sera ativado automaticamente**. "
                "O navegador abrira visivel e voce fara o acesso manualmente enquanto a IA local grava e aprende."
            )

        # Mantem compativeis com o fluxo existente (chaves SGE)
        st.session_state.sge_url = portal_url
        st.session_state.sge_cpf = cpf
        st.session_state.sge_senha = senha
        st.session_state.po_url = portal_url
        st.session_state.po_cpf = cpf
        st.session_state.po_senha = senha
        st.session_state.np_url = portal_url
        st.session_state.np_cpf = cpf
        st.session_state.np_senha = senha

    with st.expander("Chaves de conexao (Notion)", expanded=True):
        st.caption(
            "**O que e isso?** O Notion guarda suas planilhas de notas. Para o bot ler as notas "
            "de la, ele precisa de duas 'chaves' (codigos de acesso). **So preencha se voce usa Notion** — "
            "quem usa Google Forms ou Excel pode pular esta secao."
        )
        st.caption(
            "**Como obter:** 1) No navegador, acesse notion.so/profile/sessions (logado) e clique em "
            "'Criar integracao' em notion.so/my-integrations; 2) Copie a chave que comeca com **secret_** e cole abaixo; "
            "3) Abra a pagina do seu boletim no Notion, clique nos '...' > 'Conexoes' > adicione sua integracao; "
            "4) O **ID da pagina** sao os 32 caracteres no final da URL da pagina."
        )
        notion_token = st.text_input(
            "Chave do Notion (comeca com secret_)",
            value=st.session_state.get("notion_token", ""),
            key="notion_token_input",
            type="password",
        )
        st.session_state.notion_token = notion_token

        root_page_id = st.text_input(
            "ID da pagina raiz (32 caracteres)",
            value=st.session_state.get("root_page_id", ""),
            key="root_page_id_input",
        )
        st.session_state.root_page_id = root_page_id

    st.markdown("---")
    ai_provider_options = ["local", "gemini", "openai", "anthropic"]
    ai_provider_index = 0
    saved_provider = st.session_state.get("ai_provider", "")
    if saved_provider in ai_provider_options:
        ai_provider_index = ai_provider_options.index(saved_provider)
    ai_provider = st.selectbox(
        "Provedor IA",
        options=ai_provider_options,
        format_func=lambda x: {
            "local": "Local (Ollama - minicpm-v4.6) [RECOMENDADO]",
            "gemini": "Google Gemini (API Key)",
            "openai": "OpenAI GPT-4o (API Key)",
            "anthropic": "Anthropic Claude (API Key)",
        }.get(x, x),
        index=ai_provider_index,
        key="ai_provider_select",
    )
    st.session_state.ai_provider = ai_provider

    st.caption(
        "**Dica:** a IA Local (Ollama) é gratuita, privada e não precisa de internet. "
        "Porém, para **portais complexos ou novos**, conectar **sua própria API** "
        "**(Gemini/OpenAI/Anthropic)** deixa a navegação híbrida **mais rápida e precisa** "
        "(modelos em nuvem lidam melhor com HTML extenso). Fica a seu critério: "
        "Local para custo zero, API para máximo desempenho."
    )

    # Campo de API key dinamico conforme provider
    if ai_provider == "local":
        st.info(
            "**IA Local (Recomendada)** - Nao precisa de API key!\n\n"
            "O modelo `minicpm-v4.6` e baixado automaticamente na primeira execucao (~1.6GB).\n\n"
            "**O que faz:**\n"
            "- Analisa screenshots quando o bot falha\n"
            "- Sugere seletores CSS alternativos\n"
            "- Redescobre portais automaticamente\n"
            "- Aprende com cada execucao\n\n"
            "**Requisitos:** RAM minima 4GB, processador com 4+ cores"
        )
        ollama_model = st.selectbox(
            "Modelo Ollama",
            options=["openbmb/minicpm-v4.6", "llava:7b", "llava:13b", "bakllava"],
            index=0,
            key="ollama_model_select",
        )
        st.session_state.ollama_model = ollama_model
    elif ai_provider == "gemini":
        st.markdown(
            "**Como obter a API Key do Google Gemini:**\n\n"
            "1. Acesse https://aistudio.google.com/apikey\n"
            "2. Clique em **Create API Key**\n"
            "3. Copie a chave (comeca com `AIza...`)\n"
            "4. Cole abaixo\n\n"
            "**Gratuito:** 15 requests/min, 1500 requests/dia"
        )
        gemini_key = st.text_input(
            "Gemini API Key (obrigatoria)",
            value=st.session_state.get("gemini_key", ""),
            key="gemini_key_input",
            type="password",
        )
        st.session_state.gemini_key = gemini_key
        gemini_model = st.text_input(
            "Modelo Gemini",
            value=st.session_state.get("gemini_model", "gemini-2.5-flash"),
            key="gemini_model_input",
        )
        st.session_state.gemini_model = gemini_model
    elif ai_provider == "openai":
        st.markdown(
            "**Como obter a API Key da OpenAI:**\n\n"
            "1. Acesse https://platform.openai.com/api-keys\n"
            "2. Clique em **Create new secret key**\n"
            "3. Copie a chave (comeca com `sk-...`)\n"
            "4. Cole abaixo\n\n"
            "**Pago:** $2.50 por 1M tokens de input (GPT-4o-mini)"
        )
        openai_key = st.text_input(
            "OpenAI API Key (obrigatoria)",
            value=st.session_state.get("openai_key", ""),
            key="openai_key_input",
            type="password",
        )
        st.session_state.openai_key = openai_key
        openai_model = st.text_input(
            "Modelo OpenAI",
            value=st.session_state.get("openai_model", "gpt-4o"),
            key="openai_model_input",
        )
        st.session_state.openai_model = openai_model
    elif ai_provider == "anthropic":
        st.markdown(
            "**Como obter a API Key da Anthropic:**\n\n"
            "1. Acesse https://console.anthropic.com/settings/keys\n"
            "2. Clique em **Create Key**\n"
            "3. Copie a chave (comeca com `sk-ant-...`)\n"
            "4. Cole abaixo\n\n"
            "**Pago:** $3 por 1M tokens de input (Claude Sonnet)"
        )
        anthropic_key = st.text_input(
            "Anthropic API Key (obrigatoria)",
            value=st.session_state.get("anthropic_key", ""),
            key="anthropic_key_input",
            type="password",
        )
        st.session_state.anthropic_key = anthropic_key
        anthropic_model = st.text_input(
            "Modelo Anthropic",
            value=st.session_state.get("anthropic_model", "claude-sonnet-4-20250514"),
            key="anthropic_model_input",
        )
        st.session_state.anthropic_model = anthropic_model

    with st.expander("Origem dos Dados", expanded=True):
        st.markdown("**De onde vao os dados das notas?**")
        fonte = st.selectbox(
            "Selecione a origem",
            options=["notion", "imagem", "planilha", "excel", "csv", "google_sheets", "google_drive"],
            format_func=lambda x: {
                "notion": "Notion (bancos de dados)",
                "imagem": "Imagem / Foto (extrair notas com IA)",
                "planilha": "Planilha no painel (editar aqui)",
                "excel": "Arquivo Excel (.xlsx)",
                "csv": "Arquivo CSV",
                "google_sheets": "Google Sheets (planilha online)",
                "google_drive": "Google Drive (arquivo compartilhado)",
            }.get(x, x),
            index=0,
            key="fonte_select",
        )
        st.session_state.fonte = fonte

        if fonte == "notion":
            st.info(
                "Os dados serao buscados automaticamente das databases do Notion "
                "configuradas nas API Keys acima."
            )
        elif fonte == "imagem":
            st.info(
                "Envie uma foto ou print na seção 'Filtros' abaixo. "
                "A IA extrairá as notas automaticamente."
            )
        elif fonte in ("excel", "csv"):
            ext = "XLSX / XLS" if fonte == "excel" else "CSV"
            st.info(f"Selecione o arquivo **{ext}** do seu computador.")
            arquivo = st.file_uploader(
                f"Clique aqui para selecionar o arquivo {fonte.upper()}",
                type=["xlsx", "xls"] if fonte == "excel" else ["csv"],
            )
            if arquivo:
                tmp_dir = Path(tempfile.gettempdir()) / "sge_bot_uploads"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / arquivo.name
                with open(tmp_path, "wb") as f:
                    f.write(arquivo.getbuffer())
                st.session_state["arquivo_path"] = str(tmp_path)
                st.success(f"Arquivo salvo: {arquivo.name}")
                valido, avisos = _validar_template(fonte, tmp_path, st.session_state.get("tipo", "notas"))
                if valido:
                    st.success("Template valido! Pronto para executar.")
                for aviso in avisos:
                    st.warning(aviso)
        elif fonte == "planilha":
            st.info(
                "Edite as notas na aba **Planilha** (no corpo do painel) ou carregue "
                "um arquivo **.xlsx/.csv** la para preencher a tabela. A coluna **Status** "
                "(Lancada/Falha) pode ser editada por linha e fica salva entre execucoes."
            )
            with st.expander("Importar relatorio do Google Forms"):
                st.markdown(
                    "Para avaliacoes aplicadas via **Google Forms**: baixe o relatorio "
                    "(.xlsx) e importe aqui. O bot converte a pontuacao bruta em nota "
                    "(acertos x valor por questao, ex.: 19 x 0,5 = 9,5), mantendo apenas "
                    "a ultima resposta de cada aluno. Nao afeta a leitura das outras "
                    "planilhas nem do Notion."
                )
                gf_arquivo = st.file_uploader(
                    "Relatorio .xlsx do Google Forms",
                    type=["xlsx"],
                    key="gf_upload",
                )
                gf_total_auto = None
                gf_turmas: List[str] = []
                if gf_arquivo:
                    gf_dir = Path(tempfile.gettempdir()) / "sge_bot_uploads"
                    gf_dir.mkdir(exist_ok=True)
                    gf_path = gf_dir / gf_arquivo.name
                    with open(gf_path, "wb") as f:
                        f.write(gf_arquivo.getbuffer())
                    st.session_state["gf_path"] = str(gf_path)
                    try:
                        from leitor_planilhas import detectar_total_questoes_forms, detectar_turmas_forms
                        gf_total_auto = detectar_total_questoes_forms(str(gf_path))
                        gf_turmas = detectar_turmas_forms(str(gf_path))
                    except Exception:
                        gf_total_auto = None
                        gf_turmas = []
                    if gf_total_auto:
                        st.caption(f"Total de questoes detectado: **{gf_total_auto}**")
                gf_n_q = st.number_input(
                    "Total de questoes",
                    min_value=1,
                    step=1,
                    value=int(gf_total_auto or 20),
                    key="gf_n_questoes",
                )
                gf_valor = st.number_input(
                    "Valor de cada questao (pontos)",
                    min_value=0.05,
                    max_value=10.0,
                    step=0.05,
                    value=0.5,
                    key="gf_valor_questao",
                )
                gf_atividade = st.text_input(
                    "Nome da atividade (igual ao portal SGE)",
                    key="gf_atividade",
                    placeholder="Ex: Avaliacao de Humanas - Livros Sagrados",
                )
                _tri_opcoes = ["", "1o Trimestre", "2o Trimestre", "3o Trimestre"]
                _tri_atual = str(st.session_state.get("trimestre") or "")
                gf_trimestre = st.selectbox(
                    "Trimestre (preenchido na tabela)",
                    options=_tri_opcoes,
                    index=_tri_opcoes.index(_tri_atual) if _tri_atual in _tri_opcoes else 0,
                    key="gf_trimestre",
                    help="Gravado em cada linha importada e usado para abrir o "
                         "periodo certo no portal. A data nao e necessaria: a "
                         "avaliacao e localizada pelo nome.",
                )
                gf_mapa_turno: Dict[str, str] = {}
                if gf_turmas:
                    st.markdown("**Turno de cada turma** (o bot abre a turma certa no portal):")
                    from leitor_planilhas import _normalize as _norm_turma
                    for _t in gf_turmas:
                        _escolhido = st.selectbox(
                            f"Turma {_t}",
                            options=["", "Matutino", "Vespertino", "Noturno"],
                            key=f"gf_turno_{_norm_turma(_t)}",
                        )
                        if _escolhido:
                            gf_mapa_turno[_norm_turma(_t)] = _escolhido
                    if len(gf_mapa_turno) < len(gf_turmas):
                        st.caption(
                            "Deixe o filtro de Turno (aba Filtros) **vazio** para "
                            "lancar em todos os turnos; cada turma usara o turno "
                            "marcado acima."
                        )
                if st.button("Importar para a tabela", key="gf_importar", use_container_width=True):
                    gf_caminho = st.session_state.get("gf_path", "")
                    if not gf_caminho:
                        st.error("Selecione o arquivo .xlsx do Google Forms primeiro.")
                    else:
                        from leitor_planilhas import (
                            ler_notas_google_forms,
                            registros_para_linhas,
                            _normalize as _norm_linha,
                        )
                        try:
                            gf_regs = ler_notas_google_forms(
                                gf_caminho,
                                valor_questao=float(gf_valor),
                                atividade=gf_atividade,
                                n_questoes=int(gf_n_q),
                                logger=log,
                            )
                        except Exception as exc:
                            st.error(f"Erro ao ler o relatorio: {exc}")
                            gf_regs = []
                        if gf_regs:
                            gf_linhas = registros_para_linhas(gf_regs)
                            for ln in gf_linhas:
                                t_norm = _norm_linha(str(ln.get("turma") or ""))
                                if t_norm in gf_mapa_turno:
                                    ln["turno"] = gf_mapa_turno[t_norm]
                                if gf_trimestre:
                                    ln["trimestre"] = gf_trimestre
                            _aplicar_status_salvo(gf_linhas)
                            st.session_state.planilha_linhas = gf_linhas
                            sem_turno = sum(
                                1 for ln in gf_linhas if not str(ln.get("turno") or "").strip()
                            ) if gf_mapa_turno else 0
                            msg = f"{len(gf_regs)} aluno(s) importado(s)."
                            if sem_turno:
                                msg += f" Atencao: {sem_turno} linha(s) sem turno definido."
                            st.success(msg + " Confira na aba **Planilha** e execute.")
                            st.rerun()
                        else:
                            st.warning(
                                "Nenhuma resposta valida encontrada. Verifique se o "
                                "arquivo tem as colunas de nome e Score/Pontuacao."
                            )
        elif fonte == "google_sheets":
            st.info(
                "Cole o link compartilhavel do **Google Sheets** abaixo.\n\n"
                "Ex: `https://docs.google.com/spreadsheets/d/...`"
            )
            link = st.text_input(
                "URL compartilhavel do Google Sheets:",
                key="link_input",
                placeholder="https://docs.google.com/spreadsheets/d/...",
            )
            st.session_state["link_url"] = link
        elif fonte == "google_drive":
            st.info(
                "Cole os links dos arquivos do Drive na seção **'Filtros'** abaixo, "
                "um para cada ano."
            )

    with st.expander("Download de Templates", expanded=False):
        st.markdown("**Baixe o modelo, preencha, e use no bot:**")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            st.markdown("**Notas (alunos)**")
            st.download_button(
                "📥 .xlsx",
                data=gerar_template_notas_xlsx(),
                file_name="template_notas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.download_button(
                "📥 .csv",
                data=gerar_template_notas_csv(),
                file_name="template_notas.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with col_t2:
            st.markdown("**Sequência Didática**")
            st.download_button(
                "📥 .xlsx",
                data=gerar_template_sequencias_xlsx(),
                file_name="template_sequencia_didatica.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.download_button(
                "📥 .csv",
                data=gerar_template_sequencias_csv(),
                file_name="template_sequencia_didatica.csv",
                mime="text/csv",
                use_container_width=True,
            )

    with st.expander("Filtros"):
        tipo = st.radio(
            "Tipo de lancamento:",
            options=["notas", "chamada", "sequencia", "faltas"],
            format_func=lambda x: {
                "notas": "Notas",
                "chamada": "Chamada (foto do diario)",
                "sequencia": "Sequência Didática",
                "faltas": "Faltas do mês (leitura)",
            }.get(x, x),
            horizontal=True,
            key="tipo_radio",
        )
        st.session_state.tipo = tipo

        col_f1, col_f2 = st.columns(2)
        _linhas_painel = (
            st.session_state.get("planilha_linhas", [])
            if st.session_state.get("fonte", "") == "planilha" else []
        )
        with col_f1:
            from leitor_planilhas import valores_distintos as _val_dist
            _escolas_disp = _val_dist(_linhas_painel, "escola")
            if _escolas_disp:
                escola = st.selectbox(
                    "Escola",
                    options=[""] + _escolas_disp,
                    key="escola_opcoes",
                    help="Opcoes vindas da tabela do painel.",
                )
                st.session_state["escola_input"] = escola
            else:
                escola = st.text_input(
                    "Escola",
                    key="escola_input",
                    placeholder="vazio = todas as escolas",
                )
            st.session_state.escola = escola
            _turmas_disp = _val_dist(_linhas_painel, "turma")
            if _turmas_disp:
                turma = st.selectbox(
                    "Turma",
                    options=[""] + _turmas_disp,
                    key="turma_opcoes",
                    help="Opcoes vindas da tabela do painel.",
                )
                st.session_state["turma_input"] = turma
            else:
                turma = st.text_input(
                    "Turma (ex: 6o Ano)",
                    key="turma_input",
                    placeholder="vazio = todas as turmas",
                )
            st.session_state.turma = turma
        with col_f2:
            turno = st.selectbox("Turno", options=["", "Matutino", "Vespertino", "Noturno"], key="turno_select")
            st.session_state.turno = turno
            trimestre = st.selectbox("Trimestre", options=["", "1o Trimestre", "2o Trimestre", "3o Trimestre"], key="trimestre_select")
            st.session_state.trimestre = trimestre

        # Com a tabela do painel carregada, Turno/Trimestre selecionados nos
        # filtros preenchem as CELULAS VAZIAS correspondentes SOMENTE nas
        # linhas que casam com os demais filtros (Escola/Turma), para o turno
        # de uma turma nao vazar para outra (ex.: 6°2 vespertino nao pode
        # preencher as linhas da 6°1).
        if st.session_state.get("fonte", "") == "planilha" and _linhas_painel:
            _filtros_p = {
                "escola": str(st.session_state.get("escola") or "").strip(),
                "turno": str(st.session_state.get("turno") or "").strip(),
                "turma": str(st.session_state.get("turma") or "").strip(),
                "trimestre": str(st.session_state.get("trimestre") or "").strip(),
            }
            _preenche_turno = bool(_filtros_p["turno"])
            _preenche_tri = bool(_filtros_p["trimestre"])
            if _preenche_turno or _preenche_tri:
                from leitor_planilhas import _normalize as _norm_campo

                def _linha_casa_filtros(_ln: Dict[str, Any]) -> bool:
                    for _campo, _alvo in _filtros_p.items():
                        if not _alvo:
                            continue
                        _valor = _norm_campo(str(_ln.get(_campo) or ""))
                        if _valor and _valor != _norm_campo(_alvo):
                            return False
                    return True

                _preencheu = False
                for _ln in st.session_state.planilha_linhas:
                    if not _linha_casa_filtros(_ln):
                        continue
                    if _preenche_turno and not str(_ln.get("turno") or "").strip():
                        _ln["turno"] = _filtros_p["turno"]
                        _preencheu = True
                    if _preenche_tri and not str(_ln.get("trimestre") or "").strip():
                        _ln["trimestre"] = _filtros_p["trimestre"]
                        _preencheu = True
                if _preencheu:
                    st.caption(
                        "Turno/Trimestre do filtro preenchidos apenas nas linhas "
                        "que batem com os filtros de Escola/Turma. Para turnos "
                        "diferentes por turma, deixe o Turno vazio e use o "
                        "mapeamento do importador do Google Forms."
                    )
                    st.rerun()

        if st.session_state.get("fonte", "") == "google_drive" and tipo == "notas":
            st.markdown("**Arquivo no Google Drive**")
            st.caption("Cole o link do arquivo usado como origem das notas.")
            link_drive_notas = st.text_input(
                "Link do arquivo no Drive:",
                key="link_input_drive_notas",
                placeholder="https://drive.google.com/file/d/...",
            )
            st.session_state["link_url"] = link_drive_notas

        if tipo == "sequencia":
            filtros_ativos = [k for k in ["escola", "turno", "turma"] if st.session_state.get(k, "")]
            if not filtros_ativos:
                st.info("Publicara em **todas as escolas, turnos e turmas** disponiveis.")
            else:
                st.caption(f"Filtros ativos: {', '.join(filtros_ativos)}")

            if st.session_state.get("fonte", "") == "google_drive":
                st.markdown("**Links do Google Drive por Turma**")
                st.caption(
                    "Digite o nome da turma exatamente como aparece no portal (ex.: '6º Ano' ou '6º Ano 1') "
                    "e cole o link do arquivo compartilhado. Deixe vazio o que nao for lancar."
                )

                if "seq_drive_links" not in st.session_state:
                    st.session_state.seq_drive_links = [
                        {"ano": "6º Ano", "link": ""},
                        {"ano": "7º Ano", "link": ""},
                        {"ano": "8º Ano", "link": ""},
                        {"ano": "9º Ano", "link": ""},
                    ]

                for i, entry in enumerate(st.session_state.seq_drive_links):
                    cols = st.columns([1.4, 4, 0.5])
                    with cols[0]:
                        entry["ano"] = st.text_input(
                            "Ano/Turma",
                            value=entry["ano"],
                            key=f"seq_drive_ano_{i}",
                            placeholder="Ex.: 6º Ano 1",
                            label_visibility="collapsed",
                        )
                    with cols[1]:
                        entry["link"] = st.text_input(
                            "Link do Drive",
                            value=entry["link"],
                            key=f"seq_drive_link_{i}",
                            placeholder="https://drive.google.com/file/d/...",
                            label_visibility="collapsed",
                        )
                    with cols[2]:
                        if len(st.session_state.seq_drive_links) > 1:
                            if st.button("✕", key=f"seq_drive_rm_{i}", help="Remover"):
                                st.session_state.seq_drive_links.pop(i)
                                st.rerun()

                if st.button("+ Adicionar turma", key="add_drive_link", use_container_width=True):
                    st.session_state.seq_drive_links.append({"ano": "", "link": ""})
                    st.rerun()

                st.markdown("**Campos comuns a todos os anos:**")
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    st.text_input(
                        "Título do Documento",
                        key="seq_titulo_documento",
                        placeholder="Sequência - Matemática",
                    )
                    st.text_input(
                        "Período início (dd/mm/aaaa)",
                        key="seq_periodo_inicio",
                        placeholder="01/03/2026",
                    )
                with col_g2:
                    st.number_input(
                        "Nº de aulas",
                        min_value=1, max_value=40, value=4,
                        key="seq_n_aulas",
                    )
                    st.text_input(
                        "Período fim (dd/mm/aaaa)",
                        key="seq_periodo_fim",
                        placeholder="31/03/2026",
                    )

        if tipo == "chamada":
            st.markdown("---")
            st.markdown("**Chamada diaria por foto do diario de classe**")
            st.caption(
                "Envie a foto da chamada do dia. A IA le a grade (aluno x dia), o bot "
                "compara com o que ja esta lancado no SGE e NAO repreenche dias/duplicados."
            )
            foto_chamada = st.file_uploader(
                "Foto do diario de classe (chamada do dia)",
                type=["jpg", "jpeg", "png", "webp"],
                key="chamada_foto_upload",
            )
            if foto_chamada:
                tmp_dir = Path(tempfile.gettempdir()) / "sge_bot_uploads"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / foto_chamada.name
                with open(tmp_path, "wb") as f:
                    f.write(foto_chamada.getbuffer())
                st.session_state["chamada_foto_path"] = str(tmp_path)
                st.success(f"Foto salva: {foto_chamada.name}")
            else:
                st.session_state.pop("chamada_foto_path", None)

            col_c1, col_c2 = st.columns(2)
            with col_c1:
                chamada_dia = st.text_input(
                    "Dia da chamada (dd/mm/aaaa)",
                    key="chamada_dia_input",
                    placeholder="05/08/2026",
                    help="Data do dia da chamada no diario. Usada para abrir o dia no SGE.",
                )
                st.session_state.chamada_dia = chamada_dia
            with col_c2:
                chamada_disciplina = st.text_input(
                    "Disciplina (opcional, ex: ENSINO RELIGIOSO)",
                    key="chamada_disciplina_input",
                    help="Se deixar vazio, o bot usa a disciplina da pagina aberta no SGE.",
                )
                st.session_state.chamada_disciplina = chamada_disciplina

        if tipo == "notas":
            st.markdown("---")
            st.markdown("**Imagem / Foto (extrair notas com IA)**")
            img = st.file_uploader(
                "Envie foto ou print com notas dos alunos",
                type=["jpg", "jpeg", "png", "webp"],
                key="imagem_upload",
            )
            if img:
                tmp_dir = Path(tempfile.gettempdir()) / "sge_bot_uploads"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / img.name
                with open(tmp_path, "wb") as f:
                    f.write(img.getbuffer())
                st.session_state["imagem_path"] = str(tmp_path)
                st.success(f"Imagem salva: {img.name}")
                st.caption("Informe abaixo o **nome da atividade** e a **data** para o bot localizar no portal:")
                col_ia1, col_ia2 = st.columns(2)
                with col_ia1:
                    avaliacao_nome = st.text_input(
                        "Nome da atividade (ex: Prova 1, Trabalho 2)",
                        key="avaliacao_input",
                        help="Nome da atividade que aparece no portal. O bot identifica por este nome."
                    )
                    st.session_state.avaliacao_nome = avaliacao_nome
                with col_ia2:
                    avaliacao_data = st.text_input(
                        "Data da atividade (dd/mm/aaaa)",
                        key="avaliacao_data_input",
                        help="Data de realização para o bot localizar/validar no portal."
                    )
                    st.session_state.avaliacao_data = avaliacao_data
            else:
                st.session_state.pop("imagem_path", None)
                st.session_state.pop("avaliacao_nome", None)
                st.session_state.pop("avaliacao_data", None)

            if st.session_state.get("fonte", "notion") != "imagem":
                st.markdown("---")
                st.markdown("**Atalho: Lançar apenas 1 avaliação** (reduz tempo drasticamente)")
                col_a1, col_a2 = st.columns(2)
                with col_a1:
                    avaliacao_nome = st.text_input(
                        "Nome da avaliação (ex: Prova 1, Trabalho 2)",
                        key="avaliacao_atalho_input",
                        help="Se preenchido, o bot ignora as outras avaliações e lança apenas esta."
                    )
                    st.session_state.avaliacao_nome = avaliacao_nome
                with col_a2:
                    avaliacao_data = st.text_input(
                        "Data da avaliação (dd/mm/aaaa)",
                        key="avaliacao_data_atalho_input",
                        help="Data de realização para validação. Se vazio, usa a data do Notion."
                    )
                    st.session_state.avaliacao_data = avaliacao_data

            lote = st.checkbox(
                "Modo Lote: processar todas as escolas, turmas e trimestres",
                value=False,
                key="lote_check",
                help="Quando ativo, ignora os filtros acima e processa tudo que estiver no Notion (6o ao 9o ano)."
            )
            st.session_state.lote = lote

            if not lote and st.session_state.get("fonte", "notion") == "notion":
                st.markdown("---")
                st.markdown("**Lancar uma turma por vez** (quando ha 2 turmas do mesmo ano)")
                st.caption(
                    "Ex.: filtrando '6o Ano' o bot encontra '6º Ano 1' e '6º Ano 2'. "
                    "Escolha uma turma abaixo para lancar apenas ela nesta execucao."
                )
                _filtros_atual = (
                    st.session_state.get("escola", ""),
                    st.session_state.get("turno", ""),
                    st.session_state.get("trimestre", ""),
                )
                if st.session_state.get("turmas_notion_filtros") != _filtros_atual:
                    st.session_state.pop("turmas_notion", None)
                    st.session_state.turma_especifica = ""
                if st.button("Carregar turmas disponíveis", key="btn_listar_turmas", use_container_width=True):
                    with st.spinner("Consultando o Notion..."):
                        turmas = _listar_turmas_notion()
                    st.session_state["turmas_notion"] = turmas
                    st.session_state["turmas_notion_filtros"] = _filtros_atual
                turmas = st.session_state.get("turmas_notion", [])
                if turmas:
                    turma_especifica = st.selectbox(
                        "Turma para lancar agora (uma por vez)",
                        options=[""] + list(turmas),
                        format_func=lambda x: x or "Todas as turmas",
                        key="turma_especifica_select",
                    )
                    st.session_state.turma_especifica = turma_especifica
                    if turma_especifica:
                        st.success(f"Lancando apenas: **{turma_especifica}**")
                else:
                    st.session_state.turma_especifica = ""

    ia_opcao = st.radio(
        "Assistencia IA:",
        options=["nenhuma", "assistida", "aprendizado"],
        format_func=lambda x: {"nenhuma": "Nao usar", "assistida": "Sim - Assistida", "aprendizado": "Sim - Aprendizado"}.get(x, x),
        horizontal=True,
        key="ia_radio",
    )

    if st.session_state.get("portal_selecionado", "SGE") in ("Novo Portal", "Professor Online") and ia_opcao != "aprendizado":
        if st.session_state.get("portal_selecionado", "SGE") == "Novo Portal":
            st.info(
                "Portal **Novo Portal**: o modo aprendizado sera ativado automaticamente ao executar. "
                "Nao e necessario marcar nada aqui."
            )
        else:
            st.warning(
                "Portal **Professor Online**: acoes ainda nao suportadas (ex.: criacao/upload de "
                "planejamento) usam o **Modo Aprendizado** para a IA gravar o fluxo e incluir a "
                "nova funcionalidade ao sistema. Marque **'Sim - Aprendizado'** quando for ensinar "
                "um fluxo novo."
            )

    st.markdown("---")
    st.markdown("**Opcoes de execucao**")

    with st.expander("Janela do navegador"):
        st.markdown(
            "Quando ativado, o navegador roda em **segundo plano** sem mostrar nada na tela. "
            "Deixe marcado para uso normal.\n\n"
            "**Desmarque** quando o portal pedir **captcha** no login: o navegador abre visivel, "
            "voce completa o captcha manualmente e o bot continua sozinho. Tambem use para "
            "**ver** o navegador sendo controlado passo a passo."
        )
        headless_mode = st.checkbox("Rodar em segundo plano (recomendado)", value=st.session_state.headless_mode, key="headless_check")
        st.session_state.headless_mode = headless_mode

    with st.expander("Casamento de nomes dos alunos"):
        st.markdown(
            "Quando ativado, o bot **busca o aluno pelo primeiro nome** no portal. "
            "Se dois alunos tiverem o **mesmo primeiro nome**, ele usa o **segundo nome** "
            "para diferenciar; se ainda houver duvida, o **terceiro nome**.\n\n"
            "Ideal quando a IA le o primeiro nome corretamente mas o sobrenome sai "
            "corrompido da planilha.\n\n"
            "**Desmarque** apenas se quiser exigir o casamento pelo nome completo."
        )
        primeiro_nome_match = st.checkbox(
            "Buscar por primeiro nome (desambiguar com 2o/3o nome)",
            value=st.session_state.get("primeiro_nome_match", True),
            key="primeiro_nome_match_check",
        )
        st.session_state.primeiro_nome_match = primeiro_nome_match

    with st.expander("Correção automática de erros (IA)"):
        st.markdown(
            "Quando ativado, se acontecer algum erro (ex: data no formato errado, link invalido, "
            "demora para carregar), o programa **tenta corrigir sozinho** usando inteligencia "
            "artificial local.\n\n"
            "Se conseguir corrigir, uma mensagem explica o que foi ajustado e voce pode clicar "
            "em **Tentar novamente**.\n\n"
            "Requer o Ollama instalado (ja configurado neste computador)."
        )
        autofix_enabled = st.checkbox("Corrigir erros automaticamente", value=st.session_state.get("autofix_enabled", False), key="autofix_check")
        st.session_state.autofix_enabled = autofix_enabled

    with st.expander("Revisao final IA (conferir apos salvar)"):
        st.markdown(
            "Quando ativado, apos lancar as notas o programa **reabre cada avaliacao e confere** "
            "se a nota gravou corretamente (re-leitura deterministica).\n\n"
            "Se uma nota divergir ou o campo nao for encontrado, a **IA analisa a tela** "
            "e decide se esta correta. Se nao estiver, o programa **tenta corrigir regravando**;\n"
            "se ainda assim falhar, marca como **falha** (nao conta como lancada).\n\n"
            "Usa a IA apenas em caso de duvida (maximo 3 consultas por avaliacao)."
        )
        revisar_apos = st.checkbox("Conferir notas apos salvar (recomendado)", value=st.session_state.get("revisar_apos", True), key="revisar_apos_check")
        st.session_state.revisar_apos = revisar_apos

    with st.expander("Modo escuro"):
        st.markdown("Alterna entre tema claro e escuro da interface.")
        dark_mode = st.checkbox("Usar tema escuro", value=st.session_state.dark_mode, key="dark_check")
        st.session_state.dark_mode = dark_mode
    salvar = st.button("Salvar Configuracao")

    if salvar:
        path = salvar_config()
        st.success(f"Configuracao salva em {path}")

    limpar = st.button("Limpar Logs")
    if limpar:
        st.session_state.logs = []
        st.session_state.resultado = None


# === TEMA DINAMICO ===
if st.session_state.get("dark_mode", False):
    st.markdown(_DARK_THEME_CSS, unsafe_allow_html=True)

# === AREA PRINCIPAL ===
st.markdown('<div class="main-header">Bot do Professor - Automacao de Notas</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub-header">Lance notas do Notion/planilhas no portal do professor sem precisar de terminal</div>',
    unsafe_allow_html=True,
)

# === LEMBRETE PERMANENTE: SEMPRE REVISAR AS NOTAS ===
_acerto_pct, _acerto_total = _taxa_acerto_global()
if _acerto_pct is not None:
    st.warning(
        f"**Importante:** revise sempre as notas lancadas ao final — o bot pode errar! "
        f"Taxa de acerto dos seus ultimos lancamentos: **{_acerto_pct}%** "
        f"({_acerto_total} nota(s) processada(s)). "
        f"Confira no portal e na aba **Pendencias**.",
        icon="\U0001F440",
    )
else:
    st.warning(
        "**Importante:** revise sempre as notas lancadas ao final — o bot pode errar. "
        "Confira cada nota no portal depois do envio e use a aba **Pendencias**.",
        icon="\U0001F440",
    )

# === ABAS DO PAINEL ===
tab_plan, tab_rev, tab_logs = st.tabs(["Planilha", "Pendências", "Logs"])

with tab_plan:
    if st.session_state.get("fonte", "notion") == "planilha":
        plan_arquivo = st.file_uploader(
            "Carregar .xlsx ou .csv para preencher a tabela",
            type=["xlsx", "xls", "csv"],
            key="planilha_upload",
        )
        if plan_arquivo:
            nome_arq = plan_arquivo.name
            if nome_arq != st.session_state.get("planilha_ultimo_arquivo", ""):
                from leitor_planilhas import carregar_notas, registros_para_linhas
                tmp_dir = Path(tempfile.gettempdir()) / "sge_bot_uploads"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / f"_painel_{nome_arq}"
                with open(tmp_path, "wb") as f:
                    f.write(plan_arquivo.getbuffer())
                ext = nome_arq.rsplit(".", 1)[-1].lower()
                fonte_arq = "excel" if ext in ("xlsx", "xls") else "csv"
                try:
                    registros_arq = carregar_notas(fonte_arq, str(tmp_path), logger=log)
                except Exception as exc:
                    st.error(f"Erro ao ler o arquivo: {exc}")
                    registros_arq = []
                linhas = registros_para_linhas(registros_arq)
                _aplicar_status_salvo(linhas)
                st.session_state.planilha_linhas = linhas
                st.session_state.planilha_ultimo_arquivo = nome_arq
                if registros_arq:
                    st.success(f"{len(registros_arq)} nota(s) carregada(s). Edite se precisar e execute.")
                else:
                    st.warning("Nenhuma nota encontrada no arquivo. Confira o cabecalho (Nome/Aluno + colunas de nota).")
        else:
            if st.session_state.get("planilha_ultimo_arquivo"):
                st.session_state.planilha_ultimo_arquivo = ""

        if "planilha_linhas" not in st.session_state:
            st.session_state.planilha_linhas = []

        st.markdown("**Notas (edite as celulas; use **`+`** na ultima linha para adicionar)**")
        plan_editor = st.data_editor(
            st.session_state.planilha_linhas,
            num_rows="dynamic",
            key="planilha_editor",
            use_container_width=True,
            column_config={
                "escola": st.column_config.TextColumn("Escola"),
                "turno": st.column_config.TextColumn("Turno"),
                "turma": st.column_config.TextColumn("Turma"),
                "trimestre": st.column_config.TextColumn("Trimestre"),
                "aluno": st.column_config.TextColumn("Aluno"),
                "atividade": st.column_config.TextColumn("Atividade"),
                "nota": st.column_config.NumberColumn("Nota", min_value=0.0, max_value=10.0, format="%.2f"),
                "data_realizacao": st.column_config.TextColumn("Data realizacao"),
                "status": st.column_config.SelectboxColumn("Status", options=["", "Lancada", "Falha"]),
            },
        )
        st.session_state.planilha_linhas = plan_editor
        st.caption(f"{len(st.session_state.planilha_linhas)} linha(s) na tabela. Escola/Turno/Turma/Trimestre vazios usam os filtros acima.")

        col_p1, col_p2 = st.columns(2)
        with col_p1:
            if st.button("Limpar tabela", key="planilha_limpar", use_container_width=True):
                st.session_state.planilha_linhas = []
                st.rerun()
        with col_p2:
            if st.button("Adicionar linha vazia", key="planilha_add_linha", use_container_width=True):
                st.session_state.planilha_linhas = st.session_state.planilha_linhas + [{
                    "escola": "", "turno": "", "turma": "", "trimestre": "",
                    "aluno": "", "atividade": "", "nota": "", "data_realizacao": "", "status": "",
                }]
                st.rerun()
    else:
        st.info(
            "Selecione a origem **'Planilha no painel (editar aqui)'** na barra lateral "
            "(Origem dos Dados) para editar as notas nesta aba."
        )

# === DIALOGO DE CONFIRMACAO DE ENVIO ===
@st.dialog("Confirmar envio ao portal")
def _confirmar_envio_dialog(resumo: tuple) -> None:
    st.markdown("Voce esta enviando **de verdade** para o portal. Confira o resumo:")
    for _linha in resumo:
        if _linha.endswith(": ``") or _linha.endswith(": ''"):
            continue
        st.markdown(f"- {_linha}")
    st.caption("Depois do envio, revise as notas gravadas no portal (aba Pendencias e o proprio portal).")
    _c1, _c2 = st.columns(2)
    if _c1.button("Sim, enviar agora", type="primary", use_container_width=True):
        st.session_state.confirmar_envio = True
        st.session_state.executar_agora = True
        st.rerun()
    if _c2.button("Cancelar", use_container_width=True):
        st.rerun()

# === EXECUCAO (stepper + botao + gate) ===
st.markdown('<div class="main-header">Execucao</div>', unsafe_allow_html=True)

dry_run = False

stepper_fases = ["Carregar", "Validar", "Lancar", "Revisar"]
if st.session_state.get("executando"):
    stepper_idx = 2
elif st.session_state.get("resultado"):
    stepper_idx = 3
else:
    stepper_idx = -1
stepper_cols = st.columns(len(stepper_fases))
for _i, _fase in enumerate(stepper_fases):
    with stepper_cols[_i]:
        if _i < stepper_idx:
            st.markdown(f"<div style='text-align:center;background:#e6f4ea;border:1px solid #34a853;border-radius:8px;padding:4px 6px;font-size:0.8rem;color:#137333;'>✔ {_fase}</div>", unsafe_allow_html=True)
        elif _i == stepper_idx:
            st.markdown(f"<div style='text-align:center;background:#e8f0fe;border:1px solid #1f6feb;border-radius:8px;padding:4px 6px;font-size:0.8rem;color:#174ea6;'>→ {_fase}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='text-align:center;background:#f1f3f4;border:1px solid #dadce0;border-radius:8px;padding:4px 6px;font-size:0.8rem;color:#5f6368;'>{_fase}</div>", unsafe_allow_html=True)

col_btn1, col_btn2 = st.columns([3, 1])

with col_btn1:
    portal_ativo = st.session_state.get("portal_selecionado", "SGE")
    if portal_ativo == "Professor Online":
        pode_executar = bool(
            st.session_state.get("po_cpf", "")
            and st.session_state.get("po_senha", "")
        )
    elif portal_ativo == "Novo Portal":
        pode_executar = bool(
            st.session_state.get("np_url", "")
            and st.session_state.get("np_cpf", "")
            and st.session_state.get("np_senha", "")
        )
    elif portal_ativo == "Auto (detecta pela escola)":
        pode_executar = bool(
            (st.session_state.get("sge_cpf", "") and st.session_state.get("sge_senha", ""))
            or (st.session_state.get("po_cpf", "") and st.session_state.get("po_senha", ""))
        )
    else:
        pode_executar = bool(
            st.session_state.get("sge_cpf", "")
            and st.session_state.get("sge_senha", "")
        )
    if not pode_executar:
        st.warning("Preencha URL, CPF e Senha do portal na barra lateral")

    _resumo_envio = [
        f"**Portal:** {portal_ativo}",
        f"**Escola:** {st.session_state.get('escola', '') or 'todas'}",
        f"**Turma:** {st.session_state.get('turma', '') or 'todas'}",
        f"**Turno:** {st.session_state.get('turno', '') or 'todos'}",
        f"**Trimestre:** {st.session_state.get('trimestre', '') or 'todos'}",
    ]
    _df_atual = st.session_state.get("df")
    if _df_atual is not None and hasattr(_df_atual, "empty") and not _df_atual.empty:
        _resumo_envio.append(f"**Registros na tabela:** {len(_df_atual)}")

    executar_btn = st.button(
        "ENVIAR NOTAS AO PORTAL",
        type="primary",
        disabled=(not pode_executar or st.session_state.executando),
        use_container_width=True,
    )
    if executar_btn:
        _confirmar_envio_dialog(tuple(_resumo_envio))

with col_btn2:
    ajuda_btn = st.button("Ajuda / Tutorial", use_container_width=True)

if ajuda_btn:
    st.markdown("### Como usar")
    st.markdown("""
    1. **Portal**: Escolha SGE, Professor Online ou Novo Portal
    2. **Novo Portal**: informe URL, CPF e senha. Ao executar, o modo aprendizado e ativado automaticamente — o navegador abre visivel, voce faz o acesso manualmente e a IA local aprende.
    3. **API Keys**: Cole o token do Notion e o ID da pagina raiz (veja as instrucoes na propria secao)
    4. **Origem**: Escolha de onde ler os dados (Notion, Excel, CSV, Google)
    5. **Filtros**: Opcionais - para filtrar por escola/turma
    6. **Modo Lote**: Marque para processar todas as escolas, turmas e trimestres automaticamente
    7. **Executar**: Clique no botao, confira o resumo e confirme. Depois acompanhe os logs
    8. **Importante**: Sempre revise as notas lancadas no portal ao final
    """)

# === EXECUCAO ===
_rev_aplicar = st.session_state.pop("revisao_aplicar", False)
_rev_retentar = st.session_state.pop("revisao_retentar_ausentes", False)
if st.session_state.pop("executar_agora", False) or st.session_state.pop("autofix_trigger", False) or _rev_aplicar or _rev_retentar:
    st.session_state.executando = True
    st.session_state.logs = []
    st.session_state.pop("imagem_descartes", None)
    st.session_state.log_file = _start_log_file()
    st.session_state.resultado = None

    log("Iniciando execucao...")

    # === RESOLVE PORTAL (Auto: detecta pela escola registrada) ===
    portal_escolhido = st.session_state.get("portal_selecionado", "SGE")
    portal_resolvido = portal_escolhido
    if portal_escolhido == "Auto (detecta pela escola)":
        try:
            from bot.core.escola_registry import portal_da_escola
            escola_filtro = st.session_state.get("escola", "")
            auto = portal_da_escola(escola_filtro) if escola_filtro else None
            portal_resolvido = "Professor Online" if auto == "professor_online" else "SGE"
        except Exception as exc:
            portal_resolvido = "SGE"
            log(f"[AUTO] Aviso: falha ao consultar registro de escolas: {exc}")
        log(f"[AUTO] Escola '{st.session_state.get('escola', '') or '(vazia)'}' -> portal {portal_resolvido}")
        if portal_resolvido == "Professor Online":
            st.info(f"**Auto**: escola '{st.session_state.get('escola', '')}' reconhecida do **Professor Online**. Executando neste portal.")
        else:
            st.info("**Auto**: escola nao registrada no Professor Online. Executando no **SGE**.")
    st.session_state["portal_resolvido"] = portal_resolvido

    # === BARRA DE PROGRESSO ===
    status = st.status("Iniciando...", expanded=True)
    progress_bar = st.progress(0, text="Aguardando...")
    status_text = st.empty()

    def log_progress(msg):
        log(msg)
        try:
            m = re.search(r'\[(\d+)/(\d+)\]', msg)
            if m:
                cur, tot = int(m.group(1)), int(m.group(2))
                pct = cur / tot
                progress_bar.progress(pct, text=f"{cur}/{tot} blocos")
                status.update(label=f"Processando bloco {cur}/{tot}")
            elif "Concluido" in msg or "concluido" in msg.lower():
                progress_bar.progress(1.0, text="Concluido!")
                status.update(label="Concluido!", state="complete")
            elif "ESTRUTURA-CHANGED" in msg:
                status.update(label="Estrutura do SGE mudou — nada gravado", state="error")
            elif "ERRO" in msg or "erro" in msg.lower():
                status.update(label=msg[:80], state="error")
            elif "Iniciando" in msg or "Carregando" in msg or "Login" in msg or "login" in msg:
                progress_bar.progress(0.05, text=msg[:80])
            status_text.text(msg)
        except Exception:
            pass

    def _handle_exec_error(exc: Exception, lp: callable):
        import traceback as _tb
        tb_str = "".join(_tb.format_exception(type(exc), exc, exc.__traceback__))
        lp(f"ERRO: {exc}")
        print(tb_str)

        autofix_on = st.session_state.get("autofix_enabled", False)
        attempts = st.session_state.get("autofix_attempts", 0)
        msg_exc = str(exc)
        erro_verificacao = "Nenhuma nota foi preenchida no SGE" in msg_exc
        erro_login = any(k in msg_exc for k in ("Falha no login", "Credencial invalida", "SGE_CPF invalido"))
        if autofix_on and attempts < 3 and not msg_exc.startswith("st.") and not erro_verificacao and not erro_login:
            ctx = {
                "escola": st.session_state.get("escola", ""),
                "turno": st.session_state.get("turno", ""),
                "turma": st.session_state.get("turma", ""),
                "trimestre": st.session_state.get("trimestre", ""),
                "tipo": st.session_state.get("tipo", ""),
                "fonte": st.session_state.get("fonte", ""),
                "seq_titulo_documento": st.session_state.get("seq_titulo_documento", ""),
                "seq_periodo_inicio": st.session_state.get("seq_periodo_inicio", ""),
                "seq_periodo_fim": st.session_state.get("seq_periodo_fim", ""),
                "seq_n_aulas": st.session_state.get("seq_n_aulas", 4),
                "link_url": st.session_state.get("link_url", ""),
                "headless_mode": st.session_state.get("headless_mode", True),
            }
            result = attempt_autofix(str(exc), tb_str, ctx, logger=lp, attempt=attempts)
            if result and result.get("fixable"):
                applied = apply_fix(result, st.session_state)
                if applied:
                    lp(f"Autofix aplicado: {applied}")
                    st.session_state.autofix_message = result.get("explanation", "Erro corrigido automaticamente.")
                    st.session_state.autofix_attempts = attempts + 1
                    st.session_state.autofix_trigger = True
                    st.rerun()
                else:
                    lp("Autofix: correcao sugerida nao pode ser aplicada.")
            else:
                motivo = result.get("explanation", "Erro nao corrigivel automaticamente.") if result else "Sem resposta da IA."
                lp(f"Autofix: {motivo}")

    try:
        # Prepara variaveis de ambiente
        os.environ["HEADLESS"] = "1" if st.session_state.get("headless_mode", True) else "0"
        os.environ["SGE_REVISAR_APOS"] = "1" if st.session_state.get("revisar_apos", True) else "0"
        if st.session_state.get("portal_resolvido", "SGE") == "Professor Online":
            os.environ["PO_BASE_URL"] = st.session_state.get("po_url", "")
            os.environ["PO_CPF"] = st.session_state.get("po_cpf", "")
            os.environ["PO_SENHA"] = st.session_state.get("po_senha", "")
        elif st.session_state.get("portal_resolvido", "SGE") == "Novo Portal":
            os.environ["NP_URL"] = st.session_state.get("np_url", "")
            os.environ["NP_CPF"] = st.session_state.get("np_cpf", "")
            os.environ["NP_SENHA"] = st.session_state.get("np_senha", "")
        else:
            os.environ["SGE_LOGIN_URL"] = st.session_state.sge_url
            os.environ["SGE_CPF"] = st.session_state.sge_cpf
            os.environ["SGE_SENHA"] = st.session_state.sge_senha

        if st.session_state.notion_token:
            os.environ["NOTION_TOKEN"] = st.session_state.notion_token
        if st.session_state.root_page_id:
            os.environ["ROOT_PAGE_ID"] = st.session_state.root_page_id
        if st.session_state.gemini_key:
            os.environ["GEMINI_API_KEY"] = st.session_state.gemini_key

        ai_provider = st.session_state.get("ai_provider", "local")
        os.environ["AI_PROVIDER"] = ai_provider

        # Configura API key e modelo conforme provider escolhido
        if ai_provider == "gemini":
            gemini_key = st.session_state.get("gemini_key", "")
            if gemini_key:
                os.environ["GEMINI_API_KEY"] = gemini_key
            gemini_model = st.session_state.get("gemini_model", "gemini-2.5-flash")
            os.environ["AI_MODEL"] = gemini_model
        elif ai_provider == "openai":
            openai_key = st.session_state.get("openai_key", "")
            if openai_key:
                os.environ["OPENAI_API_KEY"] = openai_key
            openai_model = st.session_state.get("openai_model", "gpt-4o")
            os.environ["OPENAI_MODEL"] = openai_model
        elif ai_provider == "anthropic":
            anthropic_key = st.session_state.get("anthropic_key", "")
            if anthropic_key:
                os.environ["ANTHROPIC_API_KEY"] = anthropic_key
            anthropic_model = st.session_state.get("anthropic_model", "claude-sonnet-4-20250514")
            os.environ["ANTHROPIC_MODEL"] = anthropic_model
        elif ai_provider == "local":
            ollama_model = st.session_state.get("ollama_model", "llama3.2-vision")
            os.environ["OLLAMA_MODEL"] = ollama_model

        if ia_opcao != "nenhuma":
            os.environ["AI_ASSIST"] = "1"
            if ia_opcao == "aprendizado":
                os.environ["AI_LEARN_MODE"] = "1"

        # Verifica IA local (Ollama) antes de iniciar o navegador
        if ai_provider == "local" and ia_opcao not in ("nenhuma", ""):
            log_progress("Verificando Ollama (IA local)...")
            log_progress("Nota: download pode levar varios minutos na primeira execucao.")
            try:
                from ai_assist import ensure_ollama
                if not ensure_ollama(logger=log_progress):
                    log_progress("Aviso: Ollama nao disponivel. IA assistida sera limitada.")
            except Exception as exc:
                log_progress(f"Aviso: falha ao configurar Ollama: {exc}")
        elif ai_provider in ("gemini", "openai", "anthropic") and ia_opcao not in ("nenhuma", ""):
            # Verificar se a API key foi fornecida
            key_available = False
            if ai_provider == "gemini" and st.session_state.get("gemini_key"):
                key_available = True
            elif ai_provider == "openai" and st.session_state.get("openai_key"):
                key_available = True
            elif ai_provider == "anthropic" and st.session_state.get("anthropic_key"):
                key_available = True
            if not key_available:
                log_progress(f"Aviso: API key para {ai_provider.upper()} nao fornecida. IA sera desabilitada.")

        # ===== NOVO PORTAL: sessao de aprendizado do acesso =====
        if st.session_state.get("portal_resolvido", "SGE") == "Novo Portal":
            if ia_opcao != "aprendizado":
                ia_opcao = "aprendizado"
                os.environ["AI_LEARN_MODE"] = "1"
                log_progress("[HIBRIDO] Modo aprendizado ativado automaticamente para portal novo.")
                st.info("Portal novo detectado: **modo aprendizado ativado automaticamente**. O navegador abrirá visível e você fará o acesso manualmente enquanto a IA local grava e aprende.")

            log_progress("Iniciando aprendizado do novo portal...")
            log_progress("Navegador visivel: faca o acesso manualmente (login, turma, grade, salvar).")
            log_progress("Cada tela sera gravada. Feche o navegador quando terminar.")
            from aprender_novo_portal import executar_aprendizado

            resultado = executar_aprendizado(
                url=st.session_state.get("np_url", ""),
                cpf=st.session_state.get("np_cpf", ""),
                senha=st.session_state.get("np_senha", ""),
                portal_name=st.session_state.get("portal_nome_value", ""),
                outdir=os.environ.get("NP_OUTDIR", "artifacts/novo-portal"),
                max_minutes=60,
                logger=log_progress,
            )
            st.session_state.resultado = resultado
            if resultado.get("ok"):
                log_progress(f"[APRENDIZADO-OK] Portal '{resultado['portal']}' aprendido com {resultado.get('steps', 0)} passo(s).")
                log_progress(f"Plano salvo em: {resultado.get('plan_path', '')}")
            else:
                log_progress(f"[APRENDIZADO] Concluido com {resultado.get('steps', 0)} passo(s) gravados, mas o plano nao foi gerado.")
            st.stop()

        # Determina tipo de execucao
        if st.session_state.tipo == "notas":
            if st.session_state.fonte == "notion" and st.session_state.get("portal_resolvido", "SGE") != "Professor Online":
                log_progress("Carregando dados do Notion...")
                import importlib
                import lancar_notas_sge as _sge_mod
                importlib.reload(_sge_mod)
                from lancar_notas_sge import executar_lancamento

                # Modo lote: ignora filtros, processa tudo
                if st.session_state.get("lote", False):
                    log_progress("[LOTE] Modo lote ativo: processando todas as escolas, turmas e trimestres...")
                    filtro = None
                else:
                    filtro = {}
                    if st.session_state.escola:
                        filtro["escola"] = st.session_state.escola
                    if st.session_state.turno:
                        filtro["turno"] = st.session_state.turno
                    if st.session_state.turma:
                        filtro["turma"] = st.session_state.turma
                    turma_especifica = st.session_state.get("turma_especifica", "")
                    if turma_especifica:
                        filtro["turma"] = turma_especifica
                        log_progress(f"[TURMA-ESPECIFICA] Lancando apenas a turma: '{turma_especifica}'")
                    if st.session_state.trimestre:
                        filtro["trimestre"] = st.session_state.trimestre
                    # Atalho: filtrar por avaliacao especifica
                    avaliacao_nome = st.session_state.get("avaliacao_nome", "").strip()
                    avaliacao_data = st.session_state.get("avaliacao_data", "").strip()
                    if avaliacao_nome:
                        filtro["atividade"] = avaliacao_nome
                        log_progress(f"[ATALHO] Filtrando apenas avaliacao: '{avaliacao_nome}'")
                    if avaliacao_data:
                        filtro["data_realizacao"] = avaliacao_data
                        log_progress(f"[ATALHO] Data da avaliacao: {avaliacao_data}")
                    if not filtro:
                        filtro = None

                resultado = executar_lancamento(
                    filtro=filtro,
                    logger=log_progress,
                    dry_run=dry_run,
                    buscar_por_primeiro_nome=st.session_state.get("primeiro_nome_match", True),
                )
                st.session_state.resultado = resultado
                _alimentar_fila_revisao(resultado, log_progress)
                if resultado.get("estrutura_changed"):
                    st.session_state.estrutura_changed = resultado.get("estrutura_evidencia") or {}
                    st.session_state.revisao_fase = "estrutura_changed"
                log_progress(f"Concluido! Notas: {resultado['notas']}, Preenchidas: {resultado['notas_preenchidas']}, Ausentes: {resultado.get('ausentes', 0)}, Falhas: {resultado['falhas']}")

            else:
                registros = []
                fonte = st.session_state.fonte
                fonte_path = ""

                if fonte == "imagem":
                    fonte_path = st.session_state.get("imagem_path", "")
                    if not fonte_path:
                        log_progress("ERRO: Nenhuma imagem selecionada.")
                        log_progress("Selecione uma imagem na seção 'Filtros' antes de executar.")
                        st.session_state.resultado = {"blocos": 0, "notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                        st.stop()

                    log_progress("Extraindo notas da imagem com IA (reforcada)...")
                    try:
                        from ai_assist import extrair_notas_imagem
                        with open(fonte_path, "rb") as f:
                            image_bytes = f.read()

                        extraidas = extrair_notas_imagem(
                            image_bytes,
                            logger=log_progress,
                        )

                        if not extraidas:
                            log_progress("AVISO: IA nao conseguiu extrair notas da imagem.")
                            st.session_state.resultado = {"notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                            st.stop()

                        log_progress(f"Extraidas {len(extraidas)} notas da imagem.")
                        escola = st.session_state.get("escola", "")
                        turno = st.session_state.get("turno", "")
                        turma = st.session_state.get("turma", "")
                        trimestre = st.session_state.get("trimestre", "")
                        atividade = st.session_state.get("avaliacao_nome", "").strip()
                        data_realizacao = st.session_state.get("avaliacao_data", "").strip()
                        if not atividade:
                            log_progress("ERRO: Informe o nome da atividade (abaixo da imagem) para o bot identificar no portal.")
                            st.error("Informe o **nome da atividade** abaixo da imagem antes de executar.")
                            st.session_state.resultado = {"notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 1}
                            st.stop()
                        log_progress(f"[IMAGEM] Atividade: '{atividade}' | Data: {data_realizacao or 'nao informada'}")

                        from leitor_planilhas import RegistroNota

                        # === SALVAGUARDAS DA EXTRACAO POR IMAGEM ===
                        _nota_max = float(os.environ.get("NOTA_MAX_IMAGEM", "10"))
                        _descartes: list = []
                        _duplicados: list = []
                        _vistos: dict = {}
                        registros_img: list = []

                        def _chave_nome(nome: str) -> str:
                            import unicodedata
                            return re.sub(
                                r"\s+", " ",
                                unicodedata.normalize("NFD", (nome or "").strip().lower())
                                .encode("ascii", "ignore").decode(),
                            )

                        for item in extraidas:
                            aluno = str(item.get("aluno", "")).strip()
                            nota = str(item.get("nota", "")).strip().replace(",", ".")

                            if not aluno:
                                _descartes.append("(sem nome): leitura incompleta")
                                continue
                            if not nota:
                                _descartes.append(f"{aluno}: nota ilegivel/vazia")
                                continue

                            chave = _chave_nome(aluno)
                            if chave in _vistos:
                                _duplicados.append(f"{aluno} (lido 2x: '{_vistos[chave]}' e '{nota}'; usando '{_vistos[chave]}')")
                                continue
                            _vistos[chave] = nota

                            try:
                                num = float(nota)
                            except ValueError:
                                _descartes.append(f"{aluno}: nota ilegivel ('{nota}')")
                                continue
                            if num < 0 or num > _nota_max:
                                _descartes.append(
                                    f"{aluno}: nota suspeita ({nota}) fora da faixa 0-{_nota_max:g}"
                                    f" — provavel leitura errada da foto"
                                )
                                continue

                            registros_img.append(RegistroNota(
                                escola=escola, turno=turno, turma=turma,
                                trimestre=trimestre, aluno=aluno,
                                atividade=atividade, nota=nota,
                                data_realizacao=data_realizacao,
                            ))

                        for d in _duplicados:
                            log_progress(f"[IMAGEM] Nome duplicado na foto: {d}")
                        for d in _descartes:
                            log_progress(f"[IMAGEM] IGNORADO -> {d}. Corrija manualmente no portal ou envie foto melhor.")
                        if _duplicados or _descartes:
                            log_progress(
                                f"[IMAGEM] RESUMO: {len(_duplicados)} duplicado(s), "
                                f"{len(_descartes)} ignorado(s), {len(registros_img)} valido(s)."
                            )
                        st.session_state["imagem_descartes"] = _descartes
                        registros.extend(registros_img)

                        # Checagem de contagem contra a tabela carregada (se houver)
                        _df_ref = st.session_state.get("df")
                        if _df_ref is not None and hasattr(_df_ref, "empty") and not _df_ref.empty:
                            _esperados = len(_df_ref)
                            if len(extraidas) < _esperados:
                                log_progress(
                                    f"AVISO [IMAGEM]: li apenas {len(extraidas)} aluno(s) na foto, mas a tabela "
                                    f"tem {_esperados}. A foto pode ter cortado o final — confira se todos foram lancados."
                                )

                    except Exception as exc:
                        log_progress(f"ERRO ao processar imagem com IA: {exc}")
                        import traceback
                        traceback.print_exc()
                        st.session_state.resultado = {"notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 1}
                        st.stop()

                elif fonte == "planilha":
                    fonte_path = _PAINEL_SOURCE_PATH
                    from leitor_planilhas import linhas_para_registros, filtrar_linhas_por_filtros
                    linhas_fonte, _descartadas = filtrar_linhas_por_filtros(
                        st.session_state.get("planilha_linhas", []),
                        {
                            "escola": st.session_state.get("escola", ""),
                            "turno": st.session_state.get("turno", ""),
                            "turma": st.session_state.get("turma", ""),
                            "trimestre": st.session_state.get("trimestre", ""),
                        },
                        logger=log_progress,
                    )
                    registros = linhas_para_registros(
                        linhas_fonte,
                        defaults={
                            "escola": st.session_state.get("escola", ""),
                            "turno": st.session_state.get("turno", ""),
                            "turma": st.session_state.get("turma", ""),
                            "trimestre": st.session_state.get("trimestre", ""),
                        },
                        logger=log_progress,
                    )
                    if not registros:
                        st.session_state.resultado = {"blocos": 0, "notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                        if _descartadas:
                            log_progress(
                                f"ERRO: Todas as {_descartadas} linha(s) foram ignoradas "
                                "pelos filtros selecionados. Use na Escola/Turma os "
                                "valores exatos da tabela (ou deixe vazio)."
                            )
                            st.error(
                                f"Nenhuma linha bateu com os filtros ({_descartadas} "
                                "ignorada(s)). Os campos Escola/Turma agora mostram os "
                                "valores exatos da tabela — reavalie os filtros ou "
                                "deixe vazio para lancar tudo."
                            )
                        else:
                            log_progress("ERRO: Nenhuma linha valida na planilha do painel.")
                            st.error("Preencha a planilha no painel: ao menos aluno, atividade e nota (0-10).")
                        st.stop()
                    log_progress(f"{len(registros)} notas validas carregadas da planilha do painel.")

                else:
                    log_progress(f"Carregando dados de {fonte}...")
                    from leitor_planilhas import carregar_notas

                    if fonte in ("excel", "csv"):
                        fonte_path = st.session_state.get("arquivo_path", "")
                        if not fonte_path:
                            log_progress("ERRO: Nenhum arquivo selecionado.")
                            st.error("Selecione um arquivo primeiro.")
                            st.session_state.executando = False
                            st.stop()
                    elif fonte in ("google_sheets", "google_drive"):
                        fonte_path = st.session_state.get("link_url", "")
                        if not fonte_path:
                            log_progress("ERRO: Nenhum link informado.")
                            st.error("Informe o link do Google Sheets/Drive.")
                            st.session_state.executando = False
                            st.stop()

                    registros = carregar_notas(fonte, fonte_path, logger=log_progress)
                    log_progress(f"{len(registros)} notas carregadas.")

                if not registros:
                    log_progress("ERRO: Nenhum registro carregado.")
                    st.session_state.resultado = {"blocos": 0, "notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                    st.stop()

                portal_ativo = st.session_state.get("portal_resolvido", "SGE")

                if portal_ativo == "Professor Online" and fonte == "notion":
                    log_progress("ERRO: Portal Professor Online nao usa fonte Notion. Use Excel/CSV/Google.")
                    st.error("Portal Professor Online: use Excel, CSV ou Google Sheets como origem.")
                    st.session_state.resultado = {"blocos": 0, "notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                    st.stop()

                if portal_ativo == "Professor Online":
                    log_progress("Portal Professor Online selecionado. Iniciando lancamento...")
                    from lancar_professor_online import executar_lancamento as po_executar
                    resultado = po_executar(
                        fonte=fonte,
                        fonte_path=fonte_path,
                        filtro=None,
                        logger=log_progress,
                        dry_run=False,
                        cpf=st.session_state.get("po_cpf", ""),
                        senha=st.session_state.get("po_senha", ""),
                        base_url=st.session_state.get("po_url", ""),
                        registros=registros,
                    )
                    st.session_state.resultado = resultado
                    log_progress(
                        f"Concluido! Notas: {resultado['notas']}, "
                        f"Preenchidas: {resultado['notas_preenchidas']}, "
                        f"Ausentes: {resultado.get('ausentes', 0)}, "
                        f"Falhas: {resultado['falhas']}"
                    )

                else:
                    log_progress("Iniciando navegador e login no portal...")
                    from lancar_notas_sge import (
                        ContextoTurma,
                        _group_for_launch,
                        _login_sge,
                        _select_context,
                        _open_assessment_for_context,
                        TurmaNaoEncontradaError,
                        _handle_assessment_period_page,
                        _select_activity,
                        _fill_grade_for_student,
                        _read_existing_grade_for_student,
                        _classificar_leitura,
                        _confirm_save,
                        HEADLESS,
                    )
                    from status_store import StatusStore

                    status_store_path = fonte_path or os.path.join(_revisao_dir(), "retentar_ausentes")
                    status_store = StatusStore(status_store_path, logger=log_progress)
                    if fonte == "planilha":
                        _semear_status_manual(status_store)
                    ja_lancadas = 0
                    ja_no_sge = 0
                    ausentes_count = 0

                    from lancar_notas_sge import RegistroNota as SGEReg
                    registros_sge = [
                        SGEReg(
                            escola=r.escola, turno=r.turno, turma=r.turma,
                            trimestre=r.trimestre, aluno=r.aluno,
                            atividade=r.atividade, nota=r.nota,
                            data_realizacao=r.data_realizacao,
                        )
                        for r in registros
                    ]

                    pendentes = st.session_state.pop("revisao_pendentes", [])
                    if _rev_retentar:
                        # Re-tentativa de ausentes (nome/imagem ilegivel) com nome
                        # corrigido pelo professor: monta o lote SO com esses registros.
                        from lancar_notas_sge import _aplicar_pendentes_ausentes
                        registros_retentativa = _aplicar_pendentes_ausentes(pendentes)
                        if not registros_retentativa:
                            log_progress("[REVISAO] Nenhum ausente corrigido valido para re-tentar.")
                            st.session_state.revisao_forcar = False
                            st.session_state.resultado = {"blocos": 0, "notas": 0, "notas_preenchidas": 0, "ausentes": 0, "falhas": 0}
                            st.stop()
                        registros_sge = registros_retentativa
                        log_progress(f"[REVISAO] Re-tentando {len(registros_sge)} ausente(s) com nome corrigido no SGE.")
                    elif pendentes and not _rev_aplicar:
                        # Decisoes de revisao so sao consumidas na execucao disparada pelo
                        # botao "Gravar decisoes" (revisao_aplicar). Em um lancamento novo,
                        # uma fila antiga nao pode filtrar/esvaziar o lote extraido.
                        log_progress(f"[REVISAO] Ignorando {len(pendentes)} decisao(oes) pendentes de execucao anterior (lancamento novo).")
                        pendentes = []
                    if pendentes and not _rev_retentar:
                        from lancar_notas_sge import _aplicar_pendentes_revisao
                        registros_sge, _rev_forcar = _aplicar_pendentes_revisao(registros_sge, pendentes)
                        if _rev_forcar:
                            st.session_state.revisao_forcar = True
                            log_progress(f"[REVISAO] Aplicando decisoes: {len(registros_sge)} registro(s) para gravacao forçada.")
                        else:
                            log_progress("[REVISAO] Nenhuma decisao aplicavel ao lote extraido; prosseguindo com o lote normal.")

                    grouped = _group_for_launch(registros_sge)
                    log_progress(f"Blocos para lancamento: {len(grouped)}")

                    notas_ok = 0
                    falhas = 0
                    divergencias = 0
                    blocos_lancados = []

                    with _sessao_navegador(HEADLESS, st.session_state.sge_url, st.session_state.sge_cpf, st.session_state.sge_senha) as page:

                        _login_sge(page, cpf=st.session_state.sge_cpf, senha=st.session_state.sge_senha, logger=log_progress)
                        from lancar_notas_sge import _reset_navigation_cache
                        _reset_navigation_cache()

                        for idx, (key, itens) in enumerate(grouped.items(), start=1):
                            escola, turno, turma, trimestre, atividade = key
                            log_progress(f"[{idx}/{len(grouped)}] {escola} | {turno} | {turma} | {trimestre} | {atividade}")

                            from lancar_notas_sge import _date_diff_days, _dates_match
                            datas_bloco = [r.data_realizacao for r in itens if r.data_realizacao]
                            data_mais_comum = ""
                            if datas_bloco:
                                from collections import Counter
                                data_mais_comum = Counter(datas_bloco).most_common(1)[0][0]
                                diff_dias = _date_diff_days(data_mais_comum)
                                if diff_dias is not None and diff_dias > 0:
                                    log_progress(f"[DATA] Atividade com data futura ({data_mais_comum}). Pulando bloco.")
                                    falhas += len(itens)
                                    continue
                                elif diff_dias is not None and diff_dias < -90:
                                    log_progress(f"[DATA] Atencao: atividade com data antiga ({data_mais_comum}). Prosseguindo...")
                                elif diff_dias is not None:
                                    log_progress(f"[DATA] Data da atividade: {data_mais_comum} ({abs(diff_dias)} dia(s) atras)")
                            else:
                                log_progress("[DATA] Nenhuma data de realizacao definida na planilha para esta atividade.")

                            contexto = ContextoTurma(escola=escola, turno=turno, turma=turma, trimestre=trimestre)
                            _select_context(page, contexto, logger=log_progress)
                            motivo_pulo = ""
                            try:
                                _open_assessment_for_context(page, contexto, logger=log_progress)
                            except TurmaNaoEncontradaError as exc:
                                motivo_pulo = str(exc)
                            except Exception:  # noqa: BLE001
                                pass
                            if not motivo_pulo and _handle_assessment_period_page(page, contexto, logger=log_progress):
                                try:
                                    _open_assessment_for_context(page, contexto, logger=log_progress)
                                except TurmaNaoEncontradaError as exc:
                                    motivo_pulo = str(exc)
                                except Exception:  # noqa: BLE001
                                    pass
                                _handle_assessment_period_page(page, contexto, logger=log_progress)
                            if motivo_pulo:
                                log_progress(f"  [PULO] Bloco ignorado: {motivo_pulo} Nenhuma nota gravada nesta turma.")
                                falhas += len(itens)
                                continue
                            atividade_encontrada, data_sge, posicao_grid = _select_activity(page, atividade, logger=log_progress)

                            if not atividade_encontrada:
                                log_progress(f"  [AVISO] Atividade '{atividade}' nao encontrada no SGE. Pulando bloco.")
                                falhas += len(itens)
                                continue

                            if data_sge and data_mais_comum and not _dates_match(data_sge, data_mais_comum):
                                log_progress(f"  [DATA] Validacao falhou: SGE {data_sge} ≠ planilha {data_mais_comum}. Pulando bloco.")
                                falhas += len(itens)
                                continue
                            if data_sge and data_mais_comum:
                                log_progress(f"  [DATA] Datas conferem: SGE {data_sge} = planilha {data_mais_comum}")
                            elif not data_sge and data_mais_comum:
                                log_progress("  [DATA] Data da atividade nao encontrada no SGE. Prosseguindo sem validacao de data.")

                            from lancar_notas_sge import _detect_coluna_from_page
                            coluna_sge = _detect_coluna_from_page(page, posicao_grid, logger=log_progress, atividade=atividade)

                            from lancar_notas_sge import _check_estrutura_sge, ESTRUTURA_DIR
                            check_estrutura = _check_estrutura_sge(page, log_progress, contexto=contexto, atividade=atividade)
                            if not check_estrutura["ok"]:
                                log_progress(f"[ESTRUTURA-CHANGED] Layout da grade do SGE mudou. NAO gravando nada. Evidencia em {ESTRUTURA_DIR}.")
                                st.session_state.estrutura_changed = check_estrutura
                                st.session_state.revisao_fase = "estrutura_changed"
                                st.session_state.resultado = {
                                    "blocos": len(grouped),
                                    "notas": len(registros_sge),
                                    "notas_preenchidas": notas_ok,
                                    "ausentes": ausentes_count,
                                    "falhas": falhas,
                                    "divergencias": divergencias,
                                    "estrutura_changed": True,
                                }
                                st.session_state.executando = False
                                st.stop()

                            novos = 0
                            revisao_forcar = bool(st.session_state.get("revisao_forcar", False))
                            from lancar_notas_sge import _set_primeiro_nome_match
                            _set_primeiro_nome_match(bool(st.session_state.get("primeiro_nome_match", True)))
                            for reg in itens:
                                if status_store.esta_lancada(reg.escola, reg.turno, reg.turma, reg.trimestre, reg.aluno, reg.atividade):
                                    ja_lancadas += 1
                                    continue

                                existing = _read_existing_grade_for_student(page, reg.aluno, logger=log_progress, coluna_sge=coluna_sge)
                                nota_texto = str(reg.nota).replace(".", ",")
                                classe = _classificar_leitura(existing, nota_texto)
                                if classe == "ok":
                                    ja_no_sge += 1
                                    status_store.marcar_lancada(reg.escola, reg.turno, reg.turma, reg.trimestre, reg.aluno, reg.atividade, reg.nota)
                                    notas_ok += 1
                                    continue
                                if classe == "divergente" and not revisao_forcar:
                                    log_progress(f"  [DIVERGENCIA] '{reg.aluno}': SGE tem '{existing}', esperado '{nota_texto}'. Nao preencheu; enviado para revisao.")
                                    _coletar_divergencia(
                                        page, contexto, atividade, data_mais_comum,
                                        reg.aluno, reg.nota, existing, coluna_sge,
                                        logger=log_progress,
                                    )
                                    divergencias += 1
                                    continue
                                if classe == "divergente" and revisao_forcar:
                                    log_progress(f"  [REVISAO-FORCAR] Sobrescrevendo '{existing}' por '{nota_texto}' para '{reg.aluno}'.")

                                filled_suffix = _fill_grade_for_student(page, reg.aluno, reg.nota, logger=log_progress, coluna_sge=coluna_sge)
                                if filled_suffix:
                                    relido = _read_existing_grade_for_student(page, reg.aluno, logger=log_progress, coluna_sge=coluna_sge)
                                    classe_pos = _classificar_leitura(relido, nota_texto)
                                    if classe_pos == "ok" or revisao_forcar:
                                        notas_ok += 1
                                        novos += 1
                                        status_store.marcar_lancada(reg.escola, reg.turno, reg.turma, reg.trimestre, reg.aluno, reg.atividade, reg.nota)
                                    else:
                                        log_progress(f"  [DIVERGENCIA-POS] '{reg.aluno}': apos preencher, campo leu '{relido or 'vazio'}'. Enviado para revisao.")
                                        _coletar_divergencia(
                                            page, contexto, atividade, data_mais_comum,
                                            reg.aluno, reg.nota, relido, coluna_sge,
                                            logger=log_progress,
                                        )
                                        divergencias += 1
                                else:
                                    log_progress(f"  [AUSENTE] Aluno '{reg.aluno}' nao localizado na grade. Enviado para revisao (nome ou imagem ilegivel).")
                                    _coletar_ausente(
                                        page, contexto, atividade, data_mais_comum,
                                        reg.aluno, reg.nota, coluna_sge,
                                        logger=log_progress,
                                    )
                                    status_store.marcar_falha(
                                        reg.escola, reg.turno, reg.turma, reg.trimestre,
                                        reg.aluno, reg.atividade, reg.nota,
                                        erro="aluno nao localizado na grade",
                                    )
                                    ausentes_count += 1

                            if novos > 0:
                                _confirm_save(page, logger=log_progress, data_realizacao=data_mais_comum)
                                blocos_lancados.append({
                                    "contexto": contexto,
                                    "atividade": atividade,
                                    "itens": [r for r in itens],
                                    "data_realizacao": data_mais_comum,
                                })

                        falhas_revisao_qtd = 0
                        if st.session_state.get("revisar_apos", True) and blocos_lancados:
                            from lancar_notas_sge import _revisar_blocos_apos_lancamento
                            log_progress(f"[REVISAO] Re-auditoria pos-lancamento de {len(blocos_lancados)} bloco(s)...")
                            revisao = _revisar_blocos_apos_lancamento(page, blocos_lancados, logger=log_progress)
                            log_progress(
                                f"[REVISAO] Fim. Revisados: {revisao['revisados']} | Confirmados: {revisao['ok']} | "
                                f"Corrigidos: {revisao['corrigidos']} | Falhas: {revisao['falhas']} | IA usada: {revisao['ai_usada']}"
                            )

                            # Reflete o resultado da revisao na coluna Status do painel:
                            # confirmados (leitura ou IA) e corrigidos viram Lancada;
                            # nao confirmados viram Falha e entram na fila de revisao.
                            for reg_ok in list(revisao.get("regs_ok", [])) + list(revisao.get("regs_corrigidos", [])):
                                status_store.marcar_lancada(
                                    reg_ok.escola, reg_ok.turno, reg_ok.turma, reg_ok.trimestre,
                                    reg_ok.aluno, reg_ok.atividade, reg_ok.nota,
                                )
                            rev_itens = revisao.get("itens_nao_confirmados", []) or []
                            for item in rev_itens:
                                try:
                                    nota_item = float(str(item.get("nota_esperada", "")).replace(",", "."))
                                except ValueError:
                                    nota_item = 0.0
                                status_store.marcar_falha(
                                    item.get("escola", ""), item.get("turno", ""), item.get("turma", ""),
                                    item.get("trimestre", ""), item.get("aluno", ""), item.get("atividade", ""),
                                    nota_item, erro="revisao: nota nao confirmada no SGE",
                                )
                            _alimentar_fila_revisao(revisao, log_progress)
                            falhas_revisao_qtd = len(rev_itens)
                            if rev_itens:
                                log_progress(f"[PAINEL] {len(rev_itens)} linha(s) marcada(s) como Falha na planilha do painel (revisao).")

                        # Navegador persistente permanece aberto (reutilizado na proxima execucao).

                    log_progress(f"Status local: {ja_lancadas} ja lancadas, {ja_no_sge} ja no SGE, {notas_ok} preenchidas, {ausentes_count} ausentes, {falhas} falhas, {divergencias} divergencias, {falhas_revisao_qtd} falhas de revisao")

                    if fonte == "planilha":
                        _sincronizar_status_painel(status_store)
                        log_progress("[PAINEL] Coluna Status atualizada com o resultado do lancamento.")

                    st.session_state.revisao_forcar = False
                    fila_atual = st.session_state.get("revisao_fila", [])
                    st.session_state.revisao_fila = [it for it in fila_atual if it.get("decisao") in (None, "")]
                    st.session_state.revisao_fase = "pendente" if (divergencias > 0 or falhas_revisao_qtd > 0) else "fim"
                    st.session_state.resultado = {
                        "blocos": len(grouped),
                        "notas": len(registros_sge),
                        "notas_preenchidas": notas_ok,
                        "ausentes": ausentes_count,
                        "falhas": falhas + falhas_revisao_qtd,
                        "falhas_revisao": falhas_revisao_qtd,
                        "divergencias": divergencias,
                    }
                    log_progress(f"Concluido! Preenchidas: {notas_ok}, Ausentes: {ausentes_count}, Falhas: {falhas + falhas_revisao_qtd} (sendo {falhas_revisao_qtd} da revisao), Divergencias para revisao: {divergencias}")

                    if fonte == "planilha":
                        # A tabela do painel foi renderizada antes do lancamento
                        # (mesmo ciclo); força novo ciclo para exibir os status.
                        st.rerun()

        elif st.session_state.tipo == "chamada":
            from lancar_chamada_sge import executar_chamada as executar_chamada_sge

            foto_path = st.session_state.get("chamada_foto_path", "")
            if not foto_path:
                log_progress("ERRO: Nenhuma foto do diario selecionada.")
                st.error("Envie a foto da chamada do dia na seção 'Filtros' antes de executar.")
                st.session_state.resultado = {"success": False, "mensagem": "Sem foto", "plano": [], "resumo": {}, "nao_encontrados": []}
                st.stop()

            dia_input = st.session_state.get("chamada_dia", "").strip()
            if not dia_input:
                log_progress("ERRO: Informe o dia da chamada (dd/mm/aaaa).")
                st.error("Informe o **dia da chamada** no formato dd/mm/aaaa antes de executar.")
                st.session_state.resultado = {"success": False, "mensagem": "Sem dia", "plano": [], "resumo": {}, "nao_encontrados": []}
                st.stop()

            m_data = re.search(r"(\d{2})[/\-](\d{2})[/\-](\d{4})", dia_input)
            if not m_data:
                log_progress(f"ERRO: Dia '{dia_input}' em formato invalido. Use dd/mm/aaaa.")
                st.error(f"Dia '{dia_input}' em formato invalido. Use dd/mm/aaaa (ex.: 05/08/2026).")
                st.session_state.resultado = {"success": False, "mensagem": "Data invalida", "plano": [], "resumo": {}, "nao_encontrados": []}
                st.stop()
            dia_sge = f"{m_data.group(3)}/{m_data.group(2)}/{m_data.group(1)}"

            log_progress(f"[CHAMADA] Dia: {dia_sge} | Escola: {st.session_state.get('escola', '') or 'todas'}")
            try:
                portal_para_chamada = st.session_state.get("portal_resolvido", "SGE")
                if portal_para_chamada == "Professor Online":
                    from lancar_professor_online import executar_chamada as executar_chamada_po
                    log_progress("[CHAMADA] Portal: Professor Online")
                    resultado = executar_chamada_po(
                        filtro={
                            "escola": st.session_state.get("escola", ""),
                            "turno": st.session_state.get("turno", ""),
                            "turma": st.session_state.get("turma", ""),
                            "dia": dia_input.strip(),
                        },
                        foto_path=foto_path,
                        logger=log_progress,
                        dry_run=dry_run,
                        cpf=st.session_state.get("po_cpf", ""),
                        senha=st.session_state.get("po_senha", ""),
                        base_url=st.session_state.get("po_url", ""),
                    )
                else:
                    log_progress("[CHAMADA] Portal: SGE")
                    resultado = executar_chamada_sge(
                        filtro={
                            "escola": st.session_state.get("escola", ""),
                            "turno": st.session_state.get("turno", ""),
                            "turma": st.session_state.get("turma", ""),
                            "disciplina": st.session_state.get("chamada_disciplina", ""),
                            "dia": dia_sge,
                        },
                        foto_path=foto_path,
                        logger=log_progress,
                        dry_run=dry_run,
                    )
            except Exception as exc:  # noqa: BLE001
                log_progress(f"[CHAMADA] ERRO: {exc}")
                st.error(f"Falha no lancamento da chamada: {exc}")
                st.session_state.resultado = {"success": False, "mensagem": str(exc), "plano": [], "resumo": {}, "nao_encontrados": []}
                st.stop()

            st.session_state.resultado = resultado
            resumo = resultado.get("resumo", {})
            log_progress(
                f"[CHAMADA] Concluido! {resumo.get('presentes', 0)} presentes, "
                f"{resumo.get('faltas', 0)} falta(s), {resumo.get('ja_lancados', 0)} ja lancado(s)."
            )
            if resultado.get("nao_encontrados"):
                log_progress(f"[CHAMADA] Sem match na grade: {', '.join(resultado['nao_encontrados'])}")

        elif st.session_state.tipo == "faltas":
            portal_para_faltas = st.session_state.get("portal_resolvido", "SGE")
            log_progress(f"[FALTAS] Portal: {portal_para_faltas} | Escola: {st.session_state.get('escola', '') or 'todas'} | Turma: {st.session_state.get('turma', '') or 'todas'}")
            try:
                if portal_para_faltas == "Professor Online":
                    from lancar_professor_online import executar_faltas_mes
                    resultado = executar_faltas_mes(
                        filtro={
                            "escola": st.session_state.get("escola", ""),
                            "turno": st.session_state.get("turno", ""),
                            "turma": st.session_state.get("turma", ""),
                        },
                        logger=log_progress,
                        dry_run=dry_run,
                        cpf=st.session_state.get("po_cpf", ""),
                        senha=st.session_state.get("po_senha", ""),
                        base_url=st.session_state.get("po_url", ""),
                    )
                else:
                    log_progress("ERRO: Leitura de faltas do mes so esta disponivel no Professor Online por enquanto.")
                    st.warning("Leitura de faltas do mês: selecione o portal **Professor Online** (ou use o portal **Auto** com a escola registrada).")
                    resultado = {"success": False, "mensagem": "Faltas do mes: disponivel apenas no Professor Online.", "faltas": None, "resumo": {}}
            except Exception as exc:  # noqa: BLE001
                log_progress(f"[FALTAS] ERRO: {exc}")
                st.error(f"Falha ao ler faltas do mes: {exc}")
                resultado = {"success": False, "mensagem": str(exc), "faltas": None, "resumo": {}}

            st.session_state.resultado = resultado
            if resultado.get("success"):
                resumo_f = resultado.get("resumo", {})
                log_progress(
                    f"[FALTAS] Concluido! {resumo_f.get('alunos', 0)} aluno(s), "
                    f"{resumo_f.get('total_faltas', 0)} falta(s) no total "
                    f"(periodo {resumo_f.get('periodo', '')})."
                )
            else:
                log_progress(f"[FALTAS] {resultado.get('mensagem', '')}")

        elif st.session_state.tipo == "sequencia":
            from lancar_sequencia_didatica_sge import executar_lancamento_sequencia, SequenciaRegistro

            registros = None
            fonte = st.session_state.fonte

            if fonte == "notion":
                log_progress("Carregando dados do Notion...")

            elif fonte in ("excel", "csv"):
                log_progress("Carregando sequencias do arquivo...")
                file_path = st.session_state.get("arquivo_path", "")
                if not file_path:
                    log_progress("ERRO: Nenhum arquivo selecionado para Sequencia Didatica.")
                    st.session_state.executando = False
                    st.stop()
                raw = (
                    ler_sequencias_excel(file_path, logger=log_progress)
                    if fonte == "excel"
                    else ler_sequencias_csv(file_path, logger=log_progress)
                )
                if not raw:
                    log_progress("ERRO: Nenhuma sequencia valida encontrada no arquivo.")
                    st.session_state.executando = False
                    st.stop()
                registros = [SequenciaRegistro(**r) for r in raw]
                log_progress(f"{len(registros)} sequencia(s) carregada(s) do arquivo.")

            elif fonte == "google_drive":
                log_progress("Preparando lancamento via Google Drive...")
                drive_links = st.session_state.get("seq_drive_links", [])
                links_preenchidos = [e for e in drive_links if e.get("link", "").strip()]
                if not links_preenchidos:
                    log_progress("ERRO: Nenhum link preenchido. Adicione links na seção 'Filtros'.")
                    st.session_state.executando = False
                    st.stop()

                periodo_inicio = st.session_state.get("seq_periodo_inicio", "").strip()
                periodo_fim = st.session_state.get("seq_periodo_fim", "").strip()
                titulo_base = st.session_state.get("seq_titulo_documento", "").strip() or "Sequencia Didatica"
                n_aulas = int(st.session_state.get("seq_n_aulas", 4) or 4)

                registros = []
                for entry in links_preenchidos:
                    ano = entry["ano"].strip()
                    if not ano:
                        continue
                    link = entry["link"].strip()
                    titulo_doc = f"{titulo_base} - {ano}"
                    registros.append(SequenciaRegistro(
                        page_id="",
                        ano=ano,
                        escola=st.session_state.get("escola", ""),
                        turno=st.session_state.get("turno", ""),
                        turma=ano,
                        titulo_documento=titulo_doc,
                        arquivo_nome=f"{ano}.pdf",
                        arquivo_url=link,
                        link_arquivo=link,
                        periodo_inicio=periodo_inicio,
                        periodo_fim=periodo_fim,
                        n_aulas=n_aulas,
                        status="",
                    ))
                log_progress(f"{len(registros)} registro(s) preparado(s) via Google Drive.")

            else:
                log_progress(f"AVISO: Fonte '{fonte}' nao suportada para sequencia didatica.")
                st.warning("Use Notion, Excel, CSV ou Google Drive para sequencia didatica.")

            if registros is not None or fonte == "notion":
                portal_para_sequencia = st.session_state.get("portal_resolvido", "SGE")

                if portal_para_sequencia == "Professor Online":
                    if fonte == "notion":
                        log_progress("ERRO: Portal Professor Online nao usa fonte Notion. Use Excel, CSV ou Google Drive.")
                        st.error("Professor Online: use Excel, CSV ou Google Drive para sequencia didatica.")
                        st.session_state.executando = False
                        st.stop()
                    if not registros:
                        log_progress("ERRO: Nenhuma sequencia carregada para o Professor Online.")
                        st.session_state.executando = False
                        st.stop()

                    from lancar_professor_online import executar_planejamento

                    def _sequencia_para_po(reg):
                        return {
                            "escola": reg.escola,
                            "turno": reg.turno,
                            "turma": reg.turma,
                            "trimestre": st.session_state.trimestre or "2o Trimestre",
                            "titulo_documento": reg.titulo_documento,
                            "arquivo_nome": reg.arquivo_nome,
                            "arquivo_url": reg.arquivo_url,
                            "periodo_inicio": reg.periodo_inicio,
                            "periodo_fim": reg.periodo_fim,
                            "n_aulas": reg.n_aulas,
                        }

                    registros_po = [_sequencia_para_po(r) for r in registros]
                    log_progress("Executando sequencia didatica no Professor Online...")
                    res_po = executar_planejamento(
                        registros=registros_po,
                        logger=log_progress,
                        dry_run=dry_run,
                        cpf=st.session_state.get("po_cpf", ""),
                        senha=st.session_state.get("po_senha", ""),
                        base_url=st.session_state.get("po_url", ""),
                    )
                    st.session_state.resultado = {
                        "contextos": res_po.get("contextos", 0),
                        "planejamentos": res_po.get("planejamentos_criados", 0),
                        "anexos": res_po.get("anexos", 0),
                        "situacoes": res_po.get("situacoes", 0),
                        "falhas": res_po.get("falhas", 0),
                        "nao_implementado": res_po.get("nao_implementado", 0),
                    }
                    pend = res_po.get("nao_implementado", 0)
                    log_progress(f"Concluido! Planejamentos: {res_po.get('planejamentos_criados', 0)}, Falhas: {res_po.get('falhas', 0)}")
                    if pend:
                        log_progress(
                            f"[PO] {pend} sequencia(s) dependem de fluxo ainda nao gravado. "
                            "Para ensinar: selecione o portal Professor Online e marque 'Sim - Aprendizado' "
                            "na Assistencia IA; a IA grava o fluxo de criacao/upload e inclui a funcionalidade."
                        )
                        st.warning(
                            f"**Professor Online**: {pend} sequencia(s) ainda nao foram publicadas porque o "
                            "fluxo de criacao/upload de planejamento nao foi gravado. Use **'Sim - Aprendizado'** "
                            "na Assistencia IA para gravar o fluxo e adicionar a funcionalidade ao sistema."
                        )
                else:
                    log_progress("Executando sequencia didatica no SGE...")
                    resumo = executar_lancamento_sequencia(
                        escola=st.session_state.escola,
                        turno=st.session_state.turno,
                        turma=st.session_state.turma,
                        trimestre=st.session_state.trimestre or "2o Trimestre",
                        dry_run=dry_run,
                        registros=registros,
                        logger=log_progress,
                    )
                    st.session_state.resultado = {
                        "contextos": resumo.contextos_total,
                        "planejamentos": resumo.planejamentos_criados,
                        "anexos": resumo.anexos_enviados,
                        "situacoes": resumo.situacoes_ativadas,
                        "falhas": resumo.falhas,
                    }
                    log_progress(f"Concluido! Planejamentos: {resumo.planejamentos_criados}, Falhas: {resumo.falhas}")

    except RuntimeError as exc:
        if "Event loop is closed" in str(exc):
            log_progress("Execucao finalizada (limpeza async concluida).")
        else:
            _handle_exec_error(exc, log_progress)
    except Exception as exc:
        _handle_exec_error(exc, log_progress)

    finally:
        st.session_state.executando = False
        st.session_state.revisao_forcar = False
        _registrar_stats_resultado(st.session_state.get("resultado") or {})
        try:
            status.update(label="Finalizado", state="complete")
        except RuntimeError:
            pass

# === FILA DE CONFIRMACAO DE DIVERGENCIAS (screenshot + decisao) ===
with tab_rev:
    # === ALARME [ESTRUTURA-CHANGED] ===
    if st.session_state.get("estrutura_changed"):
        _ev = st.session_state["estrutura_changed"]
        st.error(
            "**Alarme [ESTRUTURA-CHANGED]** — o layout da grade de notas do SGE mudou e "
            "o bot **nao gravou nada** para evitar lancar nota no lugar errado."
        )
        st.caption(
            f"Slots de aluno reconhecidos: **{_ev.get('slots')}** | "
            f"Colunas de nota: **{_ev.get('colunas') or 'nenhuma'}** | "
            f"Inputs na pagina: **{_ev.get('inputs')}**"
        )
        _shot = (_ev.get("evidencia") or {}).get("screenshot", "")
        if _shot and os.path.exists(_shot):
            try:
                st.image(_shot, caption="Screenshot da tela que mudou (evidencia)", use_container_width=True)
            except Exception:  # noqa: BLE001
                st.caption(f"Evidencia disponivel em `{_shot}` (imagem nao pôde ser renderizada).")
        with st.expander("Detalhes tecnicos (HTML, info e sugestao da IA)"):
            st.code(_ler_arquivo_texto((_ev.get("evidencia") or {}).get("info", "")) or "sem info")
            if _ev.get("sugestao_ia"):
                st.markdown("**Sugestao automatica da IA no momento do alarme:**")
                st.code(_ev["sugestao_ia"])
            st.caption(
                "A IA e consultada SOMENTE neste desvio (custo sob demanda). "
                "O 'Remodelar' abaixo pede sua autorizacao antes de salvar qualquer override local."
            )
        st.markdown("##### Remodelar estrutura (com IA, requer seu OK)")
        if st.button("1. Analisar tela com IA e sugerir novos seletores", key="estrutura_remodelar", use_container_width=True):
            sugestao = _remodelar_estrutura_com_ia(_ev)
            st.session_state.estrutura_sugestao = sugestao
            st.rerun()
        if st.session_state.get("estrutura_sugestao"):
            st.success("A IA analisou a tela. **Revise** a sugestao abaixo e, se aprovar, salve o override local.")
            st.code(st.session_state["estrutura_sugestao"])
            c_aprov, c_desc = st.columns(2)
            with c_aprov:
                if st.button("2. Aprovar e salvar override local", type="primary", key="estrutura_salvar", use_container_width=True):
                    if _salvar_estrutura_override_do_sugestao(st.session_state["estrutura_sugestao"]):
                        st.success(
                            "Override salvo em `artifacts/estrutura/estrutura_override.json`. "
                            "Re-execute o lancamento para o bot usar a nova estrutura."
                        )
                        st.session_state.pop("estrutura_sugestao", None)
                    else:
                        st.error("Nao foi possivel extrair seletores da sugestao. Salve manualmente ou tente de novo.")
            with c_desc:
                if st.button("Descartar sugestao", key="estrutura_descartar", use_container_width=True):
                    st.session_state.pop("estrutura_sugestao", None)
                    st.rerun()
        if st.button("Limpar alarme", key="estrutura_limpar", use_container_width=True):
            st.session_state.pop("estrutura_changed", None)
            st.session_state.pop("estrutura_sugestao", None)
            if st.session_state.get("resultado"):
                st.session_state.resultado["estrutura_changed"] = False
            st.rerun()

    _fila_rev = st.session_state.get("revisao_fila", [])
    _pendentes_rev = [it for it in _fila_rev if it.get("decisao") in (None, "")]
    if st.session_state.get("revisao_fase") == "pendente" and _pendentes_rev:
        st.markdown("### Confirmacao de divergencias (IA + screenshot)")
        st.caption(
            "O bot detectou divergencias entre o SGE e o esperado e **nao preencheu** esses alunos. "
            "Confira a evidencia e decida por aluno: **Confirmar** (grava a nota esperada), "
            "**Corrigir** (grava outro valor) ou **Pular** (nao grava e nao marca como lancada)."
        )

        c_top1, c_top2, c_top3 = st.columns(3)
        with c_top1:
            if st.button("Confirmar todas", type="primary", key="rev_confirmar_todas", use_container_width=True):
                for it in _fila_rev:
                    if it.get("decisao") in (None, ""):
                        it["decisao"] = "confirmar"
                        it["valor_corrigido"] = ""
                st.session_state.revisao_fila = _fila_rev
                if _aplicar_decisoes_revisao():
                    st.rerun()
        with c_top2:
            if st.button("Pular todas", key="rev_pular_todas", use_container_width=True):
                for it in _fila_rev:
                    it["decisao"] = "pular"
                st.session_state.revisao_fila = _fila_rev
                st.session_state.revisao_fase = "fim"
                if st.session_state.get("resultado"):
                    st.session_state.resultado["divergencias"] = 0
                st.rerun()
        with c_top3:
            if st.button("Limpar fila", key="rev_limpar", use_container_width=True):
                st.session_state.revisao_fila = []
                st.session_state.revisao_fase = "fim"
                if st.session_state.get("resultado"):
                    st.session_state.resultado["divergencias"] = 0
                st.rerun()

        for item in _pendentes_rev:
            fkey = f"rev_{item.get('id', 'x')}"
            with st.expander(
                f"{item.get('aluno')} — esperado **{item.get('nota_esperada')}** | SGE leu "
                f"**'{item.get('nota_lida') or 'vazio'}'** | {item.get('atividade')}",
                expanded=False,
            ):
                col_shot, col_dec = st.columns([1, 2])
                with col_shot:
                    if os.path.exists(item.get("screenshot", "")):
                        st.image(item["screenshot"], caption="Evidencia (linha do aluno no SGE)")
                    else:
                        st.caption("Evidencia indisponivel")
                    st.caption(
                        f"Escola: {item.get('escola')} | Turno: {item.get('turno')} | "
                        f"Turma: {item.get('turma')} | {item.get('trimestre')}"
                    )
                with col_dec:
                    decisao = st.radio(
                        "Decisao para este aluno",
                        ["Confirmar", "Corrigir", "Pular"],
                        key=f"{fkey}_radio",
                        horizontal=True,
                        format_func=lambda x: {
                            "Confirmar": "Confirmar (gravar esperada)",
                            "Corrigir": "Corrigir",
                            "Pular": "Pular (nao gravar)",
                        }.get(x, x),
                    )
                    valor_corrigido = ""
                    if decisao == "Corrigir":
                        valor_corrigido = st.text_input(
                            "Valor correto (0 a 10)",
                            value=str(item.get("valor_corrigido") or item.get("nota_esperada")),
                            key=f"{fkey}_valor",
                        )
                    if st.button("Aplicar decisao", key=f"{fkey}_btn"):
                        item["decisao"] = {
                            "Confirmar": "confirmar",
                            "Corrigir": "corrigir",
                            "Pular": "pular",
                        }[decisao]
                        item["valor_corrigido"] = valor_corrigido if decisao == "Corrigir" else ""
                        st.session_state.revisao_fila = _fila_rev
                        st.rerun()

        decididos = [it for it in _fila_rev if it.get("decisao") in ("confirmar", "corrigir")]
        if decididos:
            if st.button(
                f"Gravar {len(decididos)} decisao(oes) no SGE",
                type="primary",
                key="rev_gravar",
            ):
                _aplicar_decisoes_revisao()
                st.rerun()

    # === AUSENTES: nome ou imagem ilegivel ===
    _fila_aus = st.session_state.get("revisao_fila", [])
    _aus_pendentes = [it for it in _fila_aus if it.get("tipo") == "ausente" and it.get("decisao") in (None, "")]
    _aus_retentaveis = [it for it in _fila_aus if it.get("tipo") == "ausente" and it.get("decisao") == "retentar"]
    if _aus_pendentes:
        st.markdown("### Ausentes — nome ou imagem ilegível")
        st.caption(
            "O nome lido na imagem **não foi localizado** na grade do SGE. "
            "Se a **letra** estiver ilegível, digite o nome correto (como aparece na grade) "
            "e clique em 'Re-tentar' — o bot relança só esse aluno. "
            "Se a **imagem** estiver ilegível, envie outra foto e clique em **Executar** novamente."
        )
        if _aus_retentaveis:
            if st.button(
                f"Re-tentar {len(_aus_retentaveis)} ausente(s) corrigido(s) no SGE",
                type="primary",
                key="rev_retentar_ausentes",
                use_container_width=True,
            ):
                n = _coletar_retentativa_ausentes()
                if n:
                    st.success(f"{n} ausente(s) agendado(s) para re-tentativa.")
                    st.rerun()
                else:
                    st.warning("Corrija ao menos um nome antes de re-tentar.")
        for item in _aus_pendentes:
            fkey = f"aus_{item.get('id', 'x')}"
            with st.expander(
                f"🔤 {item.get('aluno')} — nota esperada **{item.get('nota_esperada')}** | {item.get('atividade')}",
                expanded=False,
            ):
                col_shot, col_dec = st.columns([1, 2])
                with col_shot:
                    if os.path.exists(item.get("screenshot", "")):
                        st.image(item["screenshot"], caption="Evidência (grade do SGE)")
                    else:
                        st.caption("Evidência indisponível")
                    st.caption(
                        f"Escola: {item.get('escola')} | Turno: {item.get('turno')} | "
                        f"Turma: {item.get('turma')} | {item.get('trimestre')}"
                    )
                with col_dec:
                    novo_nome = st.text_input(
                        "Nome correto na grade do SGE (se a letra estiver ilegível)",
                        key=f"{fkey}_nome",
                    )
                    c_nome, c_pular = st.columns(2)
                    with c_nome:
                        if st.button("Re-tentar com este nome", key=f"{fkey}_btn_nome", use_container_width=True):
                            if not novo_nome.strip():
                                st.warning("Digite o nome correto antes de re-tentar.")
                            else:
                                item["aluno_corrigido"] = novo_nome.strip()
                                item["decisao"] = "retentar"
                                item["nova_imagem"] = ""
                                st.session_state.revisao_fila = _fila_aus
                                st.rerun()
                    with c_pular:
                        if st.button("Pular (não gravar)", key=f"{fkey}_btn_pular", use_container_width=True):
                            item["decisao"] = "pular"
                            st.session_state.revisao_fila = _fila_aus
                            st.rerun()
                    nova_img = st.file_uploader(
                        "Imagem ilegível? Envie outra foto da atividade",
                        type=["png", "jpg", "jpeg"],
                        key=f"{fkey}_img",
                    )
                    if nova_img is not None:
                        if st.button("Salvar esta imagem e relançar", key=f"{fkey}_btn_img", use_container_width=True):
                            _path = _salvar_nova_imagem_ausente(item, nova_img)
                            if _path:
                                item["nova_imagem"] = _path
                                item["decisao"] = "nova_imagem"
                                st.session_state.revisao_fila = _fila_aus
                                st.success(
                                    f"Nova imagem salva em `{_path}`. Envie-a (ou outra foto mais nítida) "
                                    "na seção **Filtros → Imagem/Foto** e clique em **Executar** novamente."
                                )
                                st.rerun()
                            else:
                                st.error("Não foi possível salvar a imagem enviada.")
    else:
        st.info("Nenhuma divergencia pendente de confirmacao. Ela aparece aqui quando o bot encontrar uma nota que difere do esperado.")

# === MENSAGEM DE AUTOFIX ===
autofix_msg = st.session_state.pop("autofix_message", None)
if autofix_msg:
    st.success(f"**Autofix:** {autofix_msg}")
    col_retry1, col_retry2 = st.columns([1, 3])
    with col_retry1:
        if st.button("Tentar novamente", type="primary", key="autofix_retry_btn"):
            st.session_state.autofix_trigger = True
            st.session_state.autofix_attempts = st.session_state.get("autofix_attempts", 0)
            st.rerun()
    with col_retry2:
        if st.button("Ignorar", key="autofix_ignore_btn"):
            st.session_state.autofix_attempts = 0
            st.rerun()

# === AREA DE LOGS ===
with tab_logs:
    col_log_header, col_log_export = st.columns([3, 1])
    with col_log_header:
        st.markdown("### Logs da Execucao")
        if st.session_state.get("log_file"):
            st.caption(f"Log completo salvo em: `{st.session_state['log_file']}`")
    with col_log_export:
        if st.session_state.logs:
            log_text = "\n".join(st.session_state.logs)
            st.download_button(
                "📥 Exportar logs",
                data=log_text,
                file_name=f"logs_bot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain",
                use_container_width=True,
            )

    log_container = st.container()

    _TAGS_TECNICAS = ("[GRID-SCAN", "[COLUNA-DETECT", "[GRADE]", "[SUGESTAO]", "[CONTEXTO]")
    _logs_visiveis = [
        m for m in st.session_state.logs[-50:]
        if not m.startswith(_TAGS_TECNICAS)
    ]
    _logs_tecnicos = [
        m for m in st.session_state.logs[-50:]
        if m.startswith(_TAGS_TECNICAS)
    ]

    with log_container:
        for msg in _logs_visiveis:
            if "ERRO" in msg or "erro" in msg.lower():
                st.markdown(f":red[{msg}]")
            elif "OK" in msg or "sucesso" in msg.lower() or "concluido" in msg.lower():
                st.markdown(f":green[{msg}]")
            elif "AVISO" in msg or "aviso" in msg.lower():
                st.markdown(f":orange[{msg}]")
            else:
                st.text(msg)

        if _logs_tecnicos:
            with st.expander("🔧 Detalhes tecnicos (para suporte)", expanded=False):
                for msg in _logs_tecnicos:
                    st.text(msg)

# === RESULTADO ===
if st.session_state.resultado:
    st.markdown("### Resultado")
    res = st.session_state.resultado

    _descartes_img = st.session_state.get("imagem_descartes") or []
    if _descartes_img:
        st.warning(
            f"**{len(_descartes_img)} leitura(s) da imagem foram ignoradas** por problema de qualidade. "
            "Confira se estes alunos precisam de lancamento manual: "
            + "; ".join(d.split(":")[0].strip(" ()") for d in _descartes_img[:10])
            + ("..." if len(_descartes_img) > 10 else "")
        )

    cols = st.columns(6)
    with cols[0]:
        st.metric("Blocos/Contextos", res.get("blocos") or res.get("contextos", 0))
    with cols[1]:
        st.metric("Notas", res.get("notas", 0))
    with cols[2]:
        preenchidas = res.get("notas_preenchidas") or res.get("planejamentos", 0)
        st.metric("Sucesso", preenchidas)
    with cols[3]:
        ausentes = res.get("ausentes", 0)
        st.metric("Ausentes", ausentes)
    with cols[4]:
        falhas = res.get("falhas", 0)
        st.metric("Falhas", falhas, delta_color="inverse")
    with cols[5]:
        diverg_res = res.get("divergencias", 0)
        st.metric("Divergencias", diverg_res, delta_color="off")

    if res.get("divergencias", 0) > 0:
        st.warning(f"{res.get('divergencias')} divergencia(s) aguardando confirmacao na aba **Pendências**.")
    elif falhas > 0:
        st.warning(f"Houve {falhas} falha(s). Verifique os logs na aba **Logs**.")
    elif preenchidas > 0:
        st.success("Tudo concluido com sucesso!")

    evid_dir = _revisao_dir()
    evidencias = sorted(
        [os.path.join(evid_dir, f) for f in os.listdir(evid_dir) if f.lower().endswith((".png", ".jpg", ".jpeg"))],
        key=lambda p: os.path.getmtime(p),
        reverse=True,
    )
    if evidencias:
        with st.expander(f"Ver evidencias ({len(evidencias)} screenshot(s) da revisao)"):
            ecols = st.columns(3)
            for ei, ep in enumerate(evidencias):
                with ecols[ei % 3]:
                    st.image(ep, caption=os.path.basename(ep))

# === CHAT COM IA (comandos em linguagem natural) ===
with st.expander("💬 Assistente (comandos em linguagem natural)", expanded=False):
    st.caption("Digite um comando (ex.: 'lanca a chamada de hoje do 7o ano da tarde'). O plano e aplicado nos filtros acima.")
    if "chat_msgs" not in st.session_state:
        st.session_state.chat_msgs = []
    for _m in st.session_state.chat_msgs:
        with st.chat_message(_m["role"]):
            st.markdown(_m["text"])

    _chat_input = st.chat_input("Ex.: lanca a chamada de hoje do 7o ano da tarde")
    if _chat_input:
        st.session_state.chat_msgs.append({"role": "user", "text": _chat_input})
        try:
            from interpretar_pedido import interpretar_pedido as _interpretar_pedido
            _plano = _interpretar_pedido(_chat_input, logger=log)
            if _plano.get("error"):
                _resp = f"⚠️ {_plano['error']}"
            else:
                _tipo = _plano.get("tipo", "")
                _fonte = _plano.get("fonte", "")
                _partes = [f"**Plano:** tipo=`{_tipo or '?'}` | fonte=`{_fonte or '?'}`"]
                for _campo in ("escola", "turma", "turno", "trimestre", "atividade", "data_realizacao", "chamada_dia"):
                    if _plano.get(_campo):
                        _partes.append(f"{_campo.replace('_', ' ')}={_plano[_campo]}")
                if _plano.get("lote"):
                    _partes.append("lote=sim")
                _resp = " | ".join(_partes)
                if _plano.get("confianca"):
                    _resp += f"\n\nConfianca: {_plano['confianca']}"
                if _plano.get("procedimentos"):
                    _resp += "\n\n**Procedimentos:**\n" + "\n".join(f"- {_p}" for _p in _plano["procedimentos"])
                if _tipo in ("notas", "chamada", "sequencia", "faltas"):
                    st.session_state["tipo_radio"] = _tipo
                if _fonte in ("notion", "imagem", "planilha", "excel", "csv", "google_sheets", "google_drive"):
                    st.session_state["fonte_select"] = _fonte
                for _campo, _chave in (
                    ("escola", "escola_input"), ("turma", "turma_input"),
                    ("turno", "turno_select"), ("trimestre", "trimestre_select"),
                ):
                    if _plano.get(_campo) and _chave in st.session_state:
                        st.session_state[_chave] = _plano[_campo]
                st.session_state["chat_aplicar"] = True
                log(f"[CHAT] Plano aplicado nos filtros: tipo={_tipo} fonte={_fonte}")
        except Exception as _exc:  # noqa: BLE001
            _resp = f"⚠️ Nao consegui interpretar: {_exc}"
        st.session_state.chat_msgs.append({"role": "assistant", "text": _resp})
        st.rerun()

# === RODAPE ===
st.markdown("---")
st.markdown(
    "<small>Bot do Professor v1.4.43 - Automacao de lancamento de notas. "
    "Use com responsabilidade. Verifique os dados antes de lancar.</small>",
    unsafe_allow_html=True,
)
