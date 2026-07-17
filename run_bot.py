#!/usr/bin/env python3
"""BotDoProfessor-Local — Executavel interativo.

Executavel CLI que o usuario final roda na maquina local.
Valida licenca, conecta no SGE e executa lancamentos.
"""

import os
import re
import sys
import csv
import json
import tempfile
import logging
import subprocess
import urllib.request
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

if os.name == "nt":
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        ctypes.windll.kernel32.SetConsoleCP(65001)
    except Exception:
        pass

CONFIG_DIR = Path.home() / ".bot_local"
CONFIG_FILE = CONFIG_DIR / "config.json"
LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"

DOWNLOAD_URL = "https://github.com/FelipeMarques-bot/botdoprofessor-local/releases/latest"
SERVER_URL = os.environ.get("SERVER_URL", "https://botdoprofessor.onrender.com")


def setup_logging():
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = CONFIG_DIR / "logs" / "bot.log"
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(format=LOG_FORMAT, level=logging.INFO)
    fh = logging.FileHandler(str(log_file), encoding="utf-8")
    fh.setFormatter(logging.Formatter(LOG_FORMAT))
    logging.getLogger().addHandler(fh)


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def banner():
    print("=" * 56)
    print("          BotDoProfessor — Automatize suas notas")
    print("=" * 56)
    print()
    print("  Este programa lanca notas, planos de aula e")
    print("  sequencias didaticas automaticamente no SGE.")
    print()


def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_config(data):
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _find_python():
    import shutil

    def _is_valid_python(path):
        if not path:
            return False
        lower = path.lower()
        if "windowsapps" in lower:
            return False
        if path.endswith("python3.exe") and "local" in lower and "microsoft" in lower:
            return False
        try:
            return os.path.isfile(path) or os.access(path, os.X_OK)
        except Exception:
            return False

    py_path = shutil.which("py")
    if py_path and _is_valid_python(py_path):
        return py_path

    for name in ("python3", "python"):
        path = shutil.which(name)
        if _is_valid_python(path):
            return path

    uv_base = os.path.expandvars(r"%APPDATA%\uv\python")
    common_paths = [
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Python\Python311\python.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Python\Python310\python.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Python\Python312\python.exe"),
        r"C:\Python311\python.exe",
        r"C:\Python310\python.exe",
        r"C:\Python312\python.exe",
    ]
    if os.path.isdir(uv_base):
        import glob
        for uv_py in sorted(glob.glob(os.path.join(uv_base, "cpython-3.*", "python.exe")), reverse=True):
            common_paths.insert(0, uv_py)
    for p in common_paths:
        if os.path.isfile(p):
            return p

    return sys.executable


def check_playwright():
    ms_playwright = Path(os.environ.get("LOCALAPPDATA", "")) / "ms-playwright"
    if not ms_playwright.exists():
        ms_playwright = Path.home() / ".cache" / "ms-playwright"
    if ms_playwright.exists():
        for d in ms_playwright.iterdir():
            if d.is_dir() and d.name.startswith("chromium-"):
                chrome = d / "chrome-win64" / "chrome.exe"
                if not chrome.exists():
                    chrome = d / "chrome-win" / "chrome.exe"
                if not chrome.exists():
                    chrome = d / "chrome-linux" / "chrome"
                if not chrome.exists():
                    chrome = d / "chrome-mac" / "Chromium.app"
                if chrome.exists():
                    return True
    try:
        from playwright.sync_api import sync_playwright
        p = sync_playwright().start()
        try:
            browser = p.chromium.launch(headless=True)
            browser.close()
        except Exception:
            p.stop()
            return False
        p.stop()
        return True
    except (ImportError, Exception):
        return False


def install_playwright():
    is_frozen = getattr(sys, "frozen", False)

    if is_frozen:
        print("[!] Componentes nao encontrados.")
        print("[i] Baixando navegador Chromium (~180MB, primeira vez)...")
        print("[i] Isso demora cerca de 2 minutos e so acontece uma vez.")
        print()
        try:
            base = Path(sys._MEIPASS) / "playwright" / "driver"
            node_exe = base / "node.exe"
            cli_js = base / "package" / "cli.js"
            if not node_exe.exists():
                raise FileNotFoundError(f"node.exe nao encontrado em {node_exe}")
            subprocess.run([str(node_exe), str(cli_js), "install", "chromium"], check=True)
            print()
            print("  [OK] Navegador instalado com sucesso!")
            return
        except Exception as e:
            print(f"  [ER] Falha ao baixar navegador: {e}")
            print()
            print("  Para instalar manualmente:")
            print("    1. Abra o Prompt de Comando (Windows+R, digite 'cmd', pressione Enter)")
            print("    2. Digite: pip install playwright")
            print("    3. Digite: python -m playwright install chromium")
            print()
            print(f"  Ou baixe o programa novamente: {DOWNLOAD_URL}")
            return

    python = _find_python()
    print("[!] Componentes nao encontrados.")
    print(f"[i] Usando Python: {python}")
    print("[i] Instalando dependencias...")
    print("[i] Isso demora cerca de 2 minutos e so acontece uma vez.")
    print()
    try:
        subprocess.run([python, "-m", "pip", "install", "playwright", "--quiet"], check=True)
        print("  [OK] Dependencias instaladas")
    except Exception as e:
        print(f"  [ER] Falha ao instalar dependencias: {e}")
        print("[i] Tente manualmente: pip install playwright")
        return

    print("[i] Baixando navegador Chromium (~180MB)...")
    try:
        subprocess.run([python, "-m", "playwright", "install", "chromium"], check=True)
        print("  [OK] Navegador instalado com sucesso!")
    except Exception as e:
        print(f"  [ER] Falha ao baixar navegador: {e}")
        print("[i] Tente manualmente: python -m playwright install chromium")


def validate_license(license_key):
    try:
        import urllib.request
        import urllib.error
        req = urllib.request.Request(
            f"{SERVER_URL}/api/license/public-validate",
            data=json.dumps({"license_key": license_key}).encode(),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            return data
    except urllib.error.HTTPError as e:
        try:
            data = json.loads(e.read().decode())
            return data
        except Exception:
            return {"valid": False, "error": f"Erro HTTP {e.code}"}
    except Exception as e:
        return {"valid": False, "error": f"Servidor indisponivel: {e}"}


def get_license_key():
    config = load_config()
    saved_key = config.get("license_key", "")
    if saved_key:
        masked = saved_key[:8] + "..." + saved_key[-4:]
        print(f"[i] Chave salva: {masked}")
        use_saved = input("  Usar esta chave? (S/n): ").strip().lower()
        if use_saved != "n":
            return saved_key

    print()
    print("  Cole sua chave de licenca abaixo.")
    print("  (Voce recebeu esta chave por email apos o pagamento)")
    print()
    key = input("  Chave: ").strip()
    if not key:
        print()
        print("  [ER] Chave obrigatoria. Verifique o email que recebeu.")
        return get_license_key()
    return key


def get_sge_credentials():
    config = load_config()
    saved_cpf = config.get("cpf", "")
    if saved_cpf:
        masked = saved_cpf[:3] + "***" + saved_cpf[-2:]
        print(f"[i] CPF salvo: {masked}")
        use_saved = input("  Usar este CPF? (S/n): ").strip().lower()
        if use_saved != "n":
            return saved_cpf

    print()
    print("  Informe o CPF que voce usa para acessar o SGE.")
    print("  (So numeros, sem pontos ou traco)")
    print()
    cpf = input("  CPF: ").strip().replace(".", "").replace("-", "")
    if not cpf or len(cpf) < 11:
        print()
        print("  [ER] CPF invalido. Digite 11 numeros.")
        return get_sge_credentials()
    return cpf


def get_context():
    config = load_config()
    print()
    print("=" * 56)
    print("  Configuracao do lancamento")
    print("=" * 56)
    print()
    print("  Informe os dados da turma que voce vai lancar.")
    print("  Pressione Enter para usar o valor entre colchetes [].")
    print()

    escola = input(f"  Escola [{config.get('escola', 'ex: Escola Municipal ABC')}]: ").strip()
    if not escola:
        escola = config.get("escola", "")

    turno = input(f"  Turno [{config.get('turno', 'Manha')}]: ").strip()
    if not turno:
        turno = config.get("turno", "Manha")

    turma = input(f"  Turma [{config.get('turma', 'ex: 5o Ano A')}]: ").strip()
    if not turma:
        turma = config.get("turma", "")

    trimestre = input(f"  Trimestre [{config.get('trimestre', '1o Trimestre')}]: ").strip()
    if not trimestre:
        trimestre = config.get("trimestre", "1o Trimestre")

    print()
    print(f"  Escola: {escola or '(nao informada)'}")
    print(f"  Turno:  {turno}")
    print(f"  Turma:  {turma or '(nao informada)'}")
    print(f"  Periodo: {trimestre}")

    save_config({
        **config,
        "escola": escola,
        "turno": turno,
        "turma": turma,
        "trimestre": trimestre,
    })

    return {"escola": escola, "turno": turno, "turma": turma, "trimestre": trimestre}


def is_gsheets_url(url):
    return bool(re.search(r"docs\.google\.com/spreadsheets", url))


def download_gsheets_as_csv(url):
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        print("[ER] Nao foi possivel extrair o ID da planilha.")
        return None

    sheet_id = m.group(1)
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    print("[i] Baixando planilha do Google Sheets...")

    try:
        tmp_dir = tempfile.mkdtemp(prefix="gsheets_")
        target = os.path.join(tmp_dir, "planilha.csv")
        req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        with open(target, "wb") as f:
            f.write(data)
        print(f"[OK] Planilha baixada ({len(data)} bytes)")
        return target
    except Exception as e:
        print(f"[ER] Falha ao baixar: {e}")
        print("[i] Verifique se a planilha esta compartilhada como 'Qualquer pessoa com o link'")
        return None


def load_grades_from_file(filepath):
    ext = Path(filepath).suffix.lower()
    grades = []

    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            if not reader.fieldnames:
                reader = csv.DictReader(f, delimiter=",")
                f.seek(0)
            for row in reader:
                aluno = ""
                nota = ""
                for key, val in row.items():
                    kl = (key or "").lower().strip()
                    if "aluno" in kl or "nome" in kl:
                        aluno = (val or "").strip()
                    elif "nota" in kl or "valor" in kl or "media" in kl:
                        nota = (val or "").strip()
                if aluno and nota:
                    grades.append({"aluno": aluno, "nota": nota})
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [str(c.value or "").lower() for c in next(ws.iter_rows(max_row=1))]
            aluno_idx = next((i for i, h in enumerate(headers) if "aluno" in h or "nome" in h), None)
            nota_idx = next((i for i, h in enumerate(headers) if "nota" in h or "valor" in h or "media" in h), None)
            if aluno_idx is not None and nota_idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    aluno = str(row[aluno_idx] or "").strip()
                    nota = str(row[nota_idx] or "").strip()
                    if aluno and nota:
                        grades.append({"aluno": aluno, "nota": nota})
            wb.close()
        except ImportError:
            print("[ER] Para ler .xlsx, instale: pip install openpyxl")
            return []
    else:
        print(f"[ER] Formato nao suportado: {ext}")
        print("[i] Use CSV (.csv) ou Excel (.xlsx)")
        return []

    return grades


def print_grades_help():
    print()
    print("  FORMATO DA PLANILHA:")
    print("  Crie uma planilha com duas colunas:")
    print()
    print("    Aluno         | Nota")
    print("    Maria Silva   | 8.5")
    print("    Joao Santos   | 7.0")
    print("    Ana Oliveira  | 9.2")
    print()
    print("  Colunas aceitas:")
    print("    - Aluno: 'Aluno' ou 'Nome'")
    print("    - Nota:  'Nota', 'Valor' ou 'Media'")
    print()
    print("  Formatos aceitos:")
    print("    - CSV separado por virgula ou ponto-e-virgula")
    print("    - Excel (.xlsx)")
    print("    - Google Sheets (cole o link da planilha)")
    print()


def execute_grades(cpf, context):
    print()
    print("=" * 56)
    print("  Modo: Lancamento de notas")
    print("=" * 56)
    print()
    print("  [1] Importar de arquivo (recomendado)")
    print("  [2] Digitar notas manualmente")
    print()

    choice = input("  Opcao: ").strip()

    grades = []
    if choice == "2":
        print()
        print("  Digite as notas no formato: Nome do Aluno;Nota")
        print("  Exemplo: Maria Silva;8.5")
        print("  Pressione Enter em linha vazia para finalizar:")
        print()
        while True:
            line = input("  > ").strip()
            if not line:
                break
            parts = line.split(";")
            if len(parts) >= 2:
                grades.append({"aluno": parts[0].strip(), "nota": parts[1].strip()})
            else:
                print("  [!] Formato: Nome do Aluno;Nota  (ex: Maria Silva;8.5)")
    else:
        print()
        print("  Cole o caminho do arquivo ou o link do Google Sheets:")
        print("  (Arraste o arquivo aqui ou cole o caminho completo)")
        print()
        filepath = input("  Arquivo: ").strip().strip('"')

        if not filepath:
            print("[ER] Nenhum arquivo informado.")
            return

        if is_gsheets_url(filepath):
            downloaded = download_gsheets_as_csv(filepath)
            if not downloaded:
                return
            filepath = downloaded
        elif not os.path.isfile(filepath):
            print(f"[ER] Arquivo nao encontrado: {filepath}")
            print("[i] Verifique se o caminho esta correto.")
            return

        grades = load_grades_from_file(filepath)
        if not grades:
            print("[ER] Nenhuma nota encontrada no arquivo.")
            print_grades_help()
            return
        print(f"[OK] {len(grades)} notas carregadas.")

    if not grades:
        print("[ER] Nenhuma nota para lancar.")
        return

    print()
    print(f"  Resumo: {len(grades)} notas serao lancadas.")
    print()
    confirm = input("  Confirmar lancamento? (S/n): ").strip().lower()
    if confirm == "n":
        print("  Lancamento cancelado.")
        return

    print()
    print(f"[i] Lancando {len(grades)} notas...")
    print()

    from bot.core.sge_adapter import SGEAdapter
    from bot.core.portal_adapter import PortalContext
    from bot.core.engine import BotEngine

    adapter = SGEAdapter()
    try:
        print("  [1/4] Iniciando navegador...")
        adapter.start()
        print("  [OK] Navegador iniciado")

        print("  [2/4] Fazendo login no SGE...")
        if not adapter.login(cpf, ""):
            print("  [ER] Login falhou. Verifique o CPF.")
            return
        print("  [OK] Login realizado")

        ctx = PortalContext(
            escola=context["escola"],
            turno=context["turno"],
            turma=context["turma"],
            trimestre=context["trimestre"],
        )

        print("  [3/4] Navegando ate a turma...")
        adapter.navigate_to(ctx)
        print("  [OK] Navegacao concluida")

        print("  [4/4] Lancando notas...")
        engine = BotEngine(adapter, execution_id="cli-exec")
        result = engine.run(grades, context=ctx)

        print()
        print("=" * 40)
        print("  RESULTADO")
        print("=" * 40)
        print(f"  Lancadas:  {result.filled}")
        print(f"  Falhas:    {result.failed}")
        print(f"  Puladas:   {result.skipped}")

        if result.failed > 0:
            print()
            print("  Detalhes das falhas:")
            for d in result.details:
                if d.get("status") in ("failed", "error"):
                    print(f"    - {d['aluno']}: {d.get('reason', d.get('error', '?'))}")

        if result.filled > 0:
            print()
            save = input("  Salvar no SGE? (S/n): ").strip().lower()
            if save != "n":
                adapter.save()
                print("  [OK] Salvo com sucesso!")

    except Exception as e:
        print(f"  [ER] Erro: {e}")
        logging.exception("Erro na execucao")
    finally:
        adapter.stop()


def execute_lesson_plan(cpf, context):
    print()
    print("=" * 56)
    print("  Modo: Plano de Aula")
    print("=" * 56)
    print()
    print("  Informe os dados do plano de aula que quer criar no SGE.")
    print()

    titulo = input("  Titulo do plano: ").strip()
    if not titulo:
        print("[ER] Titulo obrigatorio.")
        return

    data_inicio = input("  Data inicio (DD/MM/AAAA): ").strip()
    data_fim = input("  Data fim (DD/MM/AAAA): ").strip()
    n_aulas = input("  Numero de aulas [1]: ").strip()
    n_aulas = int(n_aulas) if n_aulas.isdigit() else 1

    print()
    print("  Se tiver um PDF do plano de aula, informe o caminho abaixo.")
    print("  Se nao tiver, pressione Enter para pular.")
    print()
    pdf_path = input("  Caminho do PDF (vazio para pular): ").strip().strip('"')
    if pdf_path and not os.path.isfile(pdf_path):
        print(f"[ER] Arquivo nao encontrado: {pdf_path}")
        pdf_path = ""

    print()
    print(f"  Criando plano: {titulo}")
    print(f"  Periodo: {data_inicio} a {data_fim}")
    print(f"  Aulas: {n_aulas}")

    from bot.core.sge_adapter import SGEAdapter
    from bot.core.portal_adapter import PortalContext, LessonPlan

    adapter = SGEAdapter()
    try:
        print("  [1/5] Iniciando navegador...")
        adapter.start()
        print("  [OK] Navegador iniciado")

        print("  [2/5] Fazendo login no SGE...")
        if not adapter.login(cpf, ""):
            print("  [ER] Login falhou. Verifique o CPF.")
            return
        print("  [OK] Login realizado")

        ctx = PortalContext(
            escola=context["escola"],
            turno=context["turno"],
            turma=context["turma"],
            trimestre=context["trimestre"],
        )

        print("  [3/5] Navegando para Plano de Aulas...")
        if not adapter.navigate_to_lesson_plan(ctx):
            print("  [ER] Nao foi possivel navegar para Plano de Aulas.")
            return
        print("  [OK] Navegacao concluida")

        plan = LessonPlan(
            titulo=titulo,
            data_inicio=data_inicio,
            data_fim=data_fim,
            n_aulas=n_aulas,
        )

        print("  [4/5] Criando planejamento...")
        if not adapter.create_lesson_plan(plan):
            print("  [ER] Falha ao criar planejamento.")
            return
        print("  [OK] Planejamento criado")

        if pdf_path:
            print("  [5/5] Enviando PDF...")
            if adapter.upload_lesson_plan_pdf(titulo, pdf_path):
                print("  [OK] PDF enviado")
            else:
                print("  [!!] Falha ao enviar PDF (planejamento ja criado)")

        print()
        print("[OK] Plano de aula executado com sucesso!")

    except Exception as e:
        print(f"  [ER] Erro: {e}")
        logging.exception("Erro na execucao")
    finally:
        adapter.stop()


def main():
    setup_logging()
    clear()
    banner()

    if not check_playwright():
        install_playwright()
        if not check_playwright():
            print("[ER] Nao foi possivel configurar os componentes.")
            print()
            print("  Solucoes possiveis:")
            print(f"  1. Baixe o programa novamente: {DOWNLOAD_URL}")
            print("  2. Verifique sua conexao com a internet")
            print("  3. Entre em contato: labintelligenceappoiments@gmail.com")
            print()
            input("  Pressione Enter para sair...")
            sys.exit(1)

    print("[OK] Componentes configurados")
    print()

    license_key = get_license_key()
    print()
    print("[i] Validando licenca...")
    lic_result = validate_license(license_key)

    if not lic_result.get("valid"):
        print()
        print(f"[ER] Licenca invalida: {lic_result.get('error', 'desconhecido')}")
        print()
        print("  Possiveis causas:")
        print("  - Chave incorreta (verifique o email)")
        print("  - Licenca expirada")
        print("  - Servidor temporariamente indisponivel")
        print()
        input("  Pressione Enter para sair...")
        sys.exit(1)

    plan = lic_result.get("plan", "?")
    days = lic_result.get("days_remaining", "?")
    print(f"[OK] Licenca valida — plano: {plan} — {days} dias restantes")

    save_config({"license_key": license_key})
    cpf = get_sge_credentials()
    save_config({"cpf": cpf})
    context = get_context()

    print()
    print("=" * 56)
    print("  O que voce quer lancar?")
    print("=" * 56)
    print()
    print("  [1] Notas de alunos")
    print("  [2] Plano de aula")
    print()

    choice = input("  Opcao: ").strip()

    if choice == "2":
        execute_lesson_plan(cpf, context)
    else:
        execute_grades(cpf, context)

    print()
    print("=" * 56)
    print("  Operacao concluida!")
    print(f"  Logs salvos em: {CONFIG_DIR / 'logs' / 'bot.log'}")
    print("=" * 56)
    print()
    input("  Pressione Enter para sair...")


if __name__ == "__main__":
    main()
