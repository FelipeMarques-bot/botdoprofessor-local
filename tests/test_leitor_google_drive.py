import os
import sys

import pytest


def _write_xlsx(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"
    ws.append(["Escola", "Turno", "Turma", "Trimestre", "Nome do Aluno", "Atividade 1"])
    ws.append(["Juvenal", "Matutino", "6o Ano", "1o Trimestre", "Ana Oliveira", "8,5"])
    ws.append(["Juvenal", "Matutino", "6o Ano", "1o Trimestre", "Joao Santos", "7,0"])
    wb.save(path)


def _write_csv(path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("Nome do Aluno,Atividade 1\n")
        f.write("Ana Oliveira,8,5\n")
        f.write("Joao Santos,7,0\n")


def _fake_download(src):
    import shutil

    def _download(url_or_id, output=None, quiet=True, fuzzy=False):
        shutil.copy(src, output)
        return output

    return _download


class TestLerNotasGoogleDrive:
    def test_gdown_faltando_levanta_erro_claro(self, monkeypatch):
        import leitor_planilhas

        monkeypatch.setitem(sys.modules, "gdown", None)
        with pytest.raises(RuntimeError) as exc:
            leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert "pip install gdown" in str(exc.value)

    def test_baixa_xlsx_e_carrega_notas(self, tmp_path, monkeypatch):
        import gdown
        import leitor_planilhas

        src = tmp_path / "notas.xlsx"
        _write_xlsx(str(src))
        monkeypatch.setattr(gdown, "download", _fake_download(str(src)))

        registros = leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert len(registros) == 2
        assert registros[0].aluno == "Ana Oliveira"
        assert registros[0].nota == 8.5

    def test_baixa_csv_e_carrega_notas(self, tmp_path, monkeypatch):
        import gdown
        import leitor_planilhas

        src = tmp_path / "notas.csv"
        _write_csv(str(src))
        monkeypatch.setattr(gdown, "download", _fake_download(str(src)))

        registros = leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert len(registros) == 2
        assert registros[1].aluno == "Joao Santos"

    def test_download_sem_sucesso_levanta_erro_de_compartilhamento(self, tmp_path, monkeypatch):
        import gdown
        import leitor_planilhas

        def _sem_sucesso(url_or_id, output=None, quiet=True, fuzzy=False):
            return None

        monkeypatch.setattr(gdown, "download", _sem_sucesso)
        with pytest.raises(RuntimeError) as exc:
            leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert "Qualquer um com o link pode ver" in str(exc.value)

    def test_arquivo_vazio_levanta_erro_claro(self, tmp_path, monkeypatch):
        import gdown
        import leitor_planilhas

        src = tmp_path / "vazio.xlsx"
        src.write_bytes(b"")

        def _vazio(url_or_id, output=None, quiet=True, fuzzy=False):
            import shutil

            shutil.copy(str(src), output)
            return output

        monkeypatch.setattr(gdown, "download", _vazio)
        with pytest.raises(RuntimeError) as exc:
            leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert "esta vazio" in str(exc.value)

    def test_identifica_tipo_pelo_conteudo(self, tmp_path):
        from leitor_planilhas import _detect_spreadsheet_type

        xlsx = tmp_path / "a.xlsx"
        _write_xlsx(str(xlsx))
        assert _detect_spreadsheet_type(str(xlsx)) == "excel"

        csv = tmp_path / "b.csv"
        _write_csv(str(csv))
        assert _detect_spreadsheet_type(str(csv)) == "csv"

    def test_tmp_arquivo_removido_apos_sucesso(self, tmp_path, monkeypatch):
        import gdown
        import leitor_planilhas
        import tempfile

        src = tmp_path / "notas.xlsx"
        _write_xlsx(str(src))
        criados = []
        original = tempfile.NamedTemporaryFile

        def fake_tmp(*args, **kwargs):
            t = original(*args, **kwargs)
            criados.append(t.name)
            return t

        monkeypatch.setattr(tempfile, "NamedTemporaryFile", fake_tmp)
        monkeypatch.setattr(gdown, "download", _fake_download(str(src)))

        leitor_planilhas.ler_notas_google_drive("https://drive.google.com/file/d/ABC/view")
        assert criados, "esperava um arquivo temporario"
        assert not os.path.exists(criados[0])


class TestGoogleSheets:
    def test_sheets_export_xlsx_le_todas_as_abas(self, tmp_path, monkeypatch):
        import io
        import leitor_planilhas
        import urllib.request

        wb = _multi_tab_xlsx()
        buf = io.BytesIO()
        wb.save(buf)

        class FakeResp:
            status = 200

            def __init__(self, data):
                self._data = data

            def __enter__(self):
                return self

            def __exit__(self, *a):
                return False

            def read(self):
                return self._data

        def fake_urlopen(url, timeout=0):
            return FakeResp(buf.getvalue())

        monkeypatch.setattr(urllib.request, "urlopen", fake_urlopen)

        regs = leitor_planilhas.ler_notas_google_sheets("https://docs.google.com/spreadsheets/d/ABC123/edit")
        assert len(regs) == 2
        assert regs[0].aluno == "Ana Oliveira"
        assert regs[1].aluno == "Joao Santos"

    def test_sheets_ignora_aba_de_instrucoes(self):
        from leitor_planilhas import _find_col_index

        headers = ["INSTRUCOES - PLANILHA DE NOTAS", "", ""]
        assert _find_col_index(headers, ["nome", "aluno", "estudante"]) is None

    def test_csv_com_delimitador_ponto_e_virgula(self):
        from leitor_planilhas import _ler_csv_content

        content = "Nome do Aluno;Atividade 1\nAna Oliveira;8,5\nJoao Santos;7,0\n"
        regs = _ler_csv_content(content)
        assert len(regs) == 2
        assert regs[0].nota == 8.5
        assert regs[1].nota == 7.0

    def test_csv_virgula_decimal_nao_quebra(self):
        from leitor_planilhas import _ler_csv_content

        content = "Nome do Aluno,Atividade 1\nAna Oliveira,8,5\nJoao Santos,7,0\n"
        regs = _ler_csv_content(content)
        assert len(regs) == 2
        assert regs[0].aluno == "Ana Oliveira"
        assert regs[0].nota == 8.5


def _multi_tab_xlsx():
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Instrucoes"
    ws1.append(["INSTRUCOES - PLANILHA DE NOTAS", None, None])
    ws1.append(["Coluna", "O que preencher", "Exemplo"])
    ws1.append(["Escola", "Nome da escola", "Juvenal"])
    ws1.append(["Atividade 1", "Nota da atividade", "8,5"])
    ws2 = wb.create_sheet("6o Ano")
    ws2.append(["Escola", "Turno", "Turma", "Trimestre", "Nome do Aluno", "Atividade 1"])
    ws2.append(["Juvenal", "Matutino", "6o Ano", "1o Trimestre", "Ana Oliveira", "8,5"])
    ws2.append(["Juvenal", "Matutino", "6o Ano", "1o Trimestre", "Joao Santos", "7,0"])
    return wb
