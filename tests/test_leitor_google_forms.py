import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


def _monta_forms_xlsx(path, rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form Responses 1"
    ws.append(["Timestamp", "Score", "Nome completo", "Escola", "Turma", "Q1", "Q2"])
    for row in rows:
        ws.append(row)
    wb.save(path)


class TestLeitorGoogleForms:
    def test_converte_pontuacao_bruta_em_nota(self, tmp_path):
        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), 19, "Ana Silva", "Mulde", "6º1", "a", "b"],
            [datetime(2026, 8, 19, 11, 5, 0), 8, "Bruno Souza", "Tancredo", "6º2", "a", "b"],
        ])

        registros = ler_notas_google_forms(str(path), valor_questao=0.5, atividade="Avaliacao")
        by_aluno = {r.aluno: r for r in registros}
        assert by_aluno["Ana Silva"].nota == 9.5
        assert by_aluno["Bruno Souza"].nota == 4.0
        assert by_aluno["Ana Silva"].atividade == "Avaliacao"
        assert by_aluno["Ana Silva"].escola == "Mulde"
        assert by_aluno["Ana Silva"].turma == "6º1"

    def test_formato_fracao_dois_de_vinte(self, tmp_path):
        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), "2/20", "Ana Silva", "Mulde", "6º1", "a", "b"],
        ])

        registros = ler_notas_google_forms(str(path), valor_questao=0.5, atividade="Avaliacao")
        assert len(registros) == 1
        assert registros[0].nota == 1.0

    def test_resposta_duplicada_mantem_ultima(self, tmp_path):
        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 10, 0, 0), 5, "Ana Silva", "Mulde", "6º1", "a", "b"],
            [datetime(2026, 8, 19, 11, 0, 0), 18, "Ana Silva", "Mulde", "6º1", "a", "b"],
        ])

        registros = ler_notas_google_forms(str(path), valor_questao=0.5, atividade="Avaliacao")
        assert len(registros) == 1
        assert registros[0].nota == 9.0

    def test_total_questoes_parametro_sobrepoe_deteccao(self, tmp_path):
        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), 10, "Ana Silva", "Mulde", "6º1", "a", "b"],
        ])

        registros = ler_notas_google_forms(
            str(path), valor_questao=0.5, atividade="A", n_questoes=20
        )
        assert registros[0].nota == 5.0

    def test_detectar_total_questoes_por_fracao(self, tmp_path):
        from leitor_planilhas import detectar_total_questoes_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), "3/20", "Ana", "E", "T", "a", "b"],
            [datetime(2026, 8, 19, 11, 1, 0), "19/20", "Beto", "E", "T", "a", "b"],
        ])

        assert detectar_total_questoes_forms(str(path)) == 20

    def test_linha_sem_pontuacao_e_ignorada(self, tmp_path):
        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), None, "Sem Nota", "E", "T", "a", "b"],
            [datetime(2026, 8, 19, 11, 1, 0), 20, "Ana", "E", "T", "a", "b"],
        ])

        registros = ler_notas_google_forms(str(path), valor_questao=0.5, atividade="A")
        assert [r.aluno for r in registros] == ["Ana"]

    def test_aba_sem_estrutura_forms_e_pulada(self, tmp_path):
        import openpyxl

        from leitor_planilhas import ler_notas_google_forms

        path = tmp_path / "forms.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instrucoes"
        ws.append(["Leia as instrucoes"])
        wb.save(path)

        registros = ler_notas_google_forms(str(path), valor_questao=0.5, atividade="A")
        assert registros == []

    def test_detectar_turmas_unifica_grafias(self, tmp_path):
        from leitor_planilhas import detectar_turmas_forms

        path = tmp_path / "forms.xlsx"
        _monta_forms_xlsx(path, [
            [datetime(2026, 8, 19, 11, 0, 0), 10, "Ana", "Mulde", "6º1", "a", "b"],
            [datetime(2026, 8, 19, 11, 1, 0), 10, "Beto", "Tancredo", "6°1", "a", "b"],
            [datetime(2026, 8, 19, 11, 2, 0), 10, "Cris", "Tancredo", "6º2", "a", "b"],
            [datetime(2026, 8, 19, 11, 3, 0), 10, "Duda", "Tancredo", "", "a", "b"],
        ])

        turmas = detectar_turmas_forms(str(path))
        assert turmas == ["6º1", "6º2"]

    def test_detectar_turmas_sem_coluna_retorna_vazio(self, tmp_path):
        import openpyxl

        from leitor_planilhas import detectar_turmas_forms

        path = tmp_path / "forms.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Score"])
        ws.append([datetime(2026, 8, 19, 11, 0, 0), 10])
        wb.save(path)

        assert detectar_turmas_forms(str(path)) == []
